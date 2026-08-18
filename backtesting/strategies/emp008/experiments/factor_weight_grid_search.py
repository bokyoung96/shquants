from __future__ import annotations

import argparse
import json
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from itertools import product
from pathlib import Path
from typing import Iterable, Mapping

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from backtesting.run import BacktestRunner
from backtesting.strategies.emp008.data import Emp008Config
from backtesting.strategies.emp008.factor_pipeline import (
    PreparedEmp008Factors,
    load_and_prepare_emp008_factors,
)
from backtesting.strategies.emp008.factor_registry import FactorSetId
from backtesting.strategies.emp008.reports.comparison import (
    _benchmark_returns,
    excess_summary_bps,
    performance_metrics,
)
from backtesting.strategies.emp008.run_backtest import (
    active_share_summary,
    build_target_weight_spec,
    write_active_share,
)
from backtesting.strategies.emp008.run_weights import (
    build_emp008_config,
    latest_common_end,
    write_target_weights_csv,
)
from backtesting.strategies.emp008.strategy import run_emp008


FACTOR_SET = FactorSetId.SIZE_MOMENTUM_EARNINGS_VALUE
FACTOR_NAMES = (
    "ln_market_cap",
    "momentum_12m",
    "earnings_momentum",
    "value",
)
FACTOR_LABELS = {
    "ln_market_cap": "Size",
    "momentum_12m": "12M Momentum",
    "earnings_momentum": "OP Consensus Momentum",
    "value": "FCF/TEV",
}
DEFAULT_WEIGHT_OPTIONS: Mapping[str, tuple[float, ...]] = {
    "ln_market_cap": (30.0, 40.0, 50.0),
    "momentum_12m": (30.0, 40.0, 50.0),
    "earnings_momentum": (10.0, 20.0),
    "value": (10.0, 20.0),
}
DEFAULT_START = "2019-12-30"
DEFAULT_OUTPUT_DIR = Path(
    "backtesting/strategies/emp008/tests/factor_weight_grid_search_wics"
)
DEFAULT_TRACKING_ERROR_ANNUAL = 0.007
DEFAULT_FEE = 0.0002
DEFAULT_SELL_TAX = 0.0015
DEFAULT_SLIPPAGE = 0.0005
DEFAULT_CAPITAL = 100_000_000.0
SECTOR_NEUTRAL_DATASET = "wics"


@dataclass(frozen=True, slots=True)
class WeightCandidate:
    id: str
    percentages: dict[str, float]

    @property
    def multipliers(self) -> dict[str, float]:
        return percentages_to_multipliers(self.percentages).to_dict()


@dataclass(frozen=True, slots=True)
class CandidateRun:
    candidate: WeightCandidate
    returns: pd.Series
    mean_active_share_pct: float
    avg_turnover: float
    diagnostics_success_rate_pct: float
    candidate_dir: Path


def validate_percentages(percentages: Mapping[str, float]) -> None:
    if not percentages:
        raise ValueError("factor percentages must not be empty")
    values = np.asarray(
        tuple(float(value) for value in percentages.values()), dtype=float
    )
    if not np.isfinite(values).all():
        raise ValueError("factor percentages must be finite")
    if np.any(values <= 0.0):
        raise ValueError("every factor percentage must be positive")
    if not np.isclose(float(values.sum()), 100.0, rtol=0.0, atol=1e-9):
        raise ValueError("factor percentages must sum to 100")


def percentages_to_multipliers(percentages: Mapping[str, float]) -> pd.Series:
    validate_percentages(percentages)
    count = len(percentages)
    return pd.Series(percentages, dtype=float).mul(count / 100.0)


def build_weight_grid(
    weight_options: Mapping[str, Iterable[float]],
) -> tuple[WeightCandidate, ...]:
    names = tuple(weight_options)
    if not names:
        raise ValueError("weight options must not be empty")
    candidates: list[WeightCandidate] = []
    for values in product(*(tuple(weight_options[name]) for name in names)):
        percentages = dict(zip(names, map(float, values), strict=True))
        if not np.isclose(sum(percentages.values()), 100.0, rtol=0.0, atol=1e-9):
            continue
        validate_percentages(percentages)
        candidate_id = "_".join(
            (
                f"s{percentages['ln_market_cap']:g}",
                f"m{percentages['momentum_12m']:g}",
                f"e{percentages['earnings_momentum']:g}",
                f"v{percentages['value']:g}",
            )
        )
        candidates.append(WeightCandidate(id=candidate_id, percentages=percentages))
    candidates.sort(
        key=lambda candidate: (
            -(
                candidate.percentages["ln_market_cap"]
                + candidate.percentages["momentum_12m"]
            ),
            -min(
                candidate.percentages["ln_market_cap"],
                candidate.percentages["momentum_12m"],
            ),
            -candidate.percentages["momentum_12m"],
            -candidate.percentages["ln_market_cap"],
            candidate.id,
        )
    )
    return tuple(candidates)


def build_default_candidates() -> tuple[WeightCandidate, ...]:
    baseline = WeightCandidate(
        id="equal_25", percentages=dict.fromkeys(FACTOR_NAMES, 25.0)
    )
    tilted = build_weight_grid(DEFAULT_WEIGHT_OPTIONS)
    candidates = (baseline, *tilted)
    if len({candidate.id for candidate in candidates}) != len(candidates):
        raise ValueError("duplicate candidate ids")
    return candidates


def run_factor_weight_grid_search(
    *,
    parquet_dir: Path,
    output_dir: Path,
    start: str = DEFAULT_START,
    end: str | None = None,
    tracking_error_annual: float = DEFAULT_TRACKING_ERROR_ANNUAL,
    risk_model: str = "factor_idio",
    fee: float = DEFAULT_FEE,
    sell_tax: float = DEFAULT_SELL_TAX,
    slippage: float = DEFAULT_SLIPPAGE,
    force: bool = False,
    max_workers: int = 1,
) -> dict[str, object]:
    if max_workers < 1:
        raise ValueError("max_workers must be at least 1")
    output_dir.mkdir(parents=True, exist_ok=True)
    logger = _configure_logger(output_dir / "grid_search.log")
    config = build_emp008_config(
        tracking_error_annual=tracking_error_annual,
        risk_model=risk_model,
        factor_set=FACTOR_SET.value,
        sector_neutral_dataset=SECTOR_NEUTRAL_DATASET,
    )
    if config.sector_neutral_dataset is None:
        raise ValueError("factor-weight grid search must use WICS sector neutralization")
    sector_neutral_dataset = config.sector_neutral_dataset.value
    resolved_end = end or latest_common_end(parquet_dir, config)
    candidates = build_default_candidates()
    logger.info(
        "grid_search_start candidates=%d start=%s end=%s factor_set=%s sector_neutral_dataset=%s",
        len(candidates),
        start,
        resolved_end,
        FACTOR_SET.value,
        sector_neutral_dataset,
    )

    cached_runs: dict[str, CandidateRun] = {}
    pending: list[tuple[int, WeightCandidate, Path, dict[str, object]]] = []
    for position, candidate in enumerate(candidates, start=1):
        logger.info(
            "candidate_start position=%d/%d id=%s",
            position,
            len(candidates),
            candidate.id,
        )
        candidate_dir = output_dir / "candidates" / candidate.id
        signature = _candidate_signature(
            candidate=candidate,
            start=start,
            end=resolved_end,
            tracking_error_annual=tracking_error_annual,
            risk_model=risk_model,
            fee=fee,
            sell_tax=sell_tax,
            slippage=slippage,
            sector_neutral_dataset=sector_neutral_dataset,
        )
        cached = None if force else _load_cached_candidate(candidate_dir, signature)
        if cached is not None:
            cached_runs[candidate.id] = cached
            logger.info("candidate_cache_hit id=%s", candidate.id)
            logger.info("candidate_done id=%s", candidate.id)
        else:
            pending.append((position, candidate, candidate_dir, signature))

    if pending:
        logger.info("prepared_factors_start")
        prepared = load_and_prepare_emp008_factors(
            parquet_dir, start, resolved_end, config
        )
        logger.info(
            "prepared_factors_done monthly_dates=%d pending=%d max_workers=%d",
            len(prepared.monthly_dates),
            len(pending),
            max_workers,
        )
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_jobs = {
                executor.submit(
                    _run_candidate,
                    candidate=candidate,
                    candidate_dir=candidate_dir,
                    signature=signature,
                    parquet_dir=parquet_dir,
                    start=start,
                    end=resolved_end,
                    config=config,
                    prepared=prepared,
                    fee=fee,
                    sell_tax=sell_tax,
                    slippage=slippage,
                ): (position, candidate)
                for position, candidate, candidate_dir, signature in pending
            }
            for future in as_completed(future_jobs):
                position, candidate = future_jobs[future]
                cached_runs[candidate.id] = future.result()
                logger.info(
                    "candidate_done position=%d/%d id=%s",
                    position,
                    len(candidates),
                    candidate.id,
                )

    runs = [cached_runs[candidate.id] for candidate in candidates]

    benchmark = _benchmark_returns(
        parquet_dir / "qw_BM.parquet",
        "IKS200",
        runs[0].returns.index,
    )
    outputs = _write_grid_outputs(
        output_dir=output_dir,
        runs=runs,
        benchmark_returns=benchmark,
        start=start,
        end=resolved_end,
        tracking_error_annual=tracking_error_annual,
        risk_model=risk_model,
        fee=fee,
        sell_tax=sell_tax,
        slippage=slippage,
        sector_neutral_dataset=sector_neutral_dataset,
    )
    logger.info("grid_search_done summary=%s", outputs["summary_csv"])
    return outputs


def _run_candidate(
    *,
    candidate: WeightCandidate,
    candidate_dir: Path,
    signature: Mapping[str, object],
    parquet_dir: Path,
    start: str,
    end: str,
    config: Emp008Config,
    prepared: PreparedEmp008Factors,
    fee: float,
    sell_tax: float,
    slippage: float,
) -> CandidateRun:
    candidate_dir.mkdir(parents=True, exist_ok=True)
    weights_dir = candidate_dir / "weights"
    result = run_emp008(
        parquet_dir=parquet_dir,
        start=start,
        end=end,
        config=config,
        output_dir=weights_dir,
        prepared=prepared,
        factor_weights=candidate.multipliers,
    )
    weights_csv = write_target_weights_csv(
        result.target_weights, weights_dir / "target_weights.csv"
    )
    active_share_paths = write_active_share(weights_dir / "active_weights.parquet")
    weight_dates = tuple(
        pd.to_datetime(result.target_weights.index).strftime("%Y-%m-%d")
    )
    spec = build_target_weight_spec(
        name=f"emp008_weight_grid_{candidate.id}",
        weights_csv=weights_csv,
        dates=weight_dates,
        end=end,
        fill_mode="close",
        capital=DEFAULT_CAPITAL,
        fee=fee,
        sell_tax=sell_tax,
        slippage=slippage,
        allow_fractional=True,
    )
    runner = BacktestRunner(
        result_dir=candidate_dir / "backtests", write_report_assets=False, profile=True
    )
    report = runner.run_spec(runner.resolve_spec(spec))
    backtest_dir = Path(str(report.output_dir))
    returns_csv = backtest_dir / "series" / "returns.csv"
    active_share = active_share_summary(
        Path(active_share_paths["active_share_parquet"])
    )
    diagnostics_success_rate = float(result.diagnostics["success"].mean() * 100.0)
    metadata = {
        "signature": signature,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "weights_dir": str(weights_dir.resolve()),
        "weights_csv": str(weights_csv.resolve()),
        "returns_csv": str(returns_csv.resolve()),
        "backtest_dir": str(backtest_dir.resolve()),
        "backtest_summary": report.summary,
        "mean_active_share_pct": float(active_share["mean_pct"]),
        "diagnostics_success_rate_pct": diagnostics_success_rate,
    }
    (candidate_dir / "run_metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return _candidate_run_from_metadata(candidate, candidate_dir, metadata)


def _candidate_signature(
    *,
    candidate: WeightCandidate,
    start: str,
    end: str,
    tracking_error_annual: float,
    risk_model: str,
    fee: float,
    sell_tax: float,
    slippage: float,
    sector_neutral_dataset: str,
) -> dict[str, object]:
    return {
        "factor_set": FACTOR_SET.value,
        "sector_neutral_dataset": sector_neutral_dataset,
        "factor_percentages": candidate.percentages,
        "factor_multipliers": candidate.multipliers,
        "start": start,
        "end": end,
        "tracking_error_annual": tracking_error_annual,
        "risk_model": risk_model,
        "fill_mode": "close",
        "costs": {"fee": fee, "sell_tax": sell_tax, "slippage": slippage},
        "capital": DEFAULT_CAPITAL,
        "allow_fractional": True,
    }


def _load_cached_candidate(
    candidate_dir: Path,
    expected_signature: Mapping[str, object],
) -> CandidateRun | None:
    metadata_path = candidate_dir / "run_metadata.json"
    if not metadata_path.exists():
        return None
    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if metadata.get("signature") != expected_signature:
        return None
    required_paths = (
        Path(str(metadata.get("returns_csv", ""))),
        Path(str(metadata.get("weights_csv", ""))),
        Path(str(metadata.get("weights_dir", ""))) / "active_weights.parquet",
        Path(str(metadata.get("weights_dir", ""))) / "diagnostics.parquet",
    )
    if not all(path.is_file() for path in required_paths):
        return None
    percentages = metadata["signature"]["factor_percentages"]
    candidate = WeightCandidate(
        id=candidate_dir.name,
        percentages={str(name): float(value) for name, value in percentages.items()},
    )
    return _candidate_run_from_metadata(candidate, candidate_dir, metadata)


def _candidate_run_from_metadata(
    candidate: WeightCandidate,
    candidate_dir: Path,
    metadata: Mapping[str, object],
) -> CandidateRun:
    returns = pd.read_csv(
        Path(str(metadata["returns_csv"])), parse_dates=["date"]
    ).set_index("date")["returns"]
    returns = returns.astype(float).sort_index().rename(candidate.id)
    backtest_summary = metadata["backtest_summary"]
    if not isinstance(backtest_summary, Mapping):
        raise ValueError(f"invalid backtest summary for {candidate.id}")
    return CandidateRun(
        candidate=candidate,
        returns=returns,
        mean_active_share_pct=_json_number(metadata["mean_active_share_pct"]),
        avg_turnover=_json_number(backtest_summary["avg_turnover"]),
        diagnostics_success_rate_pct=_json_number(
            metadata["diagnostics_success_rate_pct"]
        ),
        candidate_dir=candidate_dir,
    )


def _json_number(value: object) -> float:
    if not isinstance(value, (int, float, str)):
        raise ValueError(f"expected a numeric JSON value, got {type(value).__name__}")
    return float(value)


def _write_grid_outputs(
    *,
    output_dir: Path,
    runs: list[CandidateRun],
    benchmark_returns: pd.Series,
    start: str,
    end: str,
    tracking_error_annual: float,
    risk_model: str,
    fee: float,
    sell_tax: float,
    slippage: float,
    sector_neutral_dataset: str,
) -> dict[str, object]:
    daily = _align_daily_returns(runs, benchmark_returns)
    summary = _build_summary(runs, daily)
    annual_returns, annual_excess = _build_annual_tables(daily)
    top_ids = tuple(summary.head(5)["candidate_id"].astype(str))
    robustness = _build_robustness(daily, top_ids)
    assumptions = pd.DataFrame(
        [
            ("factor_set", FACTOR_SET.value),
            ("sector_neutral_dataset", sector_neutral_dataset),
            ("start", start),
            ("end", end),
            ("candidate_count", len(runs)),
            ("weight_rule", "Every factor > 0%; total = 100%"),
            ("grid", "10 percentage-point tilted grid plus 25/25/25/25 baseline"),
            (
                "ranking",
                "Annualized excess return descending, then information ratio descending",
            ),
            ("tracking_error_annual", tracking_error_annual),
            ("risk_model", risk_model),
            ("fee", fee),
            ("sell_tax", sell_tax),
            ("slippage", slippage),
            ("capital", DEFAULT_CAPITAL),
            ("fill_mode", "close"),
            ("benchmark", "IKS200"),
        ],
        columns=["item", "value"],
    )
    checks = _build_checks(summary, daily, runs)

    paths = {
        "summary_csv": output_dir / "performance_summary.csv",
        "annual_returns_csv": output_dir / "annual_returns_pct.csv",
        "annual_excess_csv": output_dir / "annual_relative_excess_pct.csv",
        "robustness_csv": output_dir / "top5_subperiod_robustness.csv",
        "daily_returns_csv": output_dir / "daily_returns.csv",
        "workbook_xlsx": output_dir / "factor_weight_grid_search.xlsx",
        "manifest_json": output_dir / "manifest.json",
    }
    summary.to_csv(paths["summary_csv"], index=False, encoding="utf-8-sig")
    annual_returns.to_csv(paths["annual_returns_csv"], encoding="utf-8-sig")
    annual_excess.to_csv(paths["annual_excess_csv"], encoding="utf-8-sig")
    robustness.to_csv(paths["robustness_csv"], index=False, encoding="utf-8-sig")
    daily.to_csv(paths["daily_returns_csv"], encoding="utf-8-sig")
    _write_excel_workbook(
        path=paths["workbook_xlsx"],
        summary=summary,
        annual_returns=annual_returns,
        annual_excess=annual_excess,
        robustness=robustness,
        daily=daily,
        assumptions=assumptions,
        checks=checks,
    )
    manifest = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "factor_set": FACTOR_SET.value,
        "sector_neutral_dataset": sector_neutral_dataset,
        "factor_labels": FACTOR_LABELS,
        "start": start,
        "end": end,
        "candidate_count": len(runs),
        "top_candidate": str(summary.iloc[0]["candidate_id"]),
        "top_five": list(top_ids),
        "paths": {
            name: str(path.resolve())
            for name, path in paths.items()
            if name != "manifest_json"
        },
    }
    paths["manifest_json"].write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return {
        **{name: str(path) for name, path in paths.items()},
        "candidate_count": len(runs),
        "top_candidate": manifest["top_candidate"],
        "top_five": list(top_ids),
        "start": start,
        "end": end,
        "sector_neutral_dataset": sector_neutral_dataset,
    }


def _align_daily_returns(
    runs: list[CandidateRun], benchmark_returns: pd.Series
) -> pd.DataFrame:
    if not runs:
        raise ValueError("candidate runs must not be empty")
    common_index = pd.DatetimeIndex(pd.to_datetime(benchmark_returns.dropna().index))
    for run in runs:
        common_index = common_index.intersection(
            pd.DatetimeIndex(pd.to_datetime(run.returns.dropna().index))
        )
    common_index = common_index.sort_values()
    if common_index.empty:
        raise ValueError("no common return dates")
    daily = pd.DataFrame(index=common_index)
    daily.index.name = "date"
    daily["IKS200"] = benchmark_returns.reindex(common_index).astype(float)
    for run in runs:
        daily[run.candidate.id] = run.returns.reindex(common_index).astype(float)
    if daily.isna().any().any():
        raise ValueError("missing values after daily return alignment")
    return daily


def _build_summary(runs: list[CandidateRun], daily: pd.DataFrame) -> pd.DataFrame:
    benchmark = daily["IKS200"]
    benchmark_wealth = (1.0 + benchmark).cumprod()
    rows: list[dict[str, object]] = []
    for run in runs:
        candidate_id = run.candidate.id
        returns = daily[candidate_id]
        excess = returns.sub(benchmark)
        excess_metrics = excess_summary_bps(
            pd.DataFrame({candidate_id: excess}),
            periods_per_year=252,
        ).loc[candidate_id]
        wealth = (1.0 + returns).cumprod()
        row: dict[str, object] = {
            "candidate_id": candidate_id,
            **{
                f"{name}_pct": value
                for name, value in run.candidate.percentages.items()
            },
            **{
                f"{name}_multiplier": value
                for name, value in run.candidate.multipliers.items()
            },
            **performance_metrics(returns, periods_per_year=252),
            **{name: float(value) for name, value in excess_metrics.items()},
            "cumulative_relative_excess_bp": float(
                (wealth.iloc[-1] / benchmark_wealth.iloc[-1] - 1.0) * 10_000.0
            ),
            "mean_active_share_pct": run.mean_active_share_pct,
            "avg_turnover_pct": run.avg_turnover * 100.0,
            "diagnostics_success_rate_pct": run.diagnostics_success_rate_pct,
            "candidate_dir": str(run.candidate_dir),
        }
        rows.append(row)
    summary = pd.DataFrame(rows).sort_values(
        ["annualized_excess_bp", "information_ratio"],
        ascending=[False, False],
        kind="mergesort",
    )
    summary.insert(0, "rank", np.arange(1, len(summary) + 1))
    return summary.reset_index(drop=True)


def _build_annual_tables(daily: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    annual_input = daily
    if (
        len(daily) > 1
        and daily.iloc[0].eq(0.0).all()
        and daily.index[0].year < daily.index[1].year
    ):
        annual_input = daily.iloc[1:]
    annual_returns = (
        (1.0 + annual_input).groupby(annual_input.index.year).prod().sub(1.0)
    )
    annual_returns = annual_returns.mul(100.0)
    annual_returns.index.name = "year"
    annual_excess = (
        annual_returns.drop(columns="IKS200")
        .add(100.0)
        .div(annual_returns["IKS200"].add(100.0), axis=0)
        .sub(1.0)
        .mul(100.0)
    )
    annual_excess.index.name = "year"
    return annual_returns, annual_excess


def _build_robustness(daily: pd.DataFrame, top_ids: tuple[str, ...]) -> pd.DataFrame:
    periods = (
        ("2020-2022", "2020-01-01", "2022-12-31"),
        ("2023-2024", "2023-01-01", "2024-12-31"),
        ("2025-latest", "2025-01-01", str(daily.index.max().date())),
    )
    rows: list[dict[str, object]] = []
    for label, start, end in periods:
        start_date = pd.Timestamp(start)
        end_date = pd.Timestamp(end)
        period = daily.loc[(daily.index >= start_date) & (daily.index <= end_date)]
        if period.empty:
            continue
        for candidate_id in top_ids:
            returns = period[candidate_id]
            benchmark = period["IKS200"]
            excess = returns.sub(benchmark)
            excess_metrics = excess_summary_bps(
                pd.DataFrame({candidate_id: excess}),
                periods_per_year=252,
            ).loc[candidate_id]
            rows.append(
                {
                    "period": label,
                    "candidate_id": candidate_id,
                    "observations": len(period),
                    **performance_metrics(returns, periods_per_year=252),
                    "annualized_excess_bp": float(
                        excess_metrics["annualized_excess_bp"]
                    ),
                    "information_ratio": float(excess_metrics["information_ratio"]),
                    "cumulative_relative_excess_bp": float(
                        (((1.0 + returns).prod() / (1.0 + benchmark).prod()) - 1.0)
                        * 10_000.0
                    ),
                }
            )
    return pd.DataFrame(rows)


def _build_checks(
    summary: pd.DataFrame,
    daily: pd.DataFrame,
    runs: list[CandidateRun],
) -> pd.DataFrame:
    weight_columns = [f"{name}_pct" for name in FACTOR_NAMES]
    weight_sums = summary[weight_columns].sum(axis=1)
    checks = (
        ("candidate_count", len(runs) == 9, len(runs), 9),
        (
            "all_weights_positive",
            bool(summary[weight_columns].gt(0.0).all().all()),
            float(summary[weight_columns].min().min()),
            "> 0",
        ),
        (
            "all_weights_sum_100",
            bool(np.isclose(weight_sums, 100.0, rtol=0.0, atol=1e-9).all()),
            float((weight_sums - 100.0).abs().max()),
            0.0,
        ),
        (
            "daily_returns_complete",
            not daily.isna().any().any(),
            int(daily.isna().sum().sum()),
            0,
        ),
        (
            "optimizer_success",
            bool(summary["diagnostics_success_rate_pct"].eq(100.0).all()),
            float(summary["diagnostics_success_rate_pct"].min()),
            100.0,
        ),
    )
    return pd.DataFrame(
        [
            {
                "check": name,
                "status": "OK" if passed else "FAIL",
                "actual": actual,
                "expected": expected,
            }
            for name, passed, actual, expected in checks
        ]
    )


def _write_excel_workbook(
    *,
    path: Path,
    summary: pd.DataFrame,
    annual_returns: pd.DataFrame,
    annual_excess: pd.DataFrame,
    robustness: pd.DataFrame,
    daily: pd.DataFrame,
    assumptions: pd.DataFrame,
    checks: pd.DataFrame,
) -> None:
    compact_summary_columns = [
        "rank",
        "candidate_id",
        *(f"{name}_pct" for name in FACTOR_NAMES),
        "cagr_pct",
        "total_return_pct",
        "annualized_excess_bp",
        "information_ratio",
        "max_drawdown_pct",
        "avg_turnover_pct",
    ]
    compact_summary = summary.loc[:, compact_summary_columns].rename(
        columns={
            "rank": "Rank",
            "candidate_id": "Candidate",
            "ln_market_cap_pct": "Size %",
            "momentum_12m_pct": "12M Momentum %",
            "earnings_momentum_pct": "OP Consensus %",
            "value_pct": "FCF/TEV %",
            "cagr_pct": "CAGR %",
            "total_return_pct": "Total Return %",
            "annualized_excess_bp": "Annualized Excess (bp)",
            "information_ratio": "IR",
            "max_drawdown_pct": "MDD %",
            "avg_turnover_pct": "Avg Turnover %",
        }
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        compact_summary.to_excel(writer, sheet_name="Summary", index=False)
        summary.to_excel(writer, sheet_name="Candidate Detail", index=False)
        annual_returns.to_excel(writer, sheet_name="Annual Returns")
        annual_excess.to_excel(writer, sheet_name="Annual Excess")
        robustness.to_excel(writer, sheet_name="Top5 Robustness", index=False)
        daily.to_excel(writer, sheet_name="Daily Returns")
        assumptions.to_excel(writer, sheet_name="Assumptions", index=False)
        checks.to_excel(writer, sheet_name="Checks", index=False)

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    input_font = Font(color="0000FF")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    thin_gray = Side(style="thin", color="D9E1F2")
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin_gray)
        for column_cells in sheet.columns:
            values = [
                "" if cell.value is None else str(cell.value)
                for cell in column_cells[:200]
            ]
            width = min(max(max(map(len, values), default=0) + 2, 10), 30)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
        sheet.row_dimensions[1].height = 24

    for sheet_name in ("Summary", "Candidate Detail"):
        summary_sheet = workbook[sheet_name]
        summary_headers = {cell.value: cell.column for cell in summary_sheet[1]}
        for header, column in summary_headers.items():
            if isinstance(header, str) and (
                header.endswith("_pct")
                or header.endswith("_bp")
                or header.endswith("%")
                or header.endswith("(bp)")
            ):
                for cell in summary_sheet.iter_cols(
                    min_col=column, max_col=column, min_row=2
                ):
                    for item in cell:
                        item.number_format = "0.00"
            if isinstance(header, str) and header.endswith("_multiplier"):
                for cell in summary_sheet.iter_cols(
                    min_col=column, max_col=column, min_row=2
                ):
                    for item in cell:
                        item.number_format = "0.00x"
            if header in {"IR", "information_ratio"}:
                for cell in summary_sheet.iter_cols(
                    min_col=column, max_col=column, min_row=2
                ):
                    for item in cell:
                        item.number_format = "0.000"
            if isinstance(header, str) and (
                (header.endswith("_pct") and header.startswith(tuple(FACTOR_NAMES)))
                or header in {"Size %", "12M Momentum %", "OP Consensus %", "FCF/TEV %"}
            ):
                for cell in summary_sheet.iter_cols(
                    min_col=column, max_col=column, min_row=2
                ):
                    for item in cell:
                        item.font = input_font

    for sheet_name in ("Annual Returns", "Annual Excess"):
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, min_col=2):
            for cell in row:
                cell.number_format = "0.00"
    daily_sheet = workbook["Daily Returns"]
    for cell in daily_sheet["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for row in daily_sheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = "0.0000%"
    checks_sheet = workbook["Checks"]
    for row in checks_sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        cell = row[0]
        cell.fill = green_fill if cell.value == "OK" else red_fill
        cell.font = Font(bold=True)
    workbook.save(path)

    verified = load_workbook(path, read_only=False, data_only=False)
    expected_sheets = {
        "Summary",
        "Candidate Detail",
        "Annual Returns",
        "Annual Excess",
        "Top5 Robustness",
        "Daily Returns",
        "Assumptions",
        "Checks",
    }
    if set(verified.sheetnames) != expected_sheets:
        raise ValueError("unexpected workbook sheets")
    if any(
        cell.value == "FAIL" for row in verified["Checks"].iter_rows() for cell in row
    ):
        raise ValueError("workbook checks failed")
    verified.close()


def _configure_logger(path: Path) -> logging.Logger:
    logger = logging.getLogger("emp008_factor_weight_grid_search")
    logger.handlers.clear()
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    file_handler = logging.FileHandler(path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    return logger


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Run the EMP008 four-factor positive-weight grid search."
    )
    parser.add_argument("--parquet-dir", type=Path, default=Path("parquet"))
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--start", default=DEFAULT_START)
    parser.add_argument("--end")
    parser.add_argument(
        "--tracking-error-annual", type=float, default=DEFAULT_TRACKING_ERROR_ANNUAL
    )
    parser.add_argument(
        "--risk-model",
        choices=("factor_idio", "direct_covariance"),
        default="factor_idio",
    )
    parser.add_argument("--fee", type=float, default=DEFAULT_FEE)
    parser.add_argument("--sell-tax", type=float, default=DEFAULT_SELL_TAX)
    parser.add_argument("--slippage", type=float, default=DEFAULT_SLIPPAGE)
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--max-workers", type=int, default=1)
    return parser


def main(argv: list[str] | None = None) -> None:
    args = _parser().parse_args(argv)
    payload = run_factor_weight_grid_search(
        parquet_dir=args.parquet_dir,
        output_dir=args.output_dir,
        start=args.start,
        end=args.end,
        tracking_error_annual=args.tracking_error_annual,
        risk_model=args.risk_model,
        fee=args.fee,
        sell_tax=args.sell_tax,
        slippage=args.slippage,
        force=args.force,
        max_workers=args.max_workers,
    )
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
