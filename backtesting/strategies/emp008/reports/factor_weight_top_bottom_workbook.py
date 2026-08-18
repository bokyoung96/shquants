from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Mapping

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from backtesting.reporting.benchmarks import _load_display_name_maps


DETAIL_COLUMNS = (
    "리밸런싱일",
    "후보 순위",
    "후보 ID",
    "Size",
    "12개월 모멘텀",
    "영업이익 컨센서스",
    "FCF/TEV",
    "종목 순위",
    "종목코드",
    "종목명",
    "투자비중",
    "BM 비중",
    "Active 비중",
)


def build_top_bottom_workbook(
    *,
    run_dir: Path,
    output_path: Path,
    stock_name_map_path: Path,
) -> Path:
    conditions = json.loads((run_dir / "run_conditions.json").read_text(encoding="utf-8"))
    summary = pd.read_csv(run_dir / "performance_summary.csv").sort_values("rank")
    _validate_summary(summary, expected_count=int(conditions["candidate_count"]))
    _sector_names, stock_names = _load_display_name_maps(stock_name_map_path)

    top_rows: list[tuple[object, ...]] = []
    bottom_rows: list[tuple[object, ...]] = []
    expected_dates: pd.DatetimeIndex | None = None
    missing_names: set[str] = set()

    for candidate in summary.itertuples(index=False):
        weights_dir = run_dir / "candidates" / str(candidate.candidate_id) / "weights"
        target = _read_weights(weights_dir / "target_weights.parquet")
        active = _read_weights(weights_dir / "active_weights.parquet")
        _validate_weight_frames(target=target, active=active, candidate_id=str(candidate.candidate_id))
        if expected_dates is None:
            expected_dates = pd.DatetimeIndex(target.index)
        elif not pd.DatetimeIndex(target.index).equals(expected_dates):
            raise ValueError(f"rebalance dates differ for {candidate.candidate_id}")
        benchmark = target.sub(active).mask(lambda frame: frame.abs().lt(1e-14), 0.0)
        factor_weights = (
            float(candidate.ln_market_cap_pct) / 100.0,
            float(candidate.momentum_12m_pct) / 100.0,
            float(candidate.earnings_momentum_pct) / 100.0,
            float(candidate.value_pct) / 100.0,
        )
        for rebalance_date in target.index:
            date_target = target.loc[rebalance_date]
            date_benchmark = benchmark.loc[rebalance_date]
            date_active = active.loc[rebalance_date]
            top_tickers = _rank_tickers(date_active, ascending=False)
            bottom_tickers = _rank_tickers(date_active, ascending=True)
            top_rows.extend(
                _detail_rows(
                    rebalance_date=rebalance_date,
                    candidate_rank=int(candidate.rank),
                    candidate_id=str(candidate.candidate_id),
                    factor_weights=factor_weights,
                    tickers=top_tickers,
                    target=date_target,
                    benchmark=date_benchmark,
                    active=date_active,
                    stock_names=stock_names,
                    missing_names=missing_names,
                )
            )
            bottom_rows.extend(
                _detail_rows(
                    rebalance_date=rebalance_date,
                    candidate_rank=int(candidate.rank),
                    candidate_id=str(candidate.candidate_id),
                    factor_weights=factor_weights,
                    tickers=bottom_tickers,
                    target=date_target,
                    benchmark=date_benchmark,
                    active=date_active,
                    stock_names=stock_names,
                    missing_names=missing_names,
                )
            )

    if expected_dates is None:
        raise ValueError("no candidate weights found")
    expected_detail_rows = len(summary) * len(expected_dates) * 10
    if len(top_rows) != expected_detail_rows or len(bottom_rows) != expected_detail_rows:
        raise ValueError("unexpected top/bottom row count")
    top_rows.sort(key=lambda row: (row[0], row[1], row[7]))
    bottom_rows.sort(key=lambda row: (row[0], row[1], row[7]))

    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_guide_sheet(
        workbook,
        run_dir=run_dir,
        conditions=conditions,
        rebalance_dates=expected_dates,
        candidate_count=len(summary),
        detail_rows=expected_detail_rows,
        missing_names=missing_names,
    )
    _write_candidates_sheet(workbook, summary)
    _write_detail_sheet(workbook, title="Top 10", rows=top_rows, table_name="Top10Table", positive=True)
    _write_detail_sheet(
        workbook,
        title="Bottom 10",
        rows=bottom_rows,
        table_name="Bottom10Table",
        positive=False,
    )
    workbook.properties.title = (
        f"EMP008 {conditions['sector_neutral_dataset']} 리밸런싱별 Top/Bottom 10"
    )
    workbook.properties.subject = "팩터 가중치 후보별 투자비중, BM 비중, Active 비중"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    _verify_workbook(
        output_path,
        expected_detail_rows=expected_detail_rows,
        expected_start=str(conditions["start"]),
        expected_end=str(conditions["end"]),
    )
    return output_path


def _read_weights(path: Path) -> pd.DataFrame:
    if not path.is_file():
        raise FileNotFoundError(path)
    frame = pd.read_parquet(path).astype(float)
    frame.index = pd.to_datetime(frame.index)
    return frame.sort_index().sort_index(axis=1)


def _validate_summary(summary: pd.DataFrame, *, expected_count: int) -> None:
    required = {
        "rank",
        "candidate_id",
        "ln_market_cap_pct",
        "momentum_12m_pct",
        "earnings_momentum_pct",
        "value_pct",
    }
    missing = sorted(required.difference(summary.columns))
    if missing:
        raise ValueError(f"performance summary missing columns: {missing}")
    if len(summary) != expected_count:
        raise ValueError(f"expected {expected_count} candidates, got {len(summary)}")


def _validate_weight_frames(
    *, target: pd.DataFrame, active: pd.DataFrame, candidate_id: str
) -> None:
    if target.empty or active.empty:
        raise ValueError(f"empty weights for {candidate_id}")
    if not target.index.equals(active.index) or not target.columns.equals(active.columns):
        raise ValueError(f"target/active labels differ for {candidate_id}")
    benchmark = target.sub(active)
    if not np.isfinite(target.to_numpy()).all() or not np.isfinite(active.to_numpy()).all():
        raise ValueError(f"non-finite weights for {candidate_id}")
    if not np.allclose(target.sum(axis=1), 1.0, rtol=0.0, atol=1e-9):
        raise ValueError(f"target weights do not sum to 100% for {candidate_id}")
    if not np.allclose(active.sum(axis=1), 0.0, rtol=0.0, atol=1e-9):
        raise ValueError(f"active weights do not sum to 0% for {candidate_id}")
    if not np.allclose(benchmark.sum(axis=1), 1.0, rtol=0.0, atol=1e-9):
        raise ValueError(f"benchmark weights do not sum to 100% for {candidate_id}")
    if float(benchmark.min().min()) < -1e-9:
        raise ValueError(f"negative benchmark weights for {candidate_id}")


def _rank_tickers(active: pd.Series, *, ascending: bool) -> tuple[str, ...]:
    ranked = (
        active.rename("active_weight")
        .rename_axis("ticker")
        .reset_index()
        .sort_values(
            ["active_weight", "ticker"],
            ascending=[ascending, True],
            kind="mergesort",
        )
        .head(10)
    )
    return tuple(ranked["ticker"].astype(str))


def _detail_rows(
    *,
    rebalance_date: pd.Timestamp,
    candidate_rank: int,
    candidate_id: str,
    factor_weights: tuple[float, float, float, float],
    tickers: tuple[str, ...],
    target: pd.Series,
    benchmark: pd.Series,
    active: pd.Series,
    stock_names: Mapping[str, str],
    missing_names: set[str],
) -> list[tuple[object, ...]]:
    rows: list[tuple[object, ...]] = []
    for stock_rank, ticker in enumerate(tickers, start=1):
        name = stock_names.get(ticker)
        if not name:
            missing_names.add(ticker)
            name = ticker
        rows.append(
            (
                pd.Timestamp(rebalance_date).to_pydatetime(),
                candidate_rank,
                candidate_id,
                *factor_weights,
                stock_rank,
                ticker,
                name,
                float(target[ticker]),
                float(benchmark[ticker]),
                float(active[ticker]),
            )
        )
    return rows


def _write_guide_sheet(
    workbook: Workbook,
    *,
    run_dir: Path,
    conditions: Mapping[str, object],
    rebalance_dates: pd.DatetimeIndex,
    candidate_count: int,
    detail_rows: int,
    missing_names: set[str],
) -> None:
    sheet = workbook.create_sheet("안내")
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:D1")
    dataset_label = (
        "WICS"
        if conditions["sector_neutral_dataset"] == "qw_wics_sec_big"
        else "WI26"
    )
    sheet["A1"] = f"EMP008 {dataset_label} 리밸런싱별 Top 10 / Bottom 10"
    sheet["A1"].font = Font(name="Malgun Gothic", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="17365D")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 30
    rows = [
        ("구분", "값", "설명", "원본"),
        ("섹터 중립화", conditions["sector_neutral_dataset"], "WI26 또는 WICS", run_dir.name),
        ("백테스트 기간", f"{conditions['start']} ~ {conditions['end']}", "성과는 2020년부터 누적", "run_conditions.json"),
        ("리밸런싱 횟수", len(rebalance_dates), "월말 기준 저장 시점", "target_weights.parquet"),
        ("가중치 후보", candidate_count, "후보별 순위와 팩터 비중은 후보 목록 시트 참조", "performance_summary.csv"),
        ("Top 10 정의", "Active 비중 내림차순 10개", "Active 비중 = 투자비중 - BM 비중", "active_weights.parquet"),
        ("Bottom 10 정의", "Active 비중 오름차순 10개", "가장 큰 BM 대비 과소비중 10개", "active_weights.parquet"),
        ("상세 행 수", detail_rows, "Top 10과 Bottom 10 각각의 행 수", ""),
        ("종목명 누락", len(missing_names), "누락 시 종목코드를 종목명으로 표시", "raw/map.xlsx"),
        ("Tracking Error", float(conditions["tracking_error_annual"]), "연율", "run_conditions.json"),
        ("거래비용", f"수수료 {float(conditions['fee']):.4%} / 매도세 {float(conditions['sell_tax']):.4%} / 슬리피지 {float(conditions['slippage']):.4%}", "성과 산출 조건", "run_conditions.json"),
        ("벤치마크", conditions["benchmark"], "BM 비중은 각 후보의 투자비중-Active 비중으로 복원", "target/active weights"),
    ]
    for row in rows:
        sheet.append(row)
    _style_header(sheet, row=2, max_column=4)
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
        row[0].font = Font(name="Malgun Gothic", bold=True, color="17365D")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row[0].row].height = 30
    sheet["B11"].number_format = "0.00%"
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 42
    sheet.column_dimensions["C"].width = 46
    sheet.column_dimensions["D"].width = 48
    sheet.freeze_panes = "A3"


def _write_candidates_sheet(workbook: Workbook, summary: pd.DataFrame) -> None:
    sheet = workbook.create_sheet("후보 목록")
    sheet.sheet_view.showGridLines = False
    columns = (
        ("순위", "rank"),
        ("후보 ID", "candidate_id"),
        ("Size", "ln_market_cap_pct"),
        ("12개월 모멘텀", "momentum_12m_pct"),
        ("영업이익 컨센서스", "earnings_momentum_pct"),
        ("FCF/TEV", "value_pct"),
        ("CAGR", "cagr_pct"),
        ("누적수익률", "total_return_pct"),
        ("IR", "information_ratio"),
        ("MDD", "max_drawdown_pct"),
    )
    sheet.append([label for label, _column in columns])
    for candidate in summary.itertuples(index=False):
        values = []
        for _label, column in columns:
            value = getattr(candidate, column)
            if column.endswith("_pct"):
                value = float(value) / 100.0
            values.append(value)
        sheet.append(values)
    _style_header(sheet, row=1, max_column=len(columns))
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=8):
        for cell in row:
            cell.number_format = "0.00%"
    for cell in sheet["I"][1:]:
        cell.number_format = "0.000"
    for cell in sheet["J"][1:]:
        cell.number_format = "0.00%"
    widths = (10, 24, 12, 18, 22, 12, 12, 14, 10, 12)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _write_detail_sheet(
    workbook: Workbook,
    *,
    title: str,
    rows: list[tuple[object, ...]],
    table_name: str,
    positive: bool,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    sheet.append(DETAIL_COLUMNS)
    for row in rows:
        sheet.append(row)
    _style_header(sheet, row=1, max_column=len(DETAIL_COLUMNS))
    for cell in sheet["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
        cell.alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=7):
        for cell in row:
            cell.number_format = "0%"
    for row in sheet.iter_rows(min_row=2, min_col=11, max_col=13):
        for cell in row:
            cell.number_format = "0.0000%"
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        row[0].alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=2, min_col=8, max_col=9):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    sheet["K1"].comment = Comment("EMP008 최적화 후 실제 투자 목표비중", "User")
    sheet["L1"].comment = Comment("같은 리밸런싱 시점에 최적화가 사용한 KOSPI200 BM 비중", "User")
    sheet["M1"].comment = Comment("투자비중 - BM 비중. Top/Bottom 순위 기준", "User")
    table_ref = f"A1:M{sheet.max_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2" if positive else "TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    active_range = f"M2:M{sheet.max_row}"
    sheet.conditional_formatting.add(
        active_range,
        ColorScaleRule(
            start_type="min",
            start_color="FFF2CC" if positive else "F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB9C",
            end_type="max",
            end_color="63BE7B" if positive else "FFF2CC",
        ),
    )
    widths = (14, 11, 24, 10, 16, 20, 11, 11, 13, 22, 14, 14, 14)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = table_ref


def _style_header(sheet, *, row: int, max_column: int) -> None:
    fill = PatternFill("solid", fgColor="17365D")
    font = Font(name="Malgun Gothic", color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E1F2"))
    for cells in sheet.iter_rows(min_row=row, max_row=row, min_col=1, max_col=max_column):
        for cell in cells:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    sheet.row_dimensions[row].height = 28


def _verify_workbook(
    path: Path,
    *,
    expected_detail_rows: int,
    expected_start: str,
    expected_end: str,
) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False)
    if workbook.sheetnames != ["안내", "후보 목록", "Top 10", "Bottom 10"]:
        raise ValueError(f"unexpected workbook sheets: {workbook.sheetnames}")
    errors = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
    for title in ("Top 10", "Bottom 10"):
        sheet = workbook[title]
        if sheet.max_row != expected_detail_rows + 1:
            raise ValueError(f"unexpected row count in {title}: {sheet.max_row}")
        if sheet.max_column != len(DETAIL_COLUMNS):
            raise ValueError(f"unexpected column count in {title}: {sheet.max_column}")
        dates = [sheet["A2"].value, sheet[f"A{sheet.max_row}"].value]
        if min(dates).strftime("%Y-%m-%d") != expected_start:
            raise ValueError(f"unexpected start date in {title}")
        if max(dates).strftime("%Y-%m-%d") != expected_end:
            raise ValueError(f"unexpected end date in {title}")
        if any(cell.value in errors for row in sheet.iter_rows() for cell in row):
            raise ValueError(f"formula error in {title}")
    workbook.close()


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build EMP008 rebalance-date top/bottom active-weight workbooks."
    )
    parser.add_argument("--run-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--stock-name-map", type=Path, default=Path("raw/map.xlsx"))
    return parser


def main(argv: list[str] | None = None) -> None:
    args = _parser().parse_args(argv)
    output = build_top_bottom_workbook(
        run_dir=args.run_dir,
        output_path=args.output,
        stock_name_map_path=args.stock_name_map,
    )
    print(output)


if __name__ == "__main__":
    main()
