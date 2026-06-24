from __future__ import annotations

from dataclasses import dataclass
from datetime import time
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


SOURCE_DIR = Path("etc/data/sidecar")
OUTPUT_DIR = Path("etc/momentum")
PLOT_PATH = OUTPUT_DIR / "momentum_strategy_graph.png"
EXCEL_PATH = OUTPUT_DIR / "momentum_strategy_results.xlsx"
REPORT_PATH = OUTPUT_DIR / "momentum_strategy_report.md"

THRESHOLDS = (1, 2, 3, 4, 5)
TARGET_PERIODS = (
    ("2025-01~current", pd.Timestamp("2025-01-01"), pd.Timestamp.max.normalize()),
    ("2026-01~current", pd.Timestamp("2026-01-01"), pd.Timestamp.max.normalize()),
    ("2026-05-27~current", pd.Timestamp("2026-05-27"), pd.Timestamp.max.normalize()),
)
CHECKPOINTS = (
    "open",
    "09:30",
    "10:00",
    "10:30",
    "11:00",
    "11:30",
    "12:00",
    "13:00",
    "13:30",
    "14:00",
    "14:30",
    "15:00",
    "15:20",
    "15:30",
)
EXIT_TIMES = tuple(label for label in CHECKPOINTS if label != "open")
FULL_SESSION_END = time(15, 30)


@dataclass(frozen=True)
class Paths:
    source: Path
    plot: Path = PLOT_PATH
    excel: Path = EXCEL_PATH
    report: Path = REPORT_PATH


def find_source_file(source_dir: Path = SOURCE_DIR) -> Path:
    matches = sorted(path for path in source_dir.glob("*KOSPI200*INDEX*.xlsx") if not path.name.startswith("~$"))
    if not matches:
        raise FileNotFoundError(f"No KOSPI200 1-minute Excel file found under {source_dir}")
    return matches[0]


def load_prices(path: Path) -> pd.DataFrame:
    raw = pd.read_excel(path)
    dt = pd.to_datetime(raw.iloc[:, 0].astype(str) + " " + raw.iloc[:, 1].astype(str), errors="coerce")
    frame = pd.DataFrame(
        {
            "dt": dt,
            "open": pd.to_numeric(raw.iloc[:, 2], errors="coerce"),
            "high": pd.to_numeric(raw.iloc[:, 3], errors="coerce"),
            "low": pd.to_numeric(raw.iloc[:, 4], errors="coerce"),
            "close": pd.to_numeric(raw.iloc[:, 5], errors="coerce"),
            "volume": pd.to_numeric(raw.iloc[:, 8], errors="coerce") if raw.shape[1] > 8 else pd.NA,
        }
    )
    frame = frame.dropna(subset=["dt", "open", "high", "low", "close"]).copy()
    frame["date"] = frame["dt"].dt.date
    frame["time"] = frame["dt"].dt.time
    return frame.drop_duplicates("dt").sort_values("dt").reset_index(drop=True)


def period_names() -> list[str]:
    return [name for name, _, _ in TARGET_PERIODS]


def periods_for_date(date_value: object) -> list[str]:
    stamp = pd.Timestamp(date_value)
    return [name for name, start, end in TARGET_PERIODS if start <= stamp <= end]


def expand_target_sessions(sessions: pd.DataFrame) -> pd.DataFrame:
    rows: list[pd.DataFrame] = []
    for name, start, end in TARGET_PERIODS:
        mask = (sessions["date"] >= start) & (sessions["date"] <= end)
        frame = sessions.loc[mask].copy()
        frame["period"] = name
        rows.append(frame)
    if not rows:
        return pd.DataFrame(columns=[*sessions.columns, "period"])
    return pd.concat(rows, ignore_index=True).sort_values(["period", "date"]).reset_index(drop=True)


def build_daily_sessions(prices: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for date_value, group in prices.groupby("date", sort=True):
        session = group.sort_values("dt").reset_index(drop=True)
        first = session.iloc[0]
        last = session.iloc[-1]
        open_price = float(first["open"])
        high_price = float(session["high"].max())
        low_price = float(session["low"].min())
        close_price = float(last["close"])
        day_range = high_price - low_price
        rows.append(
            {
                "date": pd.Timestamp(date_value),
                "periods": ", ".join(periods_for_date(date_value)),
                "first_dt": first["dt"],
                "first_time": first["time"],
                "last_dt": last["dt"],
                "last_time": last["time"],
                "minute_count": len(session),
                "is_full_session": last["time"] >= FULL_SESSION_END,
                "open": open_price,
                "high": high_price,
                "low": low_price,
                "close": close_price,
                "prev_close": np.nan,
                "gap_open_pct": np.nan,
                "open_to_close_pct": (close_price / open_price - 1.0) * 100.0,
                "day_return_pct": np.nan,
                "mfe_from_open_pct": (high_price / open_price - 1.0) * 100.0,
                "mae_from_open_pct": (low_price / open_price - 1.0) * 100.0,
                "range_pct": (day_range / open_price) * 100.0,
                "close_location": (close_price - low_price) / day_range if day_range else np.nan,
            }
        )

    sessions = pd.DataFrame(rows)
    sessions["prev_close"] = sessions["close"].shift(1)
    sessions["gap_open_pct"] = (sessions["open"] / sessions["prev_close"] - 1.0) * 100.0
    sessions["day_return_pct"] = (sessions["close"] / sessions["prev_close"] - 1.0) * 100.0
    sessions["gap_bucket"] = pd.cut(
        sessions["gap_open_pct"],
        bins=[-np.inf, 1, 2, 3, 4, 5, np.inf],
        labels=["<+1%", "+1~2%", "+2~3%", "+3~4%", "+4~5%", ">=+5%"],
        right=False,
    )
    sessions["one_way_up"] = (
        sessions["is_full_session"]
        & (sessions["gap_open_pct"] >= 1.0)
        & (sessions["open_to_close_pct"] > 0.0)
        & (sessions["close_location"] >= 0.70)
        & (sessions["mae_from_open_pct"] >= -0.40)
    )
    sessions["flow_type"] = np.select(
        [
            sessions["one_way_up"],
            (sessions["gap_open_pct"] >= 1.0) & (sessions["open_to_close_pct"] > 0.0),
            (sessions["gap_open_pct"] >= 1.0) & (sessions["open_to_close_pct"] <= 0.0),
        ],
        ["one_way_up", "up_but_choppy", "gap_fade"],
        default="not_gap_target",
    )
    return sessions


def checkpoint_return(session: pd.DataFrame, exit_label: str, open_price: float) -> float | pd.NA:
    if exit_label == "open":
        return 0.0
    target = pd.Timestamp(f"{session.iloc[0]['date']} {exit_label}").time()
    available = session[session["time"] <= target]
    if available.empty:
        return pd.NA
    close_price = float(available.iloc[-1]["close"])
    return (close_price / open_price - 1.0) * 100.0


def build_threshold_events(sessions: pd.DataFrame) -> pd.DataFrame:
    target_sessions = expand_target_sessions(sessions)
    rows: list[pd.DataFrame] = []
    for threshold in THRESHOLDS:
        frame = target_sessions[target_sessions["gap_open_pct"] >= threshold].copy()
        frame["threshold_pct"] = threshold
        rows.append(frame)
    if not rows:
        return pd.DataFrame()
    events = pd.concat(rows, ignore_index=True)
    return events.sort_values(["period", "threshold_pct", "date"]).reset_index(drop=True)


def build_checkpoint_returns(prices: pd.DataFrame, events: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    if events.empty:
        return pd.DataFrame()
    sessions = {pd.Timestamp(date_value).date(): group.sort_values("dt").reset_index(drop=True) for date_value, group in prices.groupby("date")}
    for event in events.itertuples(index=False):
        if not bool(event.is_full_session):
            continue
        session = sessions[pd.Timestamp(event.date).date()]
        open_price = float(event.open)
        for checkpoint in CHECKPOINTS:
            rows.append(
                {
                    "date": pd.Timestamp(event.date),
                    "period": event.period,
                    "threshold_pct": event.threshold_pct,
                    "gap_open_pct": event.gap_open_pct,
                    "checkpoint": checkpoint,
                    "return_from_open_pct": checkpoint_return(session, checkpoint, open_price),
                    "one_way_up": event.one_way_up,
                    "flow_type": event.flow_type,
                }
            )
    return pd.DataFrame(rows).dropna(subset=["return_from_open_pct"]).reset_index(drop=True)


def summarize_thresholds(events: pd.DataFrame) -> pd.DataFrame:
    grid = pd.MultiIndex.from_product(
        [period_names(), THRESHOLDS],
        names=["period", "threshold_pct"],
    ).to_frame(index=False)
    complete = events[events["is_full_session"]].copy()
    if complete.empty:
        grid["n"] = 0
        grid["one_way_days"] = 0
        return grid
    grouped = complete.groupby(["period", "threshold_pct"], sort=True)
    summary = grouped.agg(
        n=("date", "nunique"),
        one_way_days=("one_way_up", "sum"),
        one_way_rate_pct=("one_way_up", lambda values: values.mean() * 100.0),
        mean_gap_pct=("gap_open_pct", "mean"),
        median_gap_pct=("gap_open_pct", "median"),
        mean_open_to_close_pct=("open_to_close_pct", "mean"),
        median_open_to_close_pct=("open_to_close_pct", "median"),
        hit_rate_pct=("open_to_close_pct", lambda values: (values > 0).mean() * 100.0),
        mean_mfe_pct=("mfe_from_open_pct", "mean"),
        mean_mae_pct=("mae_from_open_pct", "mean"),
        mean_close_location=("close_location", "mean"),
    )
    summary = grid.merge(summary.reset_index(), on=["period", "threshold_pct"], how="left")
    summary["n"] = summary["n"].fillna(0).astype(int)
    summary["one_way_days"] = summary["one_way_days"].fillna(0).astype(int)
    return summary


def summarize_buckets(sessions: pd.DataFrame) -> pd.DataFrame:
    target_sessions = expand_target_sessions(sessions)
    target = target_sessions[(target_sessions["is_full_session"]) & (target_sessions["gap_open_pct"] >= 1.0)].copy()
    if target.empty:
        return pd.DataFrame()
    grouped = target.groupby(["period", "gap_bucket"], observed=True, sort=True)
    return grouped.agg(
        n=("date", "nunique"),
        one_way_days=("one_way_up", "sum"),
        one_way_rate_pct=("one_way_up", lambda values: values.mean() * 100.0),
        mean_gap_pct=("gap_open_pct", "mean"),
        mean_open_to_close_pct=("open_to_close_pct", "mean"),
        median_open_to_close_pct=("open_to_close_pct", "median"),
        mean_mae_pct=("mae_from_open_pct", "mean"),
        mean_mfe_pct=("mfe_from_open_pct", "mean"),
    ).reset_index()


def summarize_exits(checkpoints: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    grid = pd.MultiIndex.from_product(
        [period_names(), THRESHOLDS],
        names=["period", "threshold_pct"],
    ).to_frame(index=False)
    exits = checkpoints[checkpoints["checkpoint"].isin(EXIT_TIMES)].copy()
    if exits.empty:
        empty_best = grid.copy()
        empty_best["checkpoint"] = "no_data"
        empty_best["n"] = 0
        return pd.DataFrame(), empty_best, pd.DataFrame()
    grouped = exits.groupby(["period", "threshold_pct", "checkpoint"], sort=True)
    exit_summary = grouped.agg(
        n=("date", "nunique"),
        mean_return_pct=("return_from_open_pct", "mean"),
        median_return_pct=("return_from_open_pct", "median"),
        win_rate_pct=("return_from_open_pct", lambda values: (values > 0).mean() * 100.0),
        p25_return_pct=("return_from_open_pct", lambda values: values.quantile(0.25)),
        p75_return_pct=("return_from_open_pct", lambda values: values.quantile(0.75)),
        min_return_pct=("return_from_open_pct", "min"),
        max_return_pct=("return_from_open_pct", "max"),
    ).reset_index()
    best_exit = (
        exit_summary.sort_values(["period", "threshold_pct", "mean_return_pct", "median_return_pct"], ascending=[True, True, False, False])
        .groupby(["period", "threshold_pct"], as_index=False)
        .head(1)
        .reset_index(drop=True)
    )
    best_exit = grid.merge(best_exit, on=["period", "threshold_pct"], how="left")
    best_exit["checkpoint"] = best_exit["checkpoint"].fillna("no_data")
    best_exit["n"] = best_exit["n"].fillna(0).astype(int)
    event_best = (
        exits.sort_values(["period", "threshold_pct", "date", "return_from_open_pct"], ascending=[True, True, True, False])
        .groupby(["period", "threshold_pct", "date"], as_index=False)
        .head(1)
        .reset_index(drop=True)
        .rename(columns={"checkpoint": "event_best_exit_time", "return_from_open_pct": "event_best_return_pct"})
    )
    return exit_summary, best_exit, event_best


def build_quality_summary(prices: pd.DataFrame, sessions: pd.DataFrame, events: pd.DataFrame, source: Path) -> pd.DataFrame:
    target_sessions = expand_target_sessions(sessions)
    incomplete = target_sessions[~target_sessions["is_full_session"]]
    incomplete_labels = (
        incomplete.assign(label=incomplete["period"] + ":" + incomplete["date"].dt.strftime("%Y-%m-%d"))["label"].str.cat(sep=", ")
    )
    rows = [
        {"item": "source_file", "value": str(source)},
        {"item": "raw_rows", "value": len(prices)},
        {"item": "first_timestamp", "value": prices["dt"].min()},
        {"item": "last_timestamp", "value": prices["dt"].max()},
        {"item": "sessions_total", "value": sessions["date"].nunique()},
        {"item": "target_unique_sessions", "value": sessions.loc[sessions["periods"].ne(""), "date"].nunique()},
        {"item": "target_period_session_rows", "value": len(target_sessions)},
        {"item": "target_incomplete_sessions", "value": incomplete_labels},
        {"item": "threshold_events_complete_rows", "value": int(events["is_full_session"].sum()) if not events.empty else 0},
        {"item": "gap_definition", "value": "first available minute open / previous session close - 1"},
        {"item": "one_way_definition", "value": "full session, gap >= +1%, open-to-close > 0, close_location >= 0.70, open-to-low >= -0.40%"},
    ]
    return pd.DataFrame(rows)


def write_excel(
    paths: Paths,
    quality: pd.DataFrame,
    sessions: pd.DataFrame,
    events: pd.DataFrame,
    bucket_summary: pd.DataFrame,
    threshold_summary: pd.DataFrame,
    exit_summary: pd.DataFrame,
    best_exit: pd.DataFrame,
    event_best: pd.DataFrame,
    checkpoints: pd.DataFrame,
) -> None:
    paths.excel.parent.mkdir(parents=True, exist_ok=True)
    readme = pd.DataFrame(
        [
            {"item": "analysis", "value": "KOSPI200 gap-up one-way day and exit-time test"},
            {"item": "periods", "value": "2025-01~current, 2026-01~current, 2026-05-27~current"},
            {"item": "thresholds", "value": "gap_open_pct >= +1%, +2%, +3%, +4%, +5%"},
            {"item": "entry", "value": "first available minute open"},
            {"item": "exit_times", "value": ", ".join(EXIT_TIMES)},
            {"item": "costs", "value": "none; index-level flow test"},
        ]
    )
    target_sessions = expand_target_sessions(sessions)
    gap_days = target_sessions[target_sessions["gap_open_pct"] >= 1.0].copy()
    one_way_summary = (
        gap_days[gap_days["is_full_session"]]
        .groupby(["period", "flow_type"], sort=True)
        .agg(n=("date", "nunique"), mean_gap_pct=("gap_open_pct", "mean"), mean_open_to_close_pct=("open_to_close_pct", "mean"))
        .reset_index()
        if not gap_days.empty
        else pd.DataFrame()
    )
    sheets = {
        "readme": readme,
        "data_quality": quality,
        "daily_sessions": sessions,
        "gap_days_unique": gap_days,
        "bucket_summary": bucket_summary,
        "threshold_summary": threshold_summary,
        "one_way_summary": one_way_summary,
        "exit_summary": exit_summary,
        "best_exit": best_exit,
        "event_best_exit": event_best,
        "checkpoint_returns": checkpoints,
    }
    with pd.ExcelWriter(paths.excel, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm:ss", date_format="yyyy-mm-dd") as writer:
        for name, frame in sheets.items():
            output = frame.copy()
            for column in output.select_dtypes(include=["datetimetz"]).columns:
                output[column] = output[column].dt.tz_localize(None)
            output.to_excel(writer, sheet_name=name[:31], index=False)
        format_workbook(writer.book)


def format_workbook(workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1)):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 10), 34)


def pivot_metric(frame: pd.DataFrame, value: str) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame()
    return frame.pivot(index="threshold_pct", columns="period", values=value).sort_index()


def plot_results(paths: Paths, threshold_summary: pd.DataFrame, exit_summary: pd.DataFrame, checkpoints: pd.DataFrame, sessions: pd.DataFrame) -> None:
    paths.plot.parent.mkdir(parents=True, exist_ok=True)
    plt.style.use("seaborn-v0_8-whitegrid")
    fig, axes = plt.subplots(3, 2, figsize=(18, 15))
    fig.suptitle("KOSPI200 Gap-Up One-Way / Exit-Time Test", fontsize=18, fontweight="bold")

    count_pivot = pivot_metric(threshold_summary, "n")
    if not count_pivot.empty:
        count_pivot.plot(kind="bar", ax=axes[0, 0])
    axes[0, 0].set_title("Event count by period and gap threshold")
    axes[0, 0].set_xlabel("Gap threshold (%)")
    axes[0, 0].set_ylabel("Days")

    one_way_pivot = pivot_metric(threshold_summary, "one_way_rate_pct")
    if not one_way_pivot.empty:
        one_way_pivot.plot(kind="bar", ax=axes[0, 1])
    axes[0, 1].set_title("One-way-up share")
    axes[0, 1].set_xlabel("Gap threshold (%)")
    axes[0, 1].set_ylabel("% of qualifying days")

    checkpoint_order = {label: idx for idx, label in enumerate(CHECKPOINTS)}
    unique_checkpoints = checkpoints.drop_duplicates(["threshold_pct", "date", "checkpoint"])
    avg_path = (
        unique_checkpoints.groupby(["threshold_pct", "checkpoint"], sort=True)["return_from_open_pct"].mean().reset_index()
        if not unique_checkpoints.empty
        else pd.DataFrame()
    )
    if not avg_path.empty:
        for threshold, group in avg_path.groupby("threshold_pct", sort=True):
            ordered = group.assign(order=group["checkpoint"].map(checkpoint_order)).sort_values("order")
            axes[1, 0].plot(ordered["checkpoint"], ordered["return_from_open_pct"], marker="o", label=f">=+{threshold}%")
    axes[1, 0].set_title("Average intraday flow from open")
    axes[1, 0].set_xlabel("Time")
    axes[1, 0].set_ylabel("Return from open (%)")
    axes[1, 0].tick_params(axis="x", rotation=45)
    axes[1, 0].legend()

    unique_exits = unique_checkpoints[unique_checkpoints["checkpoint"].isin(EXIT_TIMES)]
    exit_all = (
        unique_exits.groupby(["threshold_pct", "checkpoint"], sort=True)["return_from_open_pct"].mean().reset_index(name="mean_return_pct")
        if not unique_exits.empty
        else pd.DataFrame()
    )
    if not exit_all.empty:
        for threshold, group in exit_all.groupby("threshold_pct", sort=True):
            ordered = group.assign(order=group["checkpoint"].map(checkpoint_order)).sort_values("order")
            axes[1, 1].plot(ordered["checkpoint"], ordered["mean_return_pct"], marker="o", label=f">=+{threshold}%")
    axes[1, 1].set_title("Exit-time mean return")
    axes[1, 1].set_xlabel("Exit time")
    axes[1, 1].set_ylabel("Mean return from open (%)")
    axes[1, 1].tick_params(axis="x", rotation=45)
    axes[1, 1].legend()

    period_path = (
        checkpoints[checkpoints["threshold_pct"] == 1]
        .groupby(["period", "checkpoint"], sort=True)["return_from_open_pct"]
        .mean()
        .reset_index()
        if not checkpoints.empty
        else pd.DataFrame()
    )
    if not period_path.empty:
        for period, group in period_path.groupby("period", sort=True):
            ordered = group.assign(order=group["checkpoint"].map(checkpoint_order)).sort_values("order")
            axes[2, 0].plot(ordered["checkpoint"], ordered["return_from_open_pct"], marker="o", label=period)
    axes[2, 0].set_title("Period flow for >=+1% gaps")
    axes[2, 0].set_xlabel("Time")
    axes[2, 0].set_ylabel("Return from open (%)")
    axes[2, 0].tick_params(axis="x", rotation=45)
    axes[2, 0].legend()

    gap_days = sessions[sessions["periods"].ne("") & sessions["is_full_session"] & (sessions["gap_open_pct"] >= 1.0)].copy()
    if not gap_days.empty:
        colors = np.where(gap_days["one_way_up"], "#1b9e77", "#d95f02")
        axes[2, 1].scatter(gap_days["gap_open_pct"], gap_days["open_to_close_pct"], c=colors, alpha=0.8)
        axes[2, 1].axhline(0, color="black", linewidth=1)
    axes[2, 1].set_title("Gap size vs open-to-close return")
    axes[2, 1].set_xlabel("Opening gap vs previous close (%)")
    axes[2, 1].set_ylabel("Open-to-close return (%)")

    for axis in axes.flat:
        axis.grid(True, alpha=0.3)
    fig.tight_layout(rect=(0, 0, 1, 0.97))
    fig.savefig(paths.plot, dpi=180)
    plt.close(fig)


def markdown_table(frame: pd.DataFrame, columns: list[str], max_rows: int = 20) -> str:
    if frame.empty:
        return "_데이터 없음_"
    output = frame.loc[:, columns].head(max_rows).copy()
    for column in output.columns:
        if pd.api.types.is_float_dtype(output[column]):
            output[column] = output[column].map(lambda value: "" if pd.isna(value) else f"{value:.2f}")
        else:
            output[column] = output[column].map(lambda value: "" if pd.isna(value) else str(value))
    headers = [str(column) for column in output.columns]
    rows = [[str(value) for value in row] for row in output.to_numpy()]
    widths = [
        max(len(header), *(len(row[index]) for row in rows)) if rows else len(header)
        for index, header in enumerate(headers)
    ]
    header_line = "| " + " | ".join(header.ljust(widths[index]) for index, header in enumerate(headers)) + " |"
    separator = "| " + " | ".join("-" * widths[index] for index in range(len(headers))) + " |"
    body = [
        "| " + " | ".join(row[index].ljust(widths[index]) for index in range(len(headers))) + " |"
        for row in rows
    ]
    return "\n".join([header_line, separator, *body])


def write_report(
    paths: Paths,
    quality: pd.DataFrame,
    sessions: pd.DataFrame,
    bucket_summary: pd.DataFrame,
    threshold_summary: pd.DataFrame,
    best_exit: pd.DataFrame,
    exit_summary: pd.DataFrame,
    event_best: pd.DataFrame,
    checkpoints: pd.DataFrame,
) -> None:
    target_sessions = expand_target_sessions(sessions)
    full_gap_days = target_sessions[(target_sessions["is_full_session"]) & (target_sessions["gap_open_pct"] >= 1.0)].copy()
    incomplete = target_sessions[~target_sessions["is_full_session"]]
    exits = checkpoints[checkpoints["checkpoint"].isin(EXIT_TIMES)].copy()
    unique_exits = exits.drop_duplicates(["threshold_pct", "date", "checkpoint"])
    overall_exit = (
        unique_exits.groupby(["threshold_pct", "checkpoint"], sort=True)
        .agg(
            n=("date", "nunique"),
            mean_return_pct=("return_from_open_pct", "mean"),
            median_return_pct=("return_from_open_pct", "median"),
            win_rate_pct=("return_from_open_pct", lambda values: (values > 0).mean() * 100.0),
        )
        .reset_index()
        .sort_values(["threshold_pct", "mean_return_pct"], ascending=[True, False])
        .groupby("threshold_pct", as_index=False)
        .head(1)
        if not unique_exits.empty
        else pd.DataFrame()
    )
    overall_grid = pd.DataFrame({"threshold_pct": list(THRESHOLDS)})
    overall_best = overall_grid.merge(overall_exit, on="threshold_pct", how="left")
    overall_best["checkpoint"] = overall_best["checkpoint"].fillna("no_data")
    overall_best["n"] = overall_best["n"].fillna(0).astype(int)

    period_counts = (
        target_sessions
        .groupby("period", sort=True)
        .agg(sessions=("date", "nunique"), full_sessions=("is_full_session", "sum"))
        .reset_index()
    )
    flow_mix = (
        full_gap_days.groupby(["period", "flow_type"], sort=True)
        .agg(n=("date", "nunique"), mean_gap_pct=("gap_open_pct", "mean"), mean_open_to_close_pct=("open_to_close_pct", "mean"))
        .reset_index()
        if not full_gap_days.empty
        else pd.DataFrame()
    )
    event_best_dist = (
        event_best.drop_duplicates(["threshold_pct", "date", "event_best_exit_time"])
        .groupby(["threshold_pct", "event_best_exit_time"], sort=True)
        .agg(n=("date", "nunique"), mean_event_best_return_pct=("event_best_return_pct", "mean"))
        .reset_index()
        .sort_values(["threshold_pct", "n", "mean_event_best_return_pct"], ascending=[True, False, False])
        if not event_best.empty
        else pd.DataFrame()
    )

    last_timestamp = quality.loc[quality["item"].eq("last_timestamp"), "value"].iloc[0]
    source_file = quality.loc[quality["item"].eq("source_file"), "value"].iloc[0]
    report = f"""# KOSPI200 갭 상승 원웨이/청산 시간 테스트

## 목적

`{source_file}`의 KOSPI200 1분 데이터를 사용해 다음 세 개의 누적 구간을 다시 테스트했다. 구간은 서로 겹치며, 예를 들어 2026년 5월 27일 이후 날짜는 세 구간 모두에 포함된다.

- 2025년 1월부터 현재까지
- 2026년 1월부터 현재까지
- 2026년 5월 27일부터 현재까지

갭은 **당일 첫 체결 가능 분봉의 시가 / 직전 거래일 15:30 종가 - 1**로 계산했다. 테스트 임계값은 `+1%`, `+2%`, `+3%`, `+4%`, `+5% 이상`이다. 청산 테스트는 지수 레벨 흐름 확인용이므로 비용은 넣지 않았다.

## 데이터와 제외 기준

- 데이터 마지막 시점: `{last_timestamp}`
- 정규장 전체 테스트는 15:30 분봉이 있는 날만 사용했다.
- 최신 데이터가 장중까지만 있는 날은 시그널 확인용으로 남기되 청산 시간 통계에서는 제외했다.

{markdown_table(period_counts, ["period", "sessions", "full_sessions"])}
"""
    if not incomplete.empty:
        report += "\n청산 통계에서 제외한 미완성 세션:\n\n"
        excluded = incomplete[["date", "first_time", "last_time", "gap_open_pct"]].copy()
        report += markdown_table(excluded, ["date", "first_time", "last_time", "gap_open_pct"])
        report += "\n"

    report += f"""
## 원웨이 장 정의

이 보고서에서 `one_way_up`은 다음 조건을 모두 만족하는 날이다.

1. 정규장 15:30까지 데이터가 있다.
2. 전일 종가 대비 첫 분봉 시가 갭이 `+1%` 이상이다.
3. 첫 분봉 시가 대비 종가 수익률이 양수다.
4. 종가가 당일 고저 범위의 상단 30% 안에 있다(`close_location >= 0.70`).
5. 첫 분봉 시가 대비 장중 저점 손실이 `-0.40%`보다 깊지 않다.

이 기준은 "올라서 시작한 뒤 크게 훼손되지 않고 상단에서 마감한 날"을 잡기 위한 보수적 정의다.

## 갭 구간별 구성

{markdown_table(bucket_summary, ["period", "gap_bucket", "n", "one_way_days", "one_way_rate_pct", "mean_gap_pct", "mean_open_to_close_pct", "mean_mae_pct", "mean_mfe_pct"])}

## 임계값별 요약

아래 표는 `>= +N%`의 누적 임계값 기준이다. 예를 들어 `+2%`는 `+2% 이상`인 날만 포함한다.

{markdown_table(threshold_summary, ["period", "threshold_pct", "n", "one_way_days", "one_way_rate_pct", "mean_gap_pct", "mean_open_to_close_pct", "hit_rate_pct", "mean_mae_pct", "mean_mfe_pct"], max_rows=30)}

## 원웨이/페이드 흐름

{markdown_table(flow_mix, ["period", "flow_type", "n", "mean_gap_pct", "mean_open_to_close_pct"])}

흐름 해석은 엑셀의 `checkpoint_returns`와 PNG의 평균 경로 subplot에서 확인할 수 있다. 핵심은 갭이 커질수록 표본 수가 급격히 줄기 때문에, `+3%` 이상 구간은 방향성 결론보다 개별 날짜 점검 성격이 강하다는 점이다.

## 청산 시간 결론

기간별/임계값별 평균 수익률이 가장 높았던 청산 시간은 다음과 같다. `no_data`는 해당 구간에 관측치가 없었다는 뜻이다.

{markdown_table(best_exit, ["period", "threshold_pct", "checkpoint", "n", "mean_return_pct", "median_return_pct", "win_rate_pct"], max_rows=30)}

전체 날짜를 중복 제거해 임계값별로 보면 다음 시간이 평균 기준 최적이었다.

{markdown_table(overall_best, ["threshold_pct", "checkpoint", "n", "mean_return_pct", "median_return_pct", "win_rate_pct"])}

이 결과는 평균 기준이다. 중복 제거한 단일 날짜의 사후 최적 청산 시간 분포는 다음과 같다.

{markdown_table(event_best_dist, ["threshold_pct", "event_best_exit_time", "n", "mean_event_best_return_pct"], max_rows=30)}

## 사용 방법

- `momentum_strategy_graph.png`: 6개 subplot으로 구성된 요약 그래프.
- `momentum_strategy_results.xlsx`: 일별 세션, 갭 이벤트, 원웨이 분류, 청산 시간별 결과, 일자별 사후 최적 청산 시간을 담은 워크북.
- `momentum_strategy_report.md`: 본 보고서.

재생성 명령:

```bash
python etc/momentum/build_kospi200_gap_oneway_report.py
```
"""
    paths.report.write_text(report, encoding="utf-8")


def run() -> Paths:
    source = find_source_file()
    paths = Paths(source=source)
    prices = load_prices(paths.source)
    sessions = build_daily_sessions(prices)
    events = build_threshold_events(sessions)
    checkpoints = build_checkpoint_returns(prices, events)
    threshold_summary = summarize_thresholds(events)
    bucket_summary = summarize_buckets(sessions)
    exit_summary, best_exit, event_best = summarize_exits(checkpoints)
    quality = build_quality_summary(prices, sessions, events, paths.source)
    write_excel(paths, quality, sessions, events, bucket_summary, threshold_summary, exit_summary, best_exit, event_best, checkpoints)
    plot_results(paths, threshold_summary, exit_summary, checkpoints, sessions)
    write_report(paths, quality, sessions, bucket_summary, threshold_summary, best_exit, exit_summary, event_best, checkpoints)
    return paths


def main() -> int:
    paths = run()
    print(f"source: {paths.source}")
    print(f"plot: {paths.plot}")
    print(f"excel: {paths.excel}")
    print(f"report: {paths.report}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
