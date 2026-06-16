from __future__ import annotations

import argparse
import hashlib
import json
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import load_workbook

from etfs import paths


@dataclass(frozen=True, slots=True)
class ValidationHolding:
    ticker: str
    ticker_raw: str
    name: str
    quantity: float
    amount: float
    weight: float


@dataclass(frozen=True, slots=True)
class ValidationSnapshot:
    as_of: str
    equity_holdings: list[ValidationHolding]
    cash: dict[str, object]


@dataclass(frozen=True, slots=True)
class ValidationFixture:
    schema_version: str
    source_type: str
    etf_code: str
    etf_code_raw: str
    etf_name: str
    index_code: str
    source: dict[str, object]
    snapshots: list[ValidationSnapshot]


@dataclass(frozen=True, slots=True)
class ValidationResult:
    validation_type: str
    etf_code: str
    index_code: str
    as_of: str
    status: str
    checks: dict[str, str]
    metrics: dict[str, object]
    differences: list[dict[str, object]]


def parse_validation_workbook(
    path: Path,
    *,
    index_code_by_etf: Mapping[str, str] | None = None,
) -> ValidationFixture:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    header_row = _find_header_row(ws)
    rows = [row for row in ws.iter_rows(min_row=header_row + 1, values_only=True) if any(value is not None for value in row)]
    if not rows:
        raise ValueError(f"no validation holdings rows found: {path}")

    etf_code_raw = str(rows[0][1])
    etf_code = _normalize_etf_code(etf_code_raw)
    etf_name = str(rows[0][2])
    by_date: dict[str, list[tuple[object, ...]]] = {}
    for row in rows:
        as_of = _date_string(row[0])
        by_date.setdefault(as_of, []).append(row)

    snapshots = [_snapshot_from_rows(as_of, items) for as_of, items in sorted(by_date.items())]
    return ValidationFixture(
        schema_version="1.0",
        source_type="etf_portfolio_component_xlsx",
        etf_code=etf_code,
        etf_code_raw=etf_code_raw,
        etf_name=etf_name,
        index_code=dict(index_code_by_etf or {}).get(etf_code, ""),
        source={
            "path": str(path),
            "sha256": _sha256(path),
            "source_type": "etf_portfolio_component_xlsx",
            "extraction_status": "parsed",
        },
        snapshots=snapshots,
    )


def write_validation_fixtures(
    workbook_paths: Iterable[Path],
    output_dir: Path,
    *,
    index_code_by_etf: Mapping[str, str] | None = None,
) -> Path:
    fixtures = [
        parse_validation_workbook(path, index_code_by_etf=index_code_by_etf)
        for path in workbook_paths
    ]
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "validation_fixtures.json"
    payload = {
        "schema_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "count": len(fixtures),
        "fixtures": [asdict(fixture) for fixture in fixtures],
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def load_validation_fixtures(path: Path) -> list[ValidationFixture]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return [_fixture_from_mapping(item) for item in payload.get("fixtures", [])]


def build_validation_results(
    fixtures: Iterable[ValidationFixture],
    specs: Iterable[Mapping[str, object]],
) -> list[ValidationResult]:
    specs_by_index = {str(spec.get("index_code", "")): spec for spec in specs}
    results: list[ValidationResult] = []
    for fixture in fixtures:
        spec = specs_by_index.get(fixture.index_code, {})
        expected_count = _expected_constituent_count(spec)
        for snapshot in fixture.snapshots:
            official_count = len(snapshot.equity_holdings)
            count_passed = expected_count is None or official_count == expected_count
            cash_excluded = bool(snapshot.cash)
            status = "passed" if count_passed and cash_excluded else "failed"
            differences: list[dict[str, object]] = []
            if not count_passed:
                differences.append(
                    {
                        "type": "constituent_count_mismatch",
                        "expected_count": expected_count,
                        "official_equity_count": official_count,
                    }
                )
            results.append(
                ValidationResult(
                    validation_type="etf_holdings_constituents",
                    etf_code=fixture.etf_code,
                    index_code=fixture.index_code,
                    as_of=snapshot.as_of,
                    status=status,
                    checks={
                        "constituent_count": "passed" if count_passed else "failed",
                        "cash_excluded": "passed" if cash_excluded else "failed",
                    },
                    metrics={
                        "expected_constituent_count": expected_count,
                        "official_equity_count": official_count,
                        "official_weight_sum": sum(holding.weight for holding in snapshot.equity_holdings)
                        + float(snapshot.cash.get("weight", 0.0) or 0.0),
                    },
                    differences=differences,
                )
            )
    return results


def write_validation_results(fixtures_path: Path, specs_path: Path, output_dir: Path) -> Path:
    fixtures = load_validation_fixtures(fixtures_path)
    specs_payload = json.loads(specs_path.read_text(encoding="utf-8"))
    results = build_validation_results(fixtures, specs_payload.get("indices", []))
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "validation_results.json"
    payload = {
        "schema_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "count": len(results),
        "results": [asdict(result) for result in results],
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def build_target_weight_validation_results(
    target_weight_payload: Mapping[str, object],
    fixtures: Iterable[ValidationFixture],
    *,
    weight_tolerance: float = 0.0,
) -> list[ValidationResult]:
    snapshots = {
        (fixture.index_code, snapshot.as_of): (fixture, snapshot)
        for fixture in fixtures
        for snapshot in fixture.snapshots
    }
    results: list[ValidationResult] = []
    for item in target_weight_payload.get("results", []):
        target_result = _mapping(item)
        index_code = str(target_result.get("index_code", ""))
        as_of = str(target_result.get("as_of", ""))
        fixture, snapshot = snapshots.get((index_code, as_of), (None, None))
        if fixture is None or snapshot is None:
            results.append(
                ValidationResult(
                    validation_type="target_weights_vs_etf_holdings",
                    etf_code="",
                    index_code=index_code,
                    as_of=as_of,
                    status="failed",
                    checks={"fixture_snapshot": "failed"},
                    metrics={},
                    differences=[{"type": "fixture_snapshot_missing", "index_code": index_code, "as_of": as_of}],
                )
            )
            continue
        results.append(
            _target_weight_validation_result(
                target_result,
                fixture,
                snapshot,
                weight_tolerance=weight_tolerance,
            )
        )
    return results


def write_target_weight_validation_results(
    fixtures_path: Path,
    target_weights_path: Path,
    output_dir: Path,
    *,
    weight_tolerance: float = 0.0,
) -> Path:
    fixtures = load_validation_fixtures(fixtures_path)
    target_payload = json.loads(target_weights_path.read_text(encoding="utf-8"))
    results = build_target_weight_validation_results(target_payload, fixtures, weight_tolerance=weight_tolerance)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "target_weight_validation.json"
    payload = {
        "schema_version": "1.0",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "count": len(results),
        "weight_tolerance": weight_tolerance,
        "results": [asdict(result) for result in results],
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def _target_weight_validation_result(
    target_result: Mapping[str, object],
    fixture: ValidationFixture,
    snapshot: ValidationSnapshot,
    *,
    weight_tolerance: float,
) -> ValidationResult:
    target_weights = _target_weights_by_security(target_result.get("target_weights"))
    official_weights = {holding.ticker: holding.weight for holding in snapshot.equity_holdings}
    target_codes = set(target_weights)
    official_codes = set(official_weights)
    differences: list[dict[str, object]] = []
    for code in sorted(target_codes - official_codes):
        differences.append(
            {
                "type": "missing_in_official_holdings",
                "security_code": code,
                "target_weight": round(target_weights[code], 12),
            }
        )
    for code in sorted(official_codes - target_codes):
        differences.append(
            {
                "type": "extra_in_official_holdings",
                "security_code": code,
                "official_weight": round(official_weights[code], 12),
            }
        )
    abs_differences: list[float] = []
    for code in sorted(target_codes & official_codes):
        difference = target_weights[code] - official_weights[code]
        abs_differences.append(abs(difference))
        if abs(difference) > weight_tolerance:
            differences.append(
                {
                    "type": "weight_difference",
                    "security_code": code,
                    "target_weight": round(target_weights[code], 12),
                    "official_weight": round(official_weights[code], 12),
                    "difference": round(difference, 12),
                }
            )
    membership_passed = target_codes == official_codes
    max_abs_difference = round(max(abs_differences, default=0.0), 12)
    weight_passed = max_abs_difference <= weight_tolerance
    return ValidationResult(
        validation_type="target_weights_vs_etf_holdings",
        etf_code=fixture.etf_code,
        index_code=fixture.index_code,
        as_of=snapshot.as_of,
        status="passed" if membership_passed and weight_passed else "failed",
        checks={
            "constituent_membership": "passed" if membership_passed else "failed",
            "weight_tolerance": "passed" if weight_passed else "failed",
        },
        metrics={
            "target_constituent_count": len(target_codes),
            "official_equity_count": len(official_codes),
            "common_constituent_count": len(target_codes & official_codes),
            "max_abs_weight_difference": max_abs_difference,
            "total_abs_weight_difference": round(sum(abs_differences), 12),
        },
        differences=differences,
    )


def _target_weights_by_security(value: object) -> dict[str, float]:
    if not isinstance(value, list):
        raise ValueError("target_weights must be a list")
    weights: dict[str, float] = {}
    for item in value:
        row = _mapping(item)
        code = _normalize_security_code(str(row.get("security_code", "")))
        if not code:
            raise ValueError("target weight security_code is required")
        weights[code] = float(row.get("target_weight", 0.0) or 0.0)
    return weights


def _snapshot_from_rows(as_of: str, rows: list[tuple[object, ...]]) -> ValidationSnapshot:
    equity: list[ValidationHolding] = []
    cash: dict[str, object] = {}
    for row in rows:
        ticker_raw = "" if row[3] is None else str(row[3])
        name = str(row[4])
        weight = float(row[7] or 0.0) / 100.0
        amount = float(row[6] or 0.0)
        quantity = float(row[5] or 0.0)
        if not ticker_raw or name == "원화현금":
            cash = {"name": name, "quantity": quantity, "amount": amount, "weight": weight}
            continue
        equity.append(
            ValidationHolding(
                ticker=_normalize_security_code(ticker_raw),
                ticker_raw=ticker_raw,
                name=name,
                quantity=quantity,
                amount=amount,
                weight=weight,
            )
        )
    return ValidationSnapshot(as_of=as_of, equity_holdings=equity, cash=cash)


def _find_header_row(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        values = [cell.value for cell in row]
        if values and values[0] == "Date" and "구성종목코드" in values:
            return row[0].row
    raise ValueError("validation workbook header row not found")


def _normalize_etf_code(value: str) -> str:
    return value[1:] if value.startswith("A") else value


def _normalize_security_code(value: str) -> str:
    match = re.search(r"(\d{6})$", value)
    return match.group(1) if match else value


def _date_string(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _expected_constituent_count(spec: Mapping[str, object]) -> int | None:
    selection = _mapping(spec.get("selection"))
    value = selection.get("total_constituents")
    return int(value) if value is not None else None


def _fixture_from_mapping(item: Mapping[str, object]) -> ValidationFixture:
    return ValidationFixture(
        schema_version=str(item.get("schema_version", "")),
        source_type=str(item.get("source_type", "")),
        etf_code=str(item.get("etf_code", "")),
        etf_code_raw=str(item.get("etf_code_raw", "")),
        etf_name=str(item.get("etf_name", "")),
        index_code=str(item.get("index_code", "")),
        source=dict(_mapping(item.get("source"))),
        snapshots=[
            ValidationSnapshot(
                as_of=str(snapshot.get("as_of", "")),
                equity_holdings=[
                    ValidationHolding(
                        ticker=str(holding.get("ticker", "")),
                        ticker_raw=str(holding.get("ticker_raw", "")),
                        name=str(holding.get("name", "")),
                        quantity=float(holding.get("quantity", 0.0) or 0.0),
                        amount=float(holding.get("amount", 0.0) or 0.0),
                        weight=float(holding.get("weight", 0.0) or 0.0),
                    )
                    for holding in snapshot.get("equity_holdings", [])
                ],
                cash=dict(_mapping(snapshot.get("cash"))),
            )
            for snapshot in item.get("snapshots", [])
        ],
    )


def _mapping(value: object) -> Mapping[str, object]:
    return value if isinstance(value, Mapping) else {}


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Normalize ETF validation workbooks and compare basic fixture checks.")
    parser.add_argument("--input", nargs="+", default=["etfs/validation_A0167A0.xlsx"])
    parser.add_argument("--output-dir", default=paths.VALIDATION_OUTPUT_DIR.as_posix())
    parser.add_argument("--specs", default=paths.FNGUIDE_METHODOLOGY_SPECS_JSON.as_posix())
    parser.add_argument("--write-results", action="store_true")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    index_map = {"0167A0": "FI00.WLT.KSS"}
    fixtures_path = write_validation_fixtures(
        [Path(value) for value in args.input],
        Path(args.output_dir),
        index_code_by_etf=index_map,
    )
    if args.write_results:
        results_path = write_validation_results(fixtures_path, Path(args.specs), Path(args.output_dir))
        print(f"wrote {fixtures_path} and {results_path}")
    else:
        print(f"wrote {fixtures_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
