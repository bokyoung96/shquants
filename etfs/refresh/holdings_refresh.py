from __future__ import annotations

import argparse
import json
import shutil
import subprocess
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Iterable, Mapping, Protocol

import pandas as pd
from openpyxl import load_workbook

from etfs import paths
from etfs.fnguide.pipeline import index_code_by_etf_from_specs


RESULT_COLUMNS = [
    "as_of",
    "etf_code",
    "etf_code_raw",
    "etf_name",
    "index_code",
    "security_code",
    "security_code_raw",
    "security_name",
    "quantity",
    "market_value",
    "weight",
    "weight_percent",
    "is_cash",
]


@dataclass(frozen=True, slots=True)
class RefreshTarget:
    etf_code: str
    excel_code: str
    etf_name: str
    index_code: str
    source: str


class RefreshDriver(Protocol):
    def refresh(self, *, template_path: Path, excel_code: str) -> Path:
        ...


class BatchRefreshDriver(Protocol):
    def refresh_many(self, *, template_path: Path, targets: Iterable[RefreshTarget]) -> dict[str, Path]:
        ...


def load_refresh_targets(
    *,
    specs_path: Path = paths.FNGUIDE_METHODOLOGY_SPECS_JSON,
    rules_path: Path = paths.FNGUIDE_RULES_JSON,
) -> list[RefreshTarget]:
    specs_payload = json.loads(specs_path.read_text(encoding="utf-8")) if specs_path.exists() else {}
    rules_payload = json.loads(rules_path.read_text(encoding="utf-8")) if rules_path.exists() else {}
    return load_refresh_targets_from_payloads(
        specs_payload=specs_payload,
        rules_payload=rules_payload,
        specs_path=specs_path,
        rules_path=rules_path,
    )


def load_refresh_targets_from_payloads(
    *,
    specs_payload: Mapping[str, object],
    rules_payload: Mapping[str, object],
    specs_path: Path,
    rules_path: Path,
) -> list[RefreshTarget]:
    targets = _targets_from_specs(specs_payload)
    if targets:
        return targets
    return _targets_from_rules(rules_payload)


def load_refresh_targets_from_ticker_workbook(
    path: Path,
    *,
    visible_only: bool = True,
) -> list[RefreshTarget]:
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    header_row = _find_ticker_header_row(ws)
    targets: list[RefreshTarget] = []
    seen: set[str] = set()
    for row_number in range(header_row + 1, ws.max_row + 1):
        if visible_only and ws.row_dimensions[row_number].hidden:
            continue
        etf_code_raw = str(ws.cell(row_number, 1).value or "").strip()
        if not etf_code_raw:
            continue
        etf_code = _strip_a_prefix(etf_code_raw)
        if not etf_code or etf_code in seen:
            continue
        seen.add(etf_code)
        targets.append(
            RefreshTarget(
                etf_code=etf_code,
                excel_code=_excel_code(etf_code),
                etf_name=str(ws.cell(row_number, 2).value or "").strip(),
                index_code=str(ws.cell(row_number, 7).value or "").strip(),
                source="ticker_workbook",
            )
        )
    return targets


def refresh_targets_to_parquet(
    *,
    targets: Iterable[RefreshTarget],
    template_path: Path,
    output_path: Path,
    index_code_by_etf: Mapping[str, str],
    driver: RefreshDriver,
) -> Path:
    records: list[dict[str, object]] = []
    for target in targets:
        refreshed_workbook = driver.refresh(template_path=template_path, excel_code=target.excel_code)
        records.extend(extract_sheet_records(refreshed_workbook, index_code_by_etf=index_code_by_etf))
    return write_records_parquet(records, output_path)


def refresh_targets_to_parquet_files(
    *,
    targets: Iterable[RefreshTarget],
    template_path: Path,
    output_dir: Path,
    index_code_by_etf: Mapping[str, str],
    driver: RefreshDriver,
) -> dict[str, Path]:
    outputs: dict[str, Path] = {}
    for target in targets:
        refreshed_workbook = driver.refresh(template_path=template_path, excel_code=target.excel_code)
        records = extract_sheet_records(refreshed_workbook, index_code_by_etf=index_code_by_etf)
        _require_target_records(records, target, refreshed_workbook)
        output_path = ticker_output_path(output_dir, target.etf_code)
        write_records_parquet(records, output_path)
        outputs[target.etf_code] = output_path
    return outputs


def refresh_targets_to_parquet_files_batch(
    *,
    targets: Iterable[RefreshTarget],
    template_path: Path,
    output_dir: Path,
    index_code_by_etf: Mapping[str, str],
    driver: RefreshDriver | BatchRefreshDriver,
) -> dict[str, Path]:
    target_list = list(targets)
    if not target_list:
        return {}
    if not hasattr(driver, "refresh_many"):
        return refresh_targets_to_parquet_files(
            targets=target_list,
            template_path=template_path,
            output_dir=output_dir,
            index_code_by_etf=index_code_by_etf,
            driver=driver,
        )

    refreshed_workbooks = driver.refresh_many(template_path=template_path, targets=target_list)
    outputs: dict[str, Path] = {}
    for target in target_list:
        refreshed_workbook = refreshed_workbooks[target.etf_code]
        records = extract_sheet_records(refreshed_workbook, index_code_by_etf=index_code_by_etf)
        _require_target_records(records, target, refreshed_workbook)
        output_path = ticker_output_path(output_dir, target.etf_code)
        write_records_parquet(records, output_path)
        outputs[target.etf_code] = output_path
    return outputs


def filter_refresh_targets(
    targets: Iterable[RefreshTarget],
    *,
    tickers: Iterable[str] = (),
    limit: int | None = None,
) -> list[RefreshTarget]:
    selected = list(targets)
    ticker_set = {_strip_a_prefix(str(ticker).strip()) for ticker in tickers if str(ticker).strip()}
    if ticker_set:
        selected = [target for target in selected if target.etf_code in ticker_set]
    if limit is not None:
        if limit <= 0:
            raise ValueError("--limit must be positive")
        selected = selected[:limit]
    return selected


def extract_sheet_records(
    workbook_path: Path,
    *,
    index_code_by_etf: Mapping[str, str] | None = None,
) -> list[dict[str, object]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb.active
    records: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=7, min_col=1, max_col=8, values_only=True):
        if not any(value is not None for value in row):
            continue
        etf_code_raw = "" if row[1] is None else str(row[1]).strip()
        etf_code = _strip_a_prefix(etf_code_raw)
        security_code_raw = "" if row[3] is None else str(row[3]).strip()
        security_name = "" if row[4] is None else str(row[4]).strip()
        is_cash = not security_code_raw or security_name == "원화현금"
        records.append(
            {
                "as_of": _date_string(row[0]),
                "etf_code": etf_code,
                "etf_code_raw": etf_code_raw,
                "etf_name": "" if row[2] is None else str(row[2]).strip(),
                "index_code": dict(index_code_by_etf or {}).get(etf_code, ""),
                "security_code": "" if is_cash else _strip_a_prefix(security_code_raw),
                "security_code_raw": security_code_raw,
                "security_name": security_name,
                "quantity": _float_or_zero(row[5]),
                "market_value": _float_or_zero(row[6]),
                "weight": _float_or_zero(row[7]) / 100.0,
                "weight_percent": _float_or_zero(row[7]),
                "is_cash": is_cash,
            }
        )
    if not records:
        raise ValueError(f"no refreshed holdings rows found: {workbook_path}")
    return records


def write_records_parquet(records: Iterable[Mapping[str, object]], output_path: Path) -> Path:
    frame = pd.DataFrame(list(records), columns=RESULT_COLUMNS)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    frame.to_parquet(output_path, index=False)
    return output_path


def write_records_by_ticker_parquet(records: Iterable[Mapping[str, object]], output_dir: Path) -> dict[str, Path]:
    by_ticker: dict[str, list[Mapping[str, object]]] = {}
    for record in records:
        etf_code = _strip_a_prefix(str(record.get("etf_code", "")).strip())
        if not etf_code:
            raise ValueError("refreshed holding record requires etf_code")
        by_ticker.setdefault(etf_code, []).append(record)
    outputs: dict[str, Path] = {}
    for etf_code, items in sorted(by_ticker.items()):
        output_path = ticker_output_path(output_dir, etf_code)
        write_records_parquet(items, output_path)
        outputs[etf_code] = output_path
    return outputs


def _require_target_records(records: list[Mapping[str, object]], target: RefreshTarget, workbook_path: Path) -> None:
    etf_codes = {_strip_a_prefix(str(record.get("etf_code", "")).strip()) for record in records}
    etf_codes.discard("")
    if etf_codes != {target.etf_code}:
        raise RuntimeError(
            f"DataGuide6 refresh did not return holdings for {target.excel_code}: "
            f"found {sorted(etf_codes)} in {workbook_path}"
        )


def ticker_output_path(output_dir: Path, etf_code: str) -> Path:
    return output_dir / f"holdings_{_safe_name(_strip_a_prefix(etf_code))}.parquet"


class DataguideExcelRefreshDriver:
    def __init__(
        self,
        *,
        work_dir: Path | None = None,
        visible: bool = False,
        wait_seconds: float = 10.0,
        runner: Callable[[list[str]], subprocess.CompletedProcess[str]] | None = None,
    ) -> None:
        self.work_dir = work_dir
        self.visible = visible
        self.wait_seconds = wait_seconds
        self.runner = runner or _run_powershell

    def refresh(self, *, template_path: Path, excel_code: str) -> Path:
        work_dir = self._work_dir()
        work_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = work_dir / f"{_safe_name(excel_code)}_{template_path.name}"
        script_path = work_dir / "refresh.ps1"
        shutil.copy2(template_path, workbook_path)
        script_path.write_text(_POWERSHELL_REFRESH_SCRIPT, encoding="utf-8")
        command = [
            "powershell",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script_path),
            "-WorkbookPath",
            str(workbook_path.resolve()),
            "-ExcelCode",
            excel_code,
            "-WaitSeconds",
            str(self.wait_seconds),
        ]
        if self.visible:
            command.append("-Visible")
        result = self.runner(command)
        if result.returncode != 0:
            raise RuntimeError(
                "DataGuide6 Excel refresh failed: "
                + (result.stderr.strip() or result.stdout.strip() or f"exit code {result.returncode}")
            )
        return workbook_path

    def refresh_many(self, *, template_path: Path, targets: Iterable[RefreshTarget]) -> dict[str, Path]:
        target_list = list(targets)
        if not target_list:
            return {}

        work_dir = self._work_dir()
        workbooks_dir = work_dir / "workbooks"
        work_dir.mkdir(parents=True, exist_ok=True)
        workbooks_dir.mkdir(parents=True, exist_ok=True)

        workbook_path = work_dir / f"refresh_{template_path.name}"
        script_path = work_dir / "refresh.ps1"
        targets_path = work_dir / "targets.json"
        outputs = {
            target.etf_code: workbooks_dir / f"holdings_{_safe_name(target.etf_code)}.xlsx"
            for target in target_list
        }
        payload = [
            {
                "etf_code": target.etf_code,
                "excel_code": target.excel_code,
                "output_path": str(outputs[target.etf_code].resolve()),
            }
            for target in target_list
        ]

        shutil.copy2(template_path, workbook_path)
        script_path.write_text(_POWERSHELL_BATCH_REFRESH_SCRIPT, encoding="utf-8")
        targets_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        command = [
            "powershell",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script_path),
            "-WorkbookPath",
            str(workbook_path.resolve()),
            "-TargetsPath",
            str(targets_path.resolve()),
            "-WaitSeconds",
            str(self.wait_seconds),
        ]
        if self.visible:
            command.append("-Visible")
        result = self.runner(command)
        if result.returncode != 0:
            raise RuntimeError(
                "DataGuide6 Excel batch refresh failed: "
                + (result.stderr.strip() or result.stdout.strip() or f"exit code {result.returncode}")
            )
        missing = [str(path) for path in outputs.values() if not path.exists()]
        if missing:
            raise RuntimeError("DataGuide6 Excel batch refresh did not create workbook(s): " + ", ".join(missing))
        return outputs

    def _work_dir(self) -> Path:
        if self.work_dir is not None:
            return self.work_dir
        return paths.REFRESH_WORK_DIR


def _run_powershell(command: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(command, capture_output=True, text=True, check=False)


_POWERSHELL_REFRESH_SCRIPT = r"""
param(
    [Parameter(Mandatory=$true)][string]$WorkbookPath,
    [Parameter(Mandatory=$true)][string]$ExcelCode,
    [double]$WaitSeconds = 1.0,
    [switch]$Visible
)

$excel = $null
$workbook = $null
$createdExcel = $false
try {
    $excel = [Runtime.InteropServices.Marshal]::GetActiveObject("Excel.Application")
    $createdExcel = $false
}
catch {
    if ($null -eq $excel) {
        $dgExe = "C:\Fnguide\DataGuide6\Conpo\DataGuide6.exe"
        if (Test-Path $dgExe) {
            Start-Process -FilePath $dgExe -WindowStyle Hidden
            Start-Sleep -Seconds 8
            try {
                $excel = [Runtime.InteropServices.Marshal]::GetActiveObject("Excel.Application")
                $createdExcel = $false
            } catch {
            }
        }
    }
    if ($null -eq $excel) {
        $excel = New-Object -ComObject Excel.Application
        $createdExcel = $true
    }
}

try {
    $excel.Visible = [bool]$Visible
    $excel.DisplayAlerts = $false
    foreach ($addin in $excel.COMAddIns) {
        $progId = [string]$addin.ProgId
        $description = [string]$addin.Description
        if (($progId -like "*DG*") -or ($progId -like "*DataGuide*") -or ($description -like "*DataGuide*")) {
            try {
                $addin.Connect = $true
            } catch {
            }
        }
    }
    $workbook = $excel.Workbooks.Open($WorkbookPath)
    $firstSheet = $workbook.Worksheets.Item(1)
    $firstSheet.Range("B3").Value2 = $ExcelCode

    for ($sheetIndex = 1; $sheetIndex -le $workbook.Sheets.Count; $sheetIndex++) {
        $sheet = $workbook.Sheets.Item($sheetIndex)
        $sheet.Activate() | Out-Null
        for ($linkIndex = 1; $linkIndex -le $sheet.Hyperlinks.Count; $linkIndex++) {
            $hyperlink = $sheet.Hyperlinks.Item($linkIndex)
            $screenTip = [string]$hyperlink.ScreenTip
            if ($screenTip.ToLowerInvariant() -ne "DataGuide6".ToLowerInvariant()) {
                continue
            }
            $hyperlink.Range.Select() | Out-Null
            $hyperlink.Follow($false, $true)
            while (-not $excel.Ready) {
                Start-Sleep -Milliseconds 100
            }
            Start-Sleep -Seconds $WaitSeconds
            $deadline = (Get-Date).AddSeconds(60)
            while ((Get-Date) -lt $deadline) {
                $refreshedCode = [string]$sheet.Range("B7").Value2
                if ($refreshedCode -eq $ExcelCode) {
                    break
                }
                Start-Sleep -Seconds 1
            }
        }
    }

    $workbook.Save()
}
finally {
    if ($null -ne $workbook) {
        $workbook.Close($true)
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null
    }
    if ($null -ne $excel) {
        if ($createdExcel) {
            $excel.Quit()
        }
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
    }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
"""


_POWERSHELL_BATCH_REFRESH_SCRIPT = r"""
param(
    [Parameter(Mandatory=$true)][string]$WorkbookPath,
    [Parameter(Mandatory=$true)][string]$TargetsPath,
    [double]$WaitSeconds = 1.0,
    [switch]$Visible
)

$excel = $null
$workbook = $null
$createdExcel = $false
try {
    $excel = [Runtime.InteropServices.Marshal]::GetActiveObject("Excel.Application")
    $createdExcel = $false
}
catch {
    if ($null -eq $excel) {
        $dgExe = "C:\Fnguide\DataGuide6\Conpo\DataGuide6.exe"
        if (Test-Path $dgExe) {
            Start-Process -FilePath $dgExe -WindowStyle Hidden
            Start-Sleep -Seconds 8
            try {
                $excel = [Runtime.InteropServices.Marshal]::GetActiveObject("Excel.Application")
                $createdExcel = $false
            } catch {
            }
        }
    }
    if ($null -eq $excel) {
        $excel = New-Object -ComObject Excel.Application
        $createdExcel = $true
    }
}

try {
    $targets = Get-Content -LiteralPath $TargetsPath -Raw -Encoding UTF8 | ConvertFrom-Json
    $excel.Visible = [bool]$Visible
    $excel.DisplayAlerts = $false
    foreach ($addin in $excel.COMAddIns) {
        $progId = [string]$addin.ProgId
        $description = [string]$addin.Description
        if (($progId -like "*DG*") -or ($progId -like "*DataGuide*") -or ($description -like "*DataGuide*")) {
            try {
                $addin.Connect = $true
            } catch {
            }
        }
    }
    $workbook = $excel.Workbooks.Open($WorkbookPath)
    $firstSheet = $workbook.Worksheets.Item(1)

    foreach ($target in $targets) {
        $excelCode = [string]$target.excel_code
        $outputPath = [string]$target.output_path
        $firstSheet.Range("B3").Value2 = $excelCode

        for ($sheetIndex = 1; $sheetIndex -le $workbook.Sheets.Count; $sheetIndex++) {
            $sheet = $workbook.Sheets.Item($sheetIndex)
            $sheet.Activate() | Out-Null
            for ($linkIndex = 1; $linkIndex -le $sheet.Hyperlinks.Count; $linkIndex++) {
                $hyperlink = $sheet.Hyperlinks.Item($linkIndex)
                $screenTip = [string]$hyperlink.ScreenTip
                if ($screenTip.ToLowerInvariant() -ne "DataGuide6".ToLowerInvariant()) {
                    continue
                }
                $hyperlink.Range.Select() | Out-Null
                $hyperlink.Follow($false, $true)
                while (-not $excel.Ready) {
                    Start-Sleep -Milliseconds 100
                }
                Start-Sleep -Seconds $WaitSeconds
                $deadline = (Get-Date).AddSeconds(60)
                while ((Get-Date) -lt $deadline) {
                    $refreshedCode = [string]$sheet.Range("B7").Value2
                    if ($refreshedCode -eq $excelCode) {
                        break
                    }
                    Start-Sleep -Seconds 1
                }
            }
        }

        $outputDirectory = Split-Path -Parent $outputPath
        if (-not (Test-Path -LiteralPath $outputDirectory)) {
            New-Item -ItemType Directory -Path $outputDirectory -Force | Out-Null
        }
        $workbook.Save()
        $workbook.SaveCopyAs($outputPath)
    }
}
finally {
    if ($null -ne $workbook) {
        $workbook.Close($false)
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null
    }
    if ($null -ne $excel) {
        if ($createdExcel) {
            $excel.Quit()
        }
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
    }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
"""


def _targets_from_specs(payload: Mapping[str, object]) -> list[RefreshTarget]:
    targets: list[RefreshTarget] = []
    seen: set[str] = set()
    for spec in payload.get("indices", []):
        if not isinstance(spec, Mapping):
            continue
        index_code = str(spec.get("index_code", "")).strip()
        products = spec.get("products", [])
        if not isinstance(products, list):
            continue
        for product in products:
            if not isinstance(product, Mapping):
                continue
            etf_code_raw = str(product.get("etf_code", "")).strip()
            etf_code = _strip_a_prefix(etf_code_raw)
            if not etf_code or etf_code in seen:
                continue
            seen.add(etf_code)
            targets.append(
                RefreshTarget(
                    etf_code=etf_code,
                    excel_code=_excel_code(etf_code),
                    etf_name=str(product.get("etf_name", "")).strip(),
                    index_code=index_code,
                    source="methodology_specs",
                )
            )
    return targets


def _find_ticker_header_row(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)):
        values = [cell.value for cell in row]
        if values and values[0] == "코드" and values[1] == "코드명":
            return row[0].row
    raise ValueError("ticker workbook header row not found")


def _targets_from_rules(payload: Mapping[str, object]) -> list[RefreshTarget]:
    targets: list[RefreshTarget] = []
    seen: set[str] = set()
    for item in payload.get("items", []):
        if not isinstance(item, Mapping):
            continue
        if str(item.get("status", "")).strip() != "downloaded":
            continue
        etf_code = _strip_a_prefix(str(item.get("code", "")).strip())
        if not etf_code or etf_code in seen:
            continue
        seen.add(etf_code)
        targets.append(
            RefreshTarget(
                etf_code=etf_code,
                excel_code=_excel_code(etf_code),
                etf_name=str(item.get("name", "")).strip(),
                index_code="",
                source="rules",
            )
        )
    return targets


def _excel_code(etf_code: str) -> str:
    return etf_code if etf_code.startswith("A") else f"A{etf_code}"


def _strip_a_prefix(value: str) -> str:
    return value[1:] if value.startswith("A") else value


def _date_string(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _float_or_zero(value: object) -> float:
    return float(value or 0.0)


def _safe_name(value: str) -> str:
    return "".join(char if char.isalnum() else "_" for char in value)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Refresh FnGuide/DataGuide ETF holdings workbooks and write parquet.")
    parser.add_argument("--template", default=paths.REFRESH_TEMPLATE_XLSX.as_posix())
    parser.add_argument("--specs", default=paths.FNGUIDE_METHODOLOGY_SPECS_JSON.as_posix())
    parser.add_argument("--rules", default=paths.FNGUIDE_RULES_JSON.as_posix())
    parser.add_argument("--output-dir", default=paths.REFRESHED_HOLDINGS_FILES_DIR.as_posix())
    parser.add_argument("--combined-output", default="")
    parser.add_argument("--refresh-mode", choices=["excel", "parse-only"], default="excel")
    parser.add_argument("--mode", choices=["batch", "single"], default="batch")
    parser.add_argument("--tickers", nargs="*", default=[], help="Optional ETF codes to refresh, with or without A prefix.")
    parser.add_argument("--limit", type=int, default=None, help="Optional maximum number of ETF tickers to refresh.")
    parser.add_argument("--hidden", action="store_true", help="Run Excel hidden. DataGuide6 refresh is more reliable visible.")
    parser.add_argument("--work-dir", default=paths.REFRESH_WORK_DIR.as_posix())
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    specs_path = Path(args.specs)
    rules_path = Path(args.rules)
    template_path = Path(args.template)
    output_dir = Path(args.output_dir)
    index_map = index_code_by_etf_from_specs(specs_path) if specs_path.exists() else {}
    if args.refresh_mode == "parse-only":
        records = extract_sheet_records(template_path, index_code_by_etf=index_map)
        outputs = write_records_by_ticker_parquet(records, output_dir)
    else:
        driver = DataguideExcelRefreshDriver(work_dir=Path(args.work_dir), visible=not bool(args.hidden))
        targets = filter_refresh_targets(
            load_refresh_targets(specs_path=specs_path, rules_path=rules_path),
            tickers=args.tickers,
            limit=args.limit,
        )
        refresh_func = refresh_targets_to_parquet_files_batch if args.mode == "batch" else refresh_targets_to_parquet_files
        outputs = refresh_func(
            targets=targets,
            template_path=template_path,
            output_dir=output_dir,
            index_code_by_etf=index_map,
            driver=driver,
        )
    if args.combined_output:
        combined_records: list[dict[str, object]] = []
        for path in outputs.values():
            combined_records.extend(pd.read_parquet(path).to_dict("records"))
        write_records_parquet(combined_records, Path(args.combined_output))
    print(f"wrote {len(outputs)} parquet file(s) under {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
