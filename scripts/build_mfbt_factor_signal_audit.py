from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backtesting.catalog import DatasetId
from backtesting.data import DataLoader, LoadRequest, ParquetStore
from backtesting.run import BacktestRunner
from backtesting.strategies import build_strategy
from backtesting.strategies.mfbt import _month_end_observations, _quarter_lagged_financials, _score_row, _value_source_period


OUTPUT = Path("reports") / "mfbt_factor_signal_audit_by_factor.xlsx"
DATE_FORMAT = "yyyy-mm-dd"
SCORE_FORMAT = "0"
PERCENT_FORMAT = "0.00%"
ACCOUNTING_FORMAT = '_-* #,##0.00_-;[Red]-* #,##0.00_-;_-* "-"??_-;_-@_-'
NUMBER_FORMAT = "#,##0.00"
FACTOR_SHEET_NAMES = (
    "price_momentum",
    "earnings_momentum",
    "dividend_yield",
    "retail_flow",
    "value",
)
TICKERS = {"A005930": "삼성전자", "A000660": "SK하이닉스"}
START = "2000-01-01"
END = "2026-05-27"


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    strategy = build_strategy("mfbt")
    datasets = list(strategy.datasets)
    if DatasetId.QW_K200_YN not in datasets:
        datasets.append(DatasetId.QW_K200_YN)

    loader = DataLoader(BacktestRunner().catalog, ParquetStore(Path("parquet")))
    market = loader.load(LoadRequest(datasets=datasets, start=START, end=END))
    market.universe = market.frames["k200_yn"].fillna(0).astype(bool)
    bundle = strategy.signal_producer.build(market)
    factors = bundle.meta
    factor_names = list(factors)
    base = factors[factor_names[0]]

    close = market.frames["close"]
    market_cap = _align(market.frames["market_cap"], close)
    op_fwd_12m = _align(market.frames["op_fwd_12m"], close)
    dps_ttm = _align(market.frames["dps_ttm"], close)
    dividend_cash_ttm = _align(market.frames["dividend_cash_ttm"], close)
    retail_flow = _align(market.frames["retail_flow"], close)
    sector = market.frames["sector_big"].reindex(index=close.index, columns=close.columns).ffill()
    free_cash_flow = _align(market.frames["free_cash_flow"], close)
    debt = _align(market.frames["interest_bearing_liability"], close)
    quick_asset = _align(market.frames["quick_asset"], close)
    universe = market.universe.reindex(index=close.index, columns=close.columns).fillna(False).astype(bool)

    calc = _build_calculations(
        close=close,
        market_cap=market_cap,
        op_fwd_12m=op_fwd_12m,
        dps_ttm=dps_ttm,
        dividend_cash=dividend_cash_ttm,
        retail_flow=retail_flow,
        sector=sector,
        free_cash_flow=free_cash_flow,
        debt=debt,
        quick_asset=quick_asset,
        universe=universe,
    )

    non_na_dates = {name: set(frame.dropna(how="all").index) for name, frame in factors.items()}
    common_dates = sorted(set.intersection(*non_na_dates.values()))
    latest_common = common_dates[-1]
    _write_factor_sheet_workbook(
        OUTPUT,
        tickers=tuple(TICKERS),
        signal_date=latest_common,
        market=market,
        factors=factors,
        calc=calc,
        common_dates=common_dates,
    )

    print(OUTPUT.resolve())
    print(f"latest_common={latest_common.date()}")
    print(f"common_dates={len(common_dates)}")


def _align(frame: pd.DataFrame, close: pd.DataFrame) -> pd.DataFrame:
    return frame.reindex(index=close.index, columns=close.columns).astype(float)


def _build_calculations(
    *,
    close: pd.DataFrame,
    market_cap: pd.DataFrame,
    op_fwd_12m: pd.DataFrame,
    dps_ttm: pd.DataFrame,
    dividend_cash: pd.DataFrame,
    retail_flow: pd.DataFrame,
    sector: pd.DataFrame,
    free_cash_flow: pd.DataFrame,
    debt: pd.DataFrame,
    quick_asset: pd.DataFrame,
    universe: pd.DataFrame,
) -> dict[str, object]:
    price_high = close.rolling(252, min_periods=252).max()
    price_ratio = close.divide(price_high)

    monthly_op = _month_end_observations(op_fwd_12m)
    prev_op = monthly_op.shift(1)
    op_denom = prev_op.abs().mask(prev_op.abs().eq(0.0))
    op_growth_raw = monthly_op.sub(prev_op).divide(op_denom)
    op_growth_metric = op_growth_raw.where(monthly_op.notna() & prev_op.notna())
    op_low_extreme = monthly_op.lt(100_000_000_000.0) & op_growth_metric.gt(0.50)
    op_growth_metric = op_growth_metric.mask(op_low_extreme, 0.0)

    monthly_close = _month_end_observations(close)
    monthly_dps = _month_end_observations(dps_ttm).reindex(index=monthly_close.index, columns=monthly_close.columns)
    dividend_yield = monthly_dps.divide(monthly_close.where(monthly_close.gt(0.0)))
    monthly_dividend_cash_ttm = _month_end_observations(dividend_cash).copy()
    monthly_dividend_cash_ttm.index = monthly_dividend_cash_ttm.index.to_period("M")
    monthly_dividend_cash_ttm = monthly_dividend_cash_ttm.loc[~monthly_dividend_cash_ttm.index.duplicated(keep="last")]

    monthly_retail = _month_end_observations(retail_flow.rolling(252, min_periods=252).sum())
    monthly_sector = _month_end_observations(sector).reindex(index=monthly_retail.index, columns=monthly_retail.columns)
    monthly_universe = universe.reindex(index=monthly_retail.index, columns=monthly_retail.columns).fillna(False).astype(bool)
    retail_sector_table = _retail_sector_table(monthly_retail, monthly_sector, monthly_universe)

    signal_dates = pd.DatetimeIndex(monthly_close.index)
    lagged_fcf = _quarter_lagged_financials(free_cash_flow, signal_dates)
    lagged_debt = _quarter_lagged_financials(debt, signal_dates)
    lagged_quick = _quarter_lagged_financials(quick_asset, signal_dates)
    monthly_mktcap = _month_end_observations(market_cap)
    value_tev = monthly_mktcap.add(lagged_debt).sub(lagged_quick)
    required = monthly_mktcap.notna() & lagged_fcf.notna() & lagged_debt.notna() & lagged_quick.notna()
    value_metric = lagged_fcf.divide(value_tev).where(required).mask(required & value_tev.le(0.0), float("-inf"))

    quarter_source_frame = _quarter_source_frame(free_cash_flow)

    return {
        "close": close,
        "sector": sector,
        "free_cash_flow": free_cash_flow,
        "debt": debt,
        "quick_asset": quick_asset,
        "price_high": price_high,
        "price_ratio": price_ratio,
        "monthly_op": monthly_op,
        "prev_op": prev_op,
        "op_growth_raw": op_growth_raw,
        "op_growth_metric": op_growth_metric,
        "op_low_extreme": op_low_extreme,
        "monthly_close": monthly_close,
        "monthly_dps": monthly_dps,
        "dividend_yield": dividend_yield,
        "dividend_cash_ttm": monthly_dividend_cash_ttm,
        "monthly_retail": monthly_retail,
        "retail_sector_table": retail_sector_table,
        "monthly_mktcap": monthly_mktcap,
        "lagged_fcf": lagged_fcf,
        "lagged_debt": lagged_debt,
        "lagged_quick": lagged_quick,
        "value_tev": value_tev,
        "value_metric": value_metric,
        "quarter_source_frame": quarter_source_frame,
    }


def _retail_sector_table(
    monthly_retail: pd.DataFrame,
    monthly_sector: pd.DataFrame,
    monthly_universe: pd.DataFrame,
) -> pd.DataFrame:
    rows = []
    for date in monthly_retail.index:
        flows = monthly_retail.loc[date]
        sectors = monthly_sector.loc[date]
        valid = monthly_universe.loc[date] & flows.notna() & sectors.notna()
        if not valid.any():
            continue
        sector_avg = flows.loc[valid].groupby(sectors.loc[valid]).mean()
        sector_score = _score_row(-sector_avg, 5)
        for sector_code, avg_flow in sector_avg.items():
            rows.append(
                {
                    "date": date,
                    "sector": sector_code,
                    "sector_avg_252d_retail_flow": avg_flow,
                    "sector_score": sector_score.loc[sector_code],
                }
            )
    frame = pd.DataFrame(rows)
    if frame.empty:
        return frame
    return frame.set_index(["date", "sector"]).sort_index()


def _quarter_source_frame(free_cash_flow: pd.DataFrame) -> pd.DataFrame:
    rows = []
    monthly = _month_end_observations(free_cash_flow)
    if monthly.empty:
        return pd.DataFrame(columns=["원천_재무월", "최초_반영_시그널월", "마지막_반영_시그널월", "lag_months", "_source_period", "_signal_period"])

    signal_periods = pd.period_range(monthly.index.min().to_period("M"), monthly.index.max().to_period("M"), freq="M")
    for signal_period in signal_periods:
        source_period = _value_source_period(signal_period)
        if source_period.month not in (3, 5, 8, 11):
            continue
        rows.append(
            {
                "원천_재무월": str(source_period),
                "최초_반영_시그널월": str(signal_period),
                "마지막_반영_시그널월": str(signal_period),
                "lag_months": signal_period.month - source_period.month if signal_period.year == source_period.year else signal_period.month + 12 - source_period.month,
                "_source_period": source_period,
                "_signal_period": signal_period,
            }
        )
    frame = pd.DataFrame(rows).drop_duplicates(subset=["원천_재무월", "최초_반영_시그널월"])
    if frame.empty:
        return frame
    grouped = frame.groupby("_source_period", as_index=False).agg(
        원천_재무월=("원천_재무월", "first"),
        최초_반영_시그널월=("최초_반영_시그널월", "first"),
        마지막_반영_시그널월=("마지막_반영_시그널월", "last"),
        lag_months=("lag_months", "first"),
        _signal_period=("_signal_period", "first"),
    )
    frame = grouped[["원천_재무월", "최초_반영_시그널월", "마지막_반영_시그널월", "lag_months", "_source_period", "_signal_period"]]
    return frame


def _methodology() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"구분": "결측 처리", "설명": "raw metric을 계산할 수 없는 경우는 NaN으로 유지합니다. 계산 가능한 값이 quantile 최하위이거나 조건을 통과하지 못한 경우만 실제 0점입니다."},
            {"구분": "최종 mfbt mask", "설명": "최종 mfbt factor metadata는 모든 필수 팩터가 산출 가능한 날짜 교집합만 남깁니다. ticker별 결측은 팩터별 raw availability를 그대로 유지합니다."},
            {"구분": "기준 universe", "설명": "모든 팩터 점수는 active universe, 현재는 KOSPI200 구성종목 기준으로 산정합니다. universe 밖 종목은 NaN입니다."},
            {"구분": "날짜 기준", "설명": "모든 팩터 output은 month-end observation입니다. daily row 중 월말 observation이 아닌 날짜는 NaN입니다."},
            {"구분": "price_momentum", "설명": "현재가 / 최근 252거래일 종가 최고가 > 0.8이면 1, 아니면 0입니다. 252거래일 종가 이력이 없으면 계산 불가이므로 NaN입니다."},
            {"구분": "earnings_momentum", "설명": "(월말 12MF 영업이익 - 전월말 12MF 영업이익) / abs(전월말 12MF 영업이익)을 K200 내부 quantile 0~4로 점수화합니다. OP가 1000억 미만이고 성장률이 50% 초과면 성장률을 0으로 리셋합니다."},
            {"구분": "dividend_yield", "설명": "dps_ttm / close를 K200 내부 quantile 0~4로 점수화하고, 신호월 기준 24개월 전, 12개월 전, 현재월의 dividend_cash_ttm이 연속 증가하면 +1을 더합니다."},
            {"구분": "retail_flow", "설명": "종목별 252일 개인 순매수대금을 구한 뒤 K200 구성종목 기준 섹터 평균을 계산합니다. 개인 순매도가 큰 섹터가 높은 점수입니다."},
            {"구분": "value", "설명": "free_cash_flow / TEV를 K200 내부 quantile 0~4로 점수화합니다. TEV = market_cap + interest_bearing_liability - quick_asset입니다."},
            {"구분": "value lag 예시", "설명": "4~5월 signal은 3월말 데이터, 6~8월은 5월말 데이터, 9~11월은 8월말 데이터, 12~3월은 11월말 데이터를 사용합니다."},
        ]
    )


def _alignment_summary(factors: dict[str, pd.DataFrame], base: pd.DataFrame, latest_common: pd.Timestamp) -> pd.DataFrame:
    rows = []
    for name, frame in factors.items():
        dates = frame.dropna(how="all").index
        rows.append(
            {
                "팩터": name,
                "row수": frame.shape[0],
                "ticker수": frame.shape[1],
                "price_momentum과_index_동일": frame.index.equals(base.index),
                "price_momentum과_ticker_동일": frame.columns.equals(base.columns),
                "signal_date수": len(dates),
                "첫_signal_date": dates[0].date().isoformat(),
                "마지막_signal_date": dates[-1].date().isoformat(),
                f"{latest_common.date()}_non_na_ticker수": int(frame.loc[latest_common].notna().sum()),
            }
        )
    return pd.DataFrame(rows)


def _date_alignment(
    factors: dict[str, pd.DataFrame],
    factor_names: list[str],
    non_na_dates: dict[str, set[pd.Timestamp]],
    union_dates: list[pd.Timestamp],
) -> pd.DataFrame:
    rows = []
    for date in union_dates:
        row = {"date": date.date().isoformat()}
        for name in factor_names:
            row[f"{name}_signal_존재"] = date in non_na_dates[name]
            row[f"{name}_non_na_ticker수"] = int(factors[name].loc[date].notna().sum()) if date in factors[name].index else 0
        row["모든_factor_signal_존재"] = all(date in non_na_dates[name] for name in factor_names)
        rows.append(row)
    return pd.DataFrame(rows)


def _ticker_alignment(
    factors: dict[str, pd.DataFrame],
    base: pd.DataFrame,
    factor_names: list[str],
    latest_common: pd.Timestamp,
) -> pd.DataFrame:
    frame = pd.DataFrame({"ticker": base.columns})
    for name in factor_names:
        frame[f"{name}_{latest_common.date()}_non_na"] = factors[name].loc[latest_common].notna().to_numpy()
    non_na_cols = [column for column in frame.columns if column.endswith("_non_na")]
    frame["최신공통일_모든팩터_non_na"] = frame[non_na_cols].all(axis=1)
    return frame


def _ticker_detail(ticker: str, market, factors: dict[str, pd.DataFrame], base: pd.DataFrame, calc: dict[str, object]) -> pd.DataFrame:
    close = calc["close"]
    sector = calc["sector"]
    month_dates = base.index[~base.index.to_period("M").duplicated(keep="last")]
    rows = []
    for date in month_dates:
        current_sector = sector.loc[date, ticker] if date in sector.index and ticker in sector.columns else pd.NA
        retail_avg, retail_score = _retail_values(calc["retail_sector_table"], date, current_sector)
        rows.append(
            {
                "날짜": date,
                "티커": ticker,
                "종목명": TICKERS[ticker],
                "K200여부": bool(market.universe.loc[date, ticker]) if date in market.universe.index else False,
                "섹터": current_sector,
                "종가": close.loc[date, ticker],
                "252일_종가고가": calc["price_high"].loc[date, ticker],
                "가격모멘텀_비율_값": calc["price_ratio"].loc[date, ticker],
                "가격모멘텀_비율_엑셀수식": None,
                "price_momentum_score": factors["price_momentum"].loc[date, ticker],
                "12MF_OP": _at(calc["monthly_op"], date, ticker),
                "전월_12MF_OP": _at(calc["prev_op"], date, ticker),
                "이익성장률_원값": _at(calc["op_growth_raw"], date, ticker),
                "이익성장률_엑셀수식": None,
                "저OP_고성장_리셋여부": bool(_at(calc["op_low_extreme"], date, ticker, default=False)),
                "이익성장률_스코어용": _at(calc["op_growth_metric"], date, ticker),
                "earnings_momentum_score": factors["earnings_momentum"].loc[date, ticker],
                "DPS_TTM": _at(calc["monthly_dps"], date, ticker),
                "배당계산_종가": _at(calc["monthly_close"], date, ticker),
                "배당수익률_값": _at(calc["dividend_yield"], date, ticker),
                "배당수익률_엑셀수식": None,
                **_dividend_bonus_parts(calc["dividend_cash_ttm"], date, ticker),
                "배당증가_bonus_엑셀수식": None,
                "dividend_yield_score": factors["dividend_yield"].loc[date, ticker],
                "개인수급_252일누적": _at(calc["monthly_retail"], date, ticker),
                "섹터평균_개인수급_252일누적": retail_avg,
                "섹터수급_score": retail_score,
                "retail_flow_score": factors["retail_flow"].loc[date, ticker],
                "value_원천재무월": _source_period_for_signal(calc["quarter_source_frame"], date),
                "시가총액": _at(calc["monthly_mktcap"], date, ticker),
                "lagged_FCF": _at(calc["lagged_fcf"], date, ticker),
                "lagged_이자발생부채": _at(calc["lagged_debt"], date, ticker),
                "lagged_당좌자산": _at(calc["lagged_quick"], date, ticker),
                "TEV_값": _at(calc["value_tev"], date, ticker),
                "TEV_엑셀수식": None,
                "FCF_over_TEV_값": _at(calc["value_metric"], date, ticker),
                "FCF_over_TEV_엑셀수식": None,
                "value_score": factors["value"].loc[date, ticker],
            }
        )
    return pd.DataFrame(rows)


def _at(frame: pd.DataFrame, date: pd.Timestamp, ticker: str, default=pd.NA):
    if date not in frame.index or ticker not in frame.columns:
        return default
    return frame.loc[date, ticker]


def _retail_values(retail_sector_table: pd.DataFrame, date: pd.Timestamp, sector) -> tuple[object, object]:
    if pd.isna(sector) or retail_sector_table.empty or (date, sector) not in retail_sector_table.index:
        return pd.NA, pd.NA
    row = retail_sector_table.loc[(date, sector)]
    return row["sector_avg_252d_retail_flow"], row["sector_score"]


def _source_period_for_signal(quarter_source_frame: pd.DataFrame, signal_date: pd.Timestamp) -> str | None:
    signal_period = signal_date.to_period("M")
    candidates = quarter_source_frame.loc[quarter_source_frame["_signal_period"] <= signal_period]
    if candidates.empty:
        return None
    return str(candidates.iloc[-1]["_source_period"])


def _dividend_bonus_parts(monthly_dividend_cash_ttm: pd.DataFrame, date: pd.Timestamp, ticker: str) -> dict[str, object]:
    signal_period = date.to_period("M")
    periods = [signal_period - 24, signal_period - 12, signal_period]
    values = [
        monthly_dividend_cash_ttm.loc[period, ticker]
        if period in monthly_dividend_cash_ttm.index and ticker in monthly_dividend_cash_ttm.columns
        else pd.NA
        for period in periods
    ]
    return {
        "배당증가_연도1": str(periods[0]),
        "배당증가_현금배당1": values[0],
        "배당증가_연도2": str(periods[1]),
        "배당증가_현금배당2": values[1],
        "배당증가_연도3": str(periods[2]),
        "배당증가_현금배당3": values[2],
    }


def _price_window_sheet(
    tickers,
    signal_date: pd.Timestamp,
    calc: dict[str, object],
    factors: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    close = calc["close"]
    rows = []
    for ticker in tickers:
        window = close.loc[:signal_date, ticker].dropna().tail(252)
        final_high = window.max()
        running_high = window.expanding(min_periods=1).max()
        for raw_date, raw_close in window.items():
            rows.append(
                {
                    "종목명": TICKERS[ticker],
                    "티커": ticker,
                    "raw_date": raw_date,
                    "raw_close": raw_close,
                    "누적_종가고가": running_high.loc[raw_date],
                    "최종고가일": bool(raw_close == final_high),
                }
            )
    return pd.DataFrame(rows)


def _one_page_sheet(
    tickers,
    signal_date: pd.Timestamp,
    market,
    factors: dict[str, pd.DataFrame],
    calc: dict[str, object],
) -> pd.DataFrame:
    rows = []
    for ticker in tickers:
        sector = _at(calc["sector"], signal_date, ticker)
        retail_avg, _ = _retail_values(calc["retail_sector_table"], signal_date, sector)
        source_row = _source_row_for_signal(calc["quarter_source_frame"], signal_date)
        source_period = source_row["_source_period"] if source_row is not None else None
        rows.append(
            {
                "신호일": signal_date,
                "종목명": TICKERS[ticker],
                "티커": ticker,
                "K200여부": bool(market.universe.loc[signal_date, ticker]) if signal_date in market.universe.index else False,
                "가격_종가": _at(calc["close"], signal_date, ticker),
                "가격_252일고가": _at(calc["price_high"], signal_date, ticker),
                "가격_비율_수식": None,
                "가격_비율_값": _at(calc["price_ratio"], signal_date, ticker),
                "price_score": _at(factors["price_momentum"], signal_date, ticker),
                "이익_12MF_OP": _at(calc["monthly_op"], signal_date, ticker),
                "이익_전월_12MF_OP": _at(calc["prev_op"], signal_date, ticker),
                "이익성장률_수식": None,
                "이익성장률_값": _at(calc["op_growth_metric"], signal_date, ticker),
                "earnings_score": _at(factors["earnings_momentum"], signal_date, ticker),
                "배당_DPS_TTM": _at(calc["monthly_dps"], signal_date, ticker),
                "배당_종가": _at(calc["monthly_close"], signal_date, ticker),
                "배당수익률_수식": None,
                "배당수익률_값": _at(calc["dividend_yield"], signal_date, ticker),
                "dividend_score": _at(factors["dividend_yield"], signal_date, ticker),
                "수급_섹터": sector,
                "수급_섹터평균252일": retail_avg,
                "retail_score": _at(factors["retail_flow"], signal_date, ticker),
                "가치_원천재무월": str(source_period) if source_period is not None else pd.NA,
                "가치_FCF": _period_value(calc["free_cash_flow"], source_period, ticker),
                "가치_시가총액": _at(calc["monthly_mktcap"], signal_date, ticker),
                "가치_이자발생부채": _period_value(calc["debt"], source_period, ticker),
                "가치_당좌자산": _period_value(calc["quick_asset"], source_period, ticker),
                "가치_TEV_수식": None,
                "가치_TEV_값": _at(calc["value_tev"], signal_date, ticker),
                "가치_FCF/TEV_수식": None,
                "가치_FCF/TEV_값": _at(calc["value_metric"], signal_date, ticker),
                "value_score": _at(factors["value"], signal_date, ticker),
            }
        )
    return pd.DataFrame(rows)


def _price_summary_sheet(
    tickers,
    signal_date: pd.Timestamp,
    calc: dict[str, object],
    factors: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    close = calc["close"]
    rows = []
    for ticker in tickers:
        window = close.loc[:signal_date, ticker].dropna().tail(252)
        final_high = window.max()
        high_dates = window.loc[window.eq(final_high)].index
        signal_close = close.loc[signal_date, ticker]
        rows.append(
            {
                "신호일": signal_date,
                "종목명": TICKERS[ticker],
                "티커": ticker,
                "신호일_종가": signal_close,
                "252일_종가고가_값": final_high,
                "252일_종가고가_엑셀수식": None,
                "고가발생일": high_dates[-1] if len(high_dates) else pd.NA,
                "종가/고가_값": signal_close / final_high if final_high else pd.NA,
                "종가/고가_엑셀수식": None,
                "price_momentum_score": _at(factors["price_momentum"], signal_date, ticker),
                "판정": "통과" if _at(factors["price_momentum"], signal_date, ticker) == 1.0 else "미통과",
            }
        )
    return pd.DataFrame(rows)


def _value_lag_summary_sheet(
    tickers,
    signal_date: pd.Timestamp,
    calc: dict[str, object],
    factors: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    source_row = _source_row_for_signal(calc["quarter_source_frame"], signal_date)
    source_period = source_row["_source_period"] if source_row is not None else None
    rows = []
    for ticker in tickers:
        rows.append(
            {
                "신호일": signal_date,
                "종목명": TICKERS[ticker],
                "티커": ticker,
                "원천_재무월": str(source_period) if source_period is not None else pd.NA,
                "lag_months": source_row["lag_months"] if source_row is not None else pd.NA,
                "원천_FCF": _period_value(calc["free_cash_flow"], source_period, ticker),
                "원천_이자발생부채": _period_value(calc["debt"], source_period, ticker),
                "원천_당좌자산": _period_value(calc["quick_asset"], source_period, ticker),
                "시가총액": _at(calc["monthly_mktcap"], signal_date, ticker),
                "TEV_값": _at(calc["value_tev"], signal_date, ticker),
                "TEV_엑셀수식": None,
                "FCF_over_TEV_값": _at(calc["value_metric"], signal_date, ticker),
                "FCF_over_TEV_엑셀수식": None,
                "value_score": _at(factors["value"], signal_date, ticker),
            }
        )
    return pd.DataFrame(rows)


def _value_lag_trace_sheet(
    tickers,
    base: pd.DataFrame,
    calc: dict[str, object],
    factors: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    month_dates = base.index[~base.index.to_period("M").duplicated(keep="last")]
    rows = []
    for date in month_dates:
        source_row = _source_row_for_signal(calc["quarter_source_frame"], date)
        source_period = source_row["_source_period"] if source_row is not None else None
        for ticker in tickers:
            rows.append(
                {
                    "신호일": date,
                    "티커": ticker,
                    "종목명": TICKERS[ticker],
                    "원천_재무월": str(source_period) if source_period is not None else pd.NA,
                    "lag_months": source_row["lag_months"] if source_row is not None else pd.NA,
                    "원천_FCF": _period_value(calc["free_cash_flow"], source_period, ticker),
                    "원천_이자발생부채": _period_value(calc["debt"], source_period, ticker),
                    "원천_당좌자산": _period_value(calc["quick_asset"], source_period, ticker),
                    "lagged_FCF": _at(calc["lagged_fcf"], date, ticker),
                    "lagged_이자발생부채": _at(calc["lagged_debt"], date, ticker),
                    "lagged_당좌자산": _at(calc["lagged_quick"], date, ticker),
                    "시가총액": _at(calc["monthly_mktcap"], date, ticker),
                    "TEV_값": _at(calc["value_tev"], date, ticker),
                    "TEV_엑셀수식": None,
                    "FCF_over_TEV_값": _at(calc["value_metric"], date, ticker),
                    "FCF_over_TEV_엑셀수식": None,
                    "value_score": factors["value"].loc[date, ticker],
                }
            )
    return pd.DataFrame(rows)


def _source_row_for_signal(quarter_source_frame: pd.DataFrame, signal_date: pd.Timestamp):
    signal_period = signal_date.to_period("M")
    candidates = quarter_source_frame.loc[quarter_source_frame["_signal_period"] <= signal_period]
    if candidates.empty:
        return None
    return candidates.iloc[-1]


def _period_value(frame: pd.DataFrame, period: pd.Period | None, ticker: str):
    if period is None or ticker not in frame.columns:
        return pd.NA
    rows = frame.loc[frame.index.to_period("M") == period]
    if rows.empty:
        return pd.NA
    return rows.iloc[-1][ticker]


def _write_factor_sheet_workbook(
    path: Path,
    *,
    tickers: tuple[str, ...],
    signal_date: pd.Timestamp,
    market,
    factors: dict[str, pd.DataFrame],
    calc: dict[str, object],
    common_dates: list[pd.Timestamp],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    formula_fill = PatternFill("solid", fgColor="E2F0D9")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    raw_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def setup(sheet_name: str, title: str):
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=15)
        ws["A2"] = "신호일"
        ws["B2"] = signal_date
        ws["C2"] = "공통 signal_date 수"
        ws["D2"] = len(common_dates)
        return ws

    def write_row(ws, row: int, values: list[object], *, fill=None, bold: bool = False) -> int:
        for column, value in enumerate(values, start=1):
            target = ws.cell(row, column, value)
            target.border = border
            target.alignment = Alignment(vertical="center", wrap_text=True)
            if fill:
                target.fill = fill
            if bold:
                target.font = Font(bold=True, color="FFFFFF" if fill == header_fill else "000000")
        return row + 1

    def finish(ws, max_col: int) -> None:
        ws.freeze_panes = "A5"
        for row in ws.iter_rows():
            for current in row:
                if isinstance(current.value, str) and current.value.startswith("="):
                    current.fill = formula_fill
        _apply_number_formats(ws)
        for index in range(1, max_col + 1):
            width = 16 if index <= max_col else 12
            ws.column_dimensions[get_column_letter(index)].width = width
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 12

    _write_price_factor_sheet(setup("price_momentum", "price_momentum 계산 과정"), write_row, finish, tickers, signal_date, factors, calc, header_fill, raw_fill)
    _write_earnings_factor_sheet(setup("earnings_momentum", "earnings_momentum 계산 과정"), write_row, finish, tickers, signal_date, factors, calc, header_fill, raw_fill)
    _write_dividend_factor_sheet(setup("dividend_yield", "dividend_yield 계산 과정"), write_row, finish, tickers, signal_date, factors, calc, header_fill, raw_fill)
    _write_retail_factor_sheet(setup("retail_flow", "retail_flow 계산 과정"), write_row, finish, tickers, signal_date, factors, calc, header_fill, raw_fill)
    _write_value_factor_sheet(setup("value", "value 계산 과정"), write_row, finish, tickers, signal_date, factors, calc, header_fill, raw_fill)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _apply_number_formats(ws) -> None:
    for row in ws.iter_rows():
        for current in row:
            current.number_format = _number_format_for_cell(ws, current.row, current.column)


def _number_format_for_cell(ws, row: int, column: int) -> str:
    value = ws.cell(row, column).value
    if isinstance(value, pd.Timestamp):
        return DATE_FORMAT
    if hasattr(value, "to_pydatetime"):
        return DATE_FORMAT
    header = _nearest_header(ws, row, column)
    if header is None:
        return "General"
    if _is_date_header(header):
        return DATE_FORMAT
    if _is_percent_header(header):
        return PERCENT_FORMAT
    if _is_score_header(header):
        return SCORE_FORMAT
    if _is_accounting_header(header):
        return ACCOUNTING_FORMAT
    if isinstance(value, int | float) and not isinstance(value, bool):
        return NUMBER_FORMAT
    return "General"


def _nearest_header(ws, row: int, column: int) -> str | None:
    for header_row in range(row - 1, 0, -1):
        value = ws.cell(header_row, column).value
        if isinstance(value, str) and value.strip():
            return value
    return None


def _is_date_header(header: str) -> bool:
    normalized = header.lower()
    return normalized in {"date", "raw_date", "신호일"} or normalized.endswith("_date")


def _is_percent_header(header: str) -> bool:
    return any(token in header for token in ("비율", "성장률", "수익률", "FCF/TEV", "FCF_over_TEV", "over_TEV"))


def _is_score_header(header: str) -> bool:
    return any(token in header.lower() for token in ("score", "bonus", "점수"))


def _is_accounting_header(header: str) -> bool:
    return any(
        token in header
        for token in (
            "종가",
            "고가",
            "raw_close",
            "OP",
            "DPS",
            "배당",
            "현금배당",
            "개인수급",
            "평균",
            "시가총액",
            "FCF",
            "부채",
            "자산",
            "TEV",
            "12MF",
        )
    )


def _write_price_factor_sheet(ws, write_row, finish, tickers, signal_date, factors, calc, header_fill, raw_fill) -> None:
    summary_header = ["종목명", "티커", "신호일 종가", "252일 종가고가 수식", "비율 수식", "score 수식", "score 코드", "판정"]
    write_row(ws, 4, summary_header, fill=header_fill, bold=True)
    raw_start = 9
    raw_header_row = raw_start - 1
    write_row(ws, raw_header_row, ["종목명", "티커", "raw_date", "raw_close", "누적고가 수식", "최종고가일 수식"], fill=raw_fill, bold=True)
    row = raw_start
    ranges = {}
    close = calc["close"]
    for ticker in tickers:
        start = row
        for raw_date, raw_close in close.loc[:signal_date, ticker].dropna().tail(252).items():
            write_row(ws, row, [TICKERS[ticker], ticker, raw_date, raw_close, f"=MAX($D${start}:D{row})", None])
            row += 1
        end = row - 1
        for mark_row in range(start, end + 1):
            ws.cell(mark_row, 6, f"=D{mark_row}=MAX($D${start}:$D${end})")
        ranges[ticker] = (start, end)
    for idx, ticker in enumerate(tickers, start=5):
        start, end = ranges[ticker]
        close_value = _at(calc["close"], signal_date, ticker)
        ws.cell(idx, 1, TICKERS[ticker])
        ws.cell(idx, 2, ticker)
        ws.cell(idx, 3, close_value)
        ws.cell(idx, 4, f"=MAX($D${start}:$D${end})")
        ws.cell(idx, 5, f"=IFERROR(C{idx}/D{idx},NA())")
        ws.cell(idx, 6, f"=IF(E{idx}>0.8,1,0)")
        ws.cell(idx, 7, _at(factors["price_momentum"], signal_date, ticker))
        ws.cell(idx, 8, "통과" if _at(factors["price_momentum"], signal_date, ticker) == 1.0 else "미통과")
    finish(ws, 8)


def _write_earnings_factor_sheet(ws, write_row, finish, tickers, signal_date, factors, calc, header_fill, raw_fill) -> None:
    write_row(ws, 4, ["종목명", "티커", "12MF OP", "전월 12MF OP", "성장률 수식", "score 수식", "score 코드"], fill=header_fill, bold=True)
    support_start = 9
    write_row(ws, support_start - 1, ["티커", "12MF OP", "전월 12MF OP", "계산가능", "성장률 수식", "score 수식"], fill=raw_fill, bold=True)
    valid = factors["earnings_momentum"].loc[signal_date].notna()
    row = support_start
    for ticker in valid.index[valid]:
        write_row(ws, row, [ticker, _at(calc["monthly_op"], signal_date, ticker), _at(calc["prev_op"], signal_date, ticker), f"=AND(ISNUMBER(B{row}),ISNUMBER(C{row}),C{row}<>0)", f"=IFERROR(IF(AND(B{row}<100000000000,(B{row}-C{row})/ABS(C{row})>0.5),0,(B{row}-C{row})/ABS(C{row})),NA())", None])
        row += 1
    end = row - 1
    for support_row in range(support_start, end + 1):
        ws.cell(support_row, 6, f"=IFERROR(MIN(4,INT((RANK.AVG(E{support_row},$E${support_start}:$E${end},1)-1)*5/COUNT($E${support_start}:$E${end}))),NA())")
    for idx, ticker in enumerate(tickers, start=5):
        ws.cell(idx, 1, TICKERS[ticker])
        ws.cell(idx, 2, ticker)
        ws.cell(idx, 3, _at(calc["monthly_op"], signal_date, ticker))
        ws.cell(idx, 4, _at(calc["prev_op"], signal_date, ticker))
        ws.cell(idx, 5, f"=IFERROR(IF(AND(C{idx}<100000000000,(C{idx}-D{idx})/ABS(D{idx})>0.5),0,(C{idx}-D{idx})/ABS(D{idx})),NA())")
        ws.cell(idx, 6, f"=IFERROR(MIN(4,INT((RANK.AVG(E{idx},$E${support_start}:$E${end},1)-1)*5/COUNT($E${support_start}:$E${end}))),NA())")
        ws.cell(idx, 7, _at(factors["earnings_momentum"], signal_date, ticker))
    finish(ws, 7)


def _write_dividend_factor_sheet(ws, write_row, finish, tickers, signal_date, factors, calc, header_fill, raw_fill) -> None:
    write_row(ws, 4, ["종목명", "티커", "DPS_TTM", "종가", "배당수익률 수식", "bonus 수식", "score 수식", "score 코드"], fill=header_fill, bold=True)
    support_start = 9
    write_row(ws, support_start - 1, ["티커", "DPS_TTM", "종가", "배당수익률 수식", "연도1배당", "연도2배당", "연도3배당", "bonus 수식", "score 수식"], fill=raw_fill, bold=True)
    valid = factors["dividend_yield"].loc[signal_date].notna()
    row = support_start
    for ticker in valid.index[valid]:
        bonus = _dividend_bonus_parts(calc["dividend_cash_ttm"], signal_date, ticker)
        write_row(ws, row, [ticker, _at(calc["monthly_dps"], signal_date, ticker), _at(calc["monthly_close"], signal_date, ticker), f"=IFERROR(B{row}/C{row},NA())", bonus["배당증가_현금배당1"], bonus["배당증가_현금배당2"], bonus["배당증가_현금배당3"], f"=IF(AND(E{row}<F{row},F{row}<G{row}),1,0)", None])
        row += 1
    end = row - 1
    for support_row in range(support_start, end + 1):
        ws.cell(support_row, 9, f"=IFERROR(MIN(4,INT((RANK.AVG(D{support_row},$D${support_start}:$D${end},1)-1)*5/COUNT($D${support_start}:$D${end})))+H{support_row},NA())")
    for idx, ticker in enumerate(tickers, start=5):
        bonus = _dividend_bonus_parts(calc["dividend_cash_ttm"], signal_date, ticker)
        ws.cell(idx, 1, TICKERS[ticker])
        ws.cell(idx, 2, ticker)
        ws.cell(idx, 3, _at(calc["monthly_dps"], signal_date, ticker))
        ws.cell(idx, 4, _at(calc["monthly_close"], signal_date, ticker))
        ws.cell(idx, 5, f"=IFERROR(C{idx}/D{idx},NA())")
        ws.cell(idx, 6, f"=IF(AND({bonus['배당증가_현금배당1']}<{bonus['배당증가_현금배당2']},{bonus['배당증가_현금배당2']}<{bonus['배당증가_현금배당3']}),1,0)")
        ws.cell(idx, 7, f"=IFERROR(MIN(4,INT((RANK.AVG(E{idx},$D${support_start}:$D${end},1)-1)*5/COUNT($D${support_start}:$D${end})))+F{idx},NA())")
        ws.cell(idx, 8, _at(factors["dividend_yield"], signal_date, ticker))
    finish(ws, 9)


def _write_retail_factor_sheet(ws, write_row, finish, tickers, signal_date, factors, calc, header_fill, raw_fill) -> None:
    write_row(ws, 4, ["종목명", "티커", "섹터", "섹터평균 252일 개인수급", "score 수식", "score 코드"], fill=header_fill, bold=True)
    support_start = 9
    write_row(ws, support_start - 1, ["섹터", "섹터평균 252일 개인수급", "점수용 metric(-평균)", "섹터점수 수식"], fill=raw_fill, bold=True)
    table = calc["retail_sector_table"]
    sector_rows = {}
    row = support_start
    if not table.empty and signal_date in table.index.get_level_values(0):
        for sector, values in table.loc[signal_date].iterrows():
            sector_rows[sector] = row
            write_row(ws, row, [sector, values["sector_avg_252d_retail_flow"], f"=-B{row}", None])
            row += 1
    end = row - 1
    for support_row in range(support_start, end + 1):
        ws.cell(support_row, 4, f"=IFERROR(MIN(4,INT((RANK.AVG(C{support_row},$C${support_start}:$C${end},1)-1)*5/COUNT($C${support_start}:$C${end}))),NA())")
    for idx, ticker in enumerate(tickers, start=5):
        sector = _at(calc["sector"], signal_date, ticker)
        retail_avg, _ = _retail_values(calc["retail_sector_table"], signal_date, sector)
        sector_row = sector_rows.get(sector)
        ws.cell(idx, 1, TICKERS[ticker])
        ws.cell(idx, 2, ticker)
        ws.cell(idx, 3, sector)
        ws.cell(idx, 4, retail_avg)
        ws.cell(idx, 5, f"=IFERROR($D${sector_row},NA())" if sector_row else "=NA()")
        ws.cell(idx, 6, _at(factors["retail_flow"], signal_date, ticker))
    finish(ws, 6)


def _write_value_factor_sheet(ws, write_row, finish, tickers, signal_date, factors, calc, header_fill, raw_fill) -> None:
    source_row = _source_row_for_signal(calc["quarter_source_frame"], signal_date)
    source_period = source_row["_source_period"] if source_row is not None else None
    write_row(ws, 4, ["종목명", "티커", "원천재무월", "FCF", "시가총액", "이자발생부채", "당좌자산", "TEV 수식", "FCF/TEV 수식", "score 수식", "score 코드"], fill=header_fill, bold=True)
    support_start = 9
    write_row(ws, support_start - 1, ["티커", "원천재무월", "FCF", "시가총액", "이자발생부채", "당좌자산", "TEV 수식", "FCF/TEV 수식", "score 수식"], fill=raw_fill, bold=True)
    valid = factors["value"].loc[signal_date].notna()
    row = support_start
    for ticker in valid.index[valid]:
        write_row(ws, row, [ticker, str(source_period), _period_value(calc["free_cash_flow"], source_period, ticker), _at(calc["monthly_mktcap"], signal_date, ticker), _period_value(calc["debt"], source_period, ticker), _period_value(calc["quick_asset"], source_period, ticker), f"=D{row}+E{row}-F{row}", f"=IFERROR(IF(G{row}<=0,-1E+99,C{row}/G{row}),NA())", None])
        row += 1
    end = row - 1
    for support_row in range(support_start, end + 1):
        ws.cell(support_row, 9, f"=IFERROR(MIN(4,INT((RANK.AVG(H{support_row},$H${support_start}:$H${end},1)-1)*5/COUNT($H${support_start}:$H${end}))),NA())")
    for idx, ticker in enumerate(tickers, start=5):
        ws.cell(idx, 1, TICKERS[ticker])
        ws.cell(idx, 2, ticker)
        ws.cell(idx, 3, str(source_period))
        ws.cell(idx, 4, _period_value(calc["free_cash_flow"], source_period, ticker))
        ws.cell(idx, 5, _at(calc["monthly_mktcap"], signal_date, ticker))
        ws.cell(idx, 6, _period_value(calc["debt"], source_period, ticker))
        ws.cell(idx, 7, _period_value(calc["quick_asset"], source_period, ticker))
        ws.cell(idx, 8, f"=E{idx}+F{idx}-G{idx}")
        ws.cell(idx, 9, f"=IFERROR(IF(H{idx}<=0,-1E+99,D{idx}/H{idx}),NA())")
        ws.cell(idx, 10, f"=IFERROR(MIN(4,INT((RANK.AVG(I{idx},$H${support_start}:$H${end},1)-1)*5/COUNT($H${support_start}:$H${end}))),NA())")
        ws.cell(idx, 11, _at(factors["value"], signal_date, ticker))
    finish(ws, 11)


def _write_single_sheet_workbook(
    path: Path,
    *,
    tickers: tuple[str, ...],
    signal_date: pd.Timestamp,
    market,
    factors: dict[str, pd.DataFrame],
    calc: dict[str, object],
    common_dates: list[pd.Timestamp],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "mfbt_calc_KR"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    subheader_fill = PatternFill("solid", fgColor="D9EAF7")
    formula_fill = PatternFill("solid", fgColor="E2F0D9")
    raw_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(row: int, column: int, value=None, *, bold: bool = False, fill=None):
        target = ws.cell(row, column, value)
        target.border = border
        target.alignment = Alignment(vertical="center", wrap_text=True)
        if bold:
            target.font = Font(bold=True)
        if fill is not None:
            target.fill = fill
        return target

    def section(row: int, title: str) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)
        target = ws.cell(row, 1, title)
        target.font = Font(bold=True, size=12, color="FFFFFF")
        target.fill = header_fill
        target.alignment = Alignment(vertical="center")
        return row + 1

    def write_header(row: int, headers: list[str], *, fill=header_fill) -> int:
        for column, value in enumerate(headers, start=1):
            c = cell(row, column, value, bold=True, fill=fill)
            if fill == header_fill:
                c.font = Font(color="FFFFFF", bold=True)
        return row + 1

    ws["A1"] = "MFBT 계산 검산표"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "신호일"
    ws["B2"] = signal_date
    ws["C2"] = "공통 signal_date 수"
    ws["D2"] = len(common_dates)
    ws["E2"] = "주의"
    ws["F2"] = "날짜는 모든 팩터 공통 월만 사용하고, ticker 결측은 팩터별 raw availability 그대로 유지합니다."

    summary_header = [
        "종목명",
        "티커",
        "팩터",
        "원천/범위",
        "입력1",
        "입력1값",
        "입력2",
        "입력2값/수식",
        "입력3",
        "입력3값",
        "계산수식(엑셀)",
        "계산값(코드)",
        "점수수식(엑셀)",
        "점수(코드)",
        "설명",
        "시가총액(value)",
        "이자발생부채(value)",
        "당좌자산(value)",
    ]
    summary_title_row = 4
    summary_header_row = section(summary_title_row, "1. 한눈에 보는 팩터별 계산 과정과 결과")
    summary_start = write_header(summary_header_row, summary_header)
    summary_rows: dict[tuple[str, str], int] = {}
    row = summary_start
    for ticker in tickers:
        for factor_name in ("price_momentum", "earnings_momentum", "dividend_yield", "retail_flow", "value"):
            summary_rows[(ticker, factor_name)] = row
            cell(row, 1, TICKERS[ticker])
            cell(row, 2, ticker)
            cell(row, 3, factor_name)
            row += 1
    support_row = row + 2

    refs: dict[str, object] = {}
    support_row = _write_price_support(ws, support_row, tickers, signal_date, calc, section, write_header, cell, raw_fill, refs)
    support_row = _write_earnings_support(ws, support_row + 1, signal_date, factors, calc, section, write_header, cell, raw_fill, refs)
    support_row = _write_dividend_support(ws, support_row + 1, signal_date, factors, calc, section, write_header, cell, raw_fill, refs)
    support_row = _write_retail_support(ws, support_row + 1, signal_date, calc, section, write_header, cell, raw_fill, refs)
    _write_value_support(ws, support_row + 1, signal_date, factors, calc, section, write_header, cell, raw_fill, formula_fill, refs)

    _fill_summary_rows(ws, summary_rows, tickers, signal_date, market, factors, calc, refs, formula_fill)

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{summary_header_row}:R{summary_start + len(summary_rows) - 1}"
    widths = {
        "A": 14,
        "B": 12,
        "C": 22,
        "D": 30,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 22,
        "I": 18,
        "J": 18,
        "K": 34,
        "L": 18,
        "M": 40,
        "N": 14,
        "O": 42,
        "P": 18,
        "Q": 18,
        "R": 18,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for cells in ws.iter_rows():
        for current in cells:
            if current.value is not None and isinstance(current.value, str) and current.value.startswith("="):
                current.fill = formula_fill
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_price_support(ws, row, tickers, signal_date, calc, section, write_header, cell, raw_fill, refs):
    row = section(row, "2. price_momentum 원천: 신호일 기준 직전 252개 raw close")
    row = write_header(row, ["종목명", "티커", "raw_date", "raw_close", "누적_종가고가", "최종고가일"], fill=raw_fill)
    price_ranges = {}
    close = calc["close"]
    for ticker in tickers:
        start = row
        window = close.loc[:signal_date, ticker].dropna().tail(252)
        for raw_date, raw_close in window.items():
            cell(row, 1, TICKERS[ticker])
            cell(row, 2, ticker)
            cell(row, 3, raw_date)
            cell(row, 4, raw_close)
            cell(row, 5, f"=MAX($D${start}:D{row})")
            row += 1
        end = row - 1
        for mark_row in range(start, end + 1):
            cell(mark_row, 6, f"=D{mark_row}=MAX($D${start}:$D${end})")
        price_ranges[ticker] = (start, end)
    refs["price_ranges"] = price_ranges
    return row


def _write_earnings_support(ws, row, signal_date, factors, calc, section, write_header, cell, raw_fill, refs):
    row = section(row, "3. earnings_momentum 원천: K200 내 12MF OP 성장률 분포")
    row = write_header(row, ["티커", "12MF_OP", "전월_12MF_OP", "계산가능", "이익성장률", "점수수식"], fill=raw_fill)
    start = row
    row_by_ticker = {}
    valid = factors["earnings_momentum"].loc[signal_date].notna()
    for ticker in valid.index[valid]:
        row_by_ticker[ticker] = row
        cell(row, 1, ticker)
        cell(row, 2, _at(calc["monthly_op"], signal_date, ticker))
        cell(row, 3, _at(calc["prev_op"], signal_date, ticker))
        cell(row, 4, f"=AND(ISNUMBER(B{row}),ISNUMBER(C{row}),C{row}<>0)")
        cell(row, 5, f"=IFERROR(IF(AND(B{row}<100000000000,(B{row}-C{row})/ABS(C{row})>0.5),0,(B{row}-C{row})/ABS(C{row})),NA())")
        row += 1
    end = row - 1
    for score_row in range(start, end + 1):
        cell(score_row, 6, f"=IFERROR(MIN(4,INT((RANK.AVG(E{score_row},$E${start}:$E${end},1)-1)*5/COUNT($E${start}:$E${end}))),NA())")
    refs["earnings_metric_range"] = (start, end)
    refs["earnings_rows"] = row_by_ticker
    return row


def _write_dividend_support(ws, row, signal_date, factors, calc, section, write_header, cell, raw_fill, refs):
    row = section(row, "4. dividend_yield 원천: K200 내 배당수익률 분포와 3년 배당 증가 bonus")
    row = write_header(row, ["티커", "DPS_TTM", "종가", "배당수익률", "연도1현금배당", "연도2현금배당", "연도3현금배당", "bonus", "점수수식"], fill=raw_fill)
    start = row
    row_by_ticker = {}
    valid = factors["dividend_yield"].loc[signal_date].notna()
    for ticker in valid.index[valid]:
        bonus = _dividend_bonus_parts(calc["dividend_cash_ttm"], signal_date, ticker)
        row_by_ticker[ticker] = row
        cell(row, 1, ticker)
        cell(row, 2, _at(calc["monthly_dps"], signal_date, ticker))
        cell(row, 3, _at(calc["monthly_close"], signal_date, ticker))
        cell(row, 4, f"=IFERROR(B{row}/C{row},NA())")
        cell(row, 5, bonus["배당증가_현금배당1"])
        cell(row, 6, bonus["배당증가_현금배당2"])
        cell(row, 7, bonus["배당증가_현금배당3"])
        cell(row, 8, f"=IF(AND(E{row}<F{row},F{row}<G{row}),1,0)")
        row += 1
    end = row - 1
    for score_row in range(start, end + 1):
        cell(score_row, 9, f"=IFERROR(MIN(4,INT((RANK.AVG(D{score_row},$D${start}:$D${end},1)-1)*5/COUNT($D${start}:$D${end})))+H{score_row},NA())")
    refs["dividend_metric_range"] = (start, end)
    refs["dividend_rows"] = row_by_ticker
    return row


def _write_retail_support(ws, row, signal_date, calc, section, write_header, cell, raw_fill, refs):
    row = section(row, "5. retail_flow 원천: 섹터별 252일 개인순매수 평균과 점수")
    row = write_header(row, ["섹터", "섹터평균_252일개인수급", "점수용_metric(-평균)", "섹터점수수식"], fill=raw_fill)
    start = row
    sector_rows = {}
    table = calc["retail_sector_table"]
    if not table.empty:
        sector_frame = table.loc[signal_date] if signal_date in table.index.get_level_values(0) else pd.DataFrame()
        for sector, values in sector_frame.iterrows():
            sector_rows[sector] = row
            cell(row, 1, sector)
            cell(row, 2, values["sector_avg_252d_retail_flow"])
            cell(row, 3, f"=-B{row}")
            row += 1
    end = row - 1
    for score_row in range(start, end + 1):
        cell(score_row, 4, f"=IFERROR(MIN(4,INT((RANK.AVG(C{score_row},$C${start}:$C${end},1)-1)*5/COUNT($C${start}:$C${end}))),NA())")
    refs["retail_sector_rows"] = sector_rows
    refs["retail_sector_range"] = (start, end)
    return row


def _write_value_support(ws, row, signal_date, factors, calc, section, write_header, cell, raw_fill, formula_fill, refs):
    row = section(row, "6. value 원천: K200 내 FCF/TEV 분포")
    source_row = _source_row_for_signal(calc["quarter_source_frame"], signal_date)
    source_period = source_row["_source_period"] if source_row is not None else None
    row = write_header(row, ["티커", "원천재무월", "FCF", "시가총액", "이자발생부채", "당좌자산", "TEV수식", "FCF/TEV수식", "점수수식"], fill=raw_fill)
    start = row
    row_by_ticker = {}
    valid = factors["value"].loc[signal_date].notna()
    for ticker in valid.index[valid]:
        row_by_ticker[ticker] = row
        cell(row, 1, ticker)
        cell(row, 2, str(source_period) if source_period is not None else pd.NA)
        cell(row, 3, _period_value(calc["free_cash_flow"], source_period, ticker))
        cell(row, 4, _at(calc["monthly_mktcap"], signal_date, ticker))
        cell(row, 5, _period_value(calc["debt"], source_period, ticker))
        cell(row, 6, _period_value(calc["quick_asset"], source_period, ticker))
        cell(row, 7, f"=D{row}+E{row}-F{row}", fill=formula_fill)
        cell(row, 8, f"=IFERROR(IF(G{row}<=0,-1E+99,C{row}/G{row}),NA())", fill=formula_fill)
        row += 1
    end = row - 1
    for score_row in range(start, end + 1):
        cell(score_row, 9, f"=IFERROR(MIN(4,INT((RANK.AVG(H{score_row},$H${start}:$H${end},1)-1)*5/COUNT($H${start}:$H${end}))),NA())", fill=formula_fill)
    refs["value_rows"] = row_by_ticker
    refs["value_metric_range"] = (start, end)
    return row


def _fill_summary_rows(ws, summary_rows, tickers, signal_date, market, factors, calc, refs, formula_fill):
    for ticker in tickers:
        sector = _at(calc["sector"], signal_date, ticker)
        retail_avg, _ = _retail_values(calc["retail_sector_table"], signal_date, sector)

        row = summary_rows[(ticker, "price_momentum")]
        start, end = refs["price_ranges"][ticker]
        _fill_summary_row(
            ws,
            row,
            ticker,
            "price_momentum",
            f"동일 시트 raw close D{start}:D{end}",
            "신호일 종가",
            _at(calc["close"], signal_date, ticker),
            "252일 종가고가",
            f"=MAX($D${start}:$D${end})",
            "threshold",
            0.8,
            f"=IFERROR(F{row}/H{row},NA())",
            _at(calc["price_ratio"], signal_date, ticker),
            f"=IF(K{row}>J{row},1,0)",
            _at(factors["price_momentum"], signal_date, ticker),
            "252일 raw close range에서 MAX를 직접 계산합니다.",
            formula_fill,
        )

        row = summary_rows[(ticker, "earnings_momentum")]
        earnings_start, earnings_end = refs["earnings_metric_range"]
        _fill_summary_row(
            ws,
            row,
            ticker,
            "earnings_momentum",
            f"K200 성장률 분포 E{earnings_start}:E{earnings_end}",
            "12MF OP",
            _at(calc["monthly_op"], signal_date, ticker),
            "전월 12MF OP",
            _at(calc["prev_op"], signal_date, ticker),
            "저OP 고성장 reset",
            bool(_at(calc["op_low_extreme"], signal_date, ticker, default=False)),
            f"=IFERROR(IF(AND(F{row}<100000000000,(F{row}-H{row})/ABS(H{row})>0.5),0,(F{row}-H{row})/ABS(H{row})),NA())",
            _at(calc["op_growth_metric"], signal_date, ticker),
            f"=IFERROR(MIN(4,INT((RANK.AVG(K{row},$E${earnings_start}:$E${earnings_end},1)-1)*5/COUNT($E${earnings_start}:$E${earnings_end}))),NA())",
            _at(factors["earnings_momentum"], signal_date, ticker),
            "성장률 산출 후 K200 분포에서 0~4점으로 bucket화합니다.",
            formula_fill,
        )

        row = summary_rows[(ticker, "dividend_yield")]
        dividend_start, dividend_end = refs["dividend_metric_range"]
        dividend_row = refs["dividend_rows"].get(ticker)
        bonus_ref = f"$H${dividend_row}" if dividend_row else "0"
        _fill_summary_row(
            ws,
            row,
            ticker,
            "dividend_yield",
            f"K200 배당수익률 분포 D{dividend_start}:D{dividend_end}",
            "DPS_TTM",
            _at(calc["monthly_dps"], signal_date, ticker),
            "종가",
            _at(calc["monthly_close"], signal_date, ticker),
            "3년 증가 bonus",
            f"={bonus_ref}",
            f"=IFERROR(F{row}/H{row},NA())",
            _at(calc["dividend_yield"], signal_date, ticker),
            f"=IFERROR(MIN(4,INT((RANK.AVG(K{row},$D${dividend_start}:$D${dividend_end},1)-1)*5/COUNT($D${dividend_start}:$D${dividend_end})))+J{row},NA())",
            _at(factors["dividend_yield"], signal_date, ticker),
            "배당수익률 bucket에 3년 연속 연말 현금배당 증가 bonus를 더합니다.",
            formula_fill,
        )

        row = summary_rows[(ticker, "retail_flow")]
        sector_row = refs["retail_sector_rows"].get(sector)
        _fill_summary_row(
            ws,
            row,
            ticker,
            "retail_flow",
            "섹터별 평균 개인수급 분포",
            "섹터",
            sector,
            "섹터평균 252일 개인수급",
            retail_avg,
            "점수용 metric",
            f"=-H{row}",
            f"=J{row}",
            -retail_avg if pd.notna(retail_avg) else pd.NA,
            f"=IFERROR($D${sector_row},NA())" if sector_row else "=NA()",
            _at(factors["retail_flow"], signal_date, ticker),
            "개인 순매도가 클수록 높은 점수입니다. 종목에는 섹터 점수를 배분합니다.",
            formula_fill,
        )

        row = summary_rows[(ticker, "value")]
        value_start, value_end = refs["value_metric_range"]
        source_period = _source_period_for_signal(calc["quarter_source_frame"], signal_date)
        _fill_summary_row(
            ws,
            row,
            ticker,
            "value",
            f"원천재무월 {source_period}, K200 FCF/TEV 분포 H{value_start}:H{value_end}",
            "FCF",
            _at(calc["lagged_fcf"], signal_date, ticker),
            "TEV",
            f"=P{row}+Q{row}-R{row}",
            "원천재무월",
            source_period,
            f"=IFERROR(IF(H{row}<=0,-1E+99,F{row}/H{row}),NA())",
            _at(calc["value_metric"], signal_date, ticker),
            f"=IFERROR(MIN(4,INT((RANK.AVG(K{row},$H${value_start}:$H${value_end},1)-1)*5/COUNT($H${value_start}:$H${value_end}))),NA())",
            _at(factors["value"], signal_date, ticker),
            "TEV = 시가총액 + 이자발생부채 - 당좌자산, TEV<=0은 최저 metric으로 처리합니다.",
            formula_fill,
        )
        ws.cell(row, 16, _at(calc["monthly_mktcap"], signal_date, ticker))
        ws.cell(row, 17, _at(calc["lagged_debt"], signal_date, ticker))
        ws.cell(row, 18, _at(calc["lagged_quick"], signal_date, ticker))


def _fill_summary_row(
    ws,
    row: int,
    ticker: str,
    factor_name: str,
    source_range,
    input1_label,
    input1_value,
    input2_label,
    input2_value,
    input3_label,
    input3_value,
    metric_formula,
    metric_value,
    score_formula,
    score_value,
    explanation,
    formula_fill,
) -> None:
    values = [
        TICKERS[ticker],
        ticker,
        factor_name,
        source_range,
        input1_label,
        input1_value,
        input2_label,
        input2_value,
        input3_label,
        input3_value,
        metric_formula,
        metric_value,
        score_formula,
        score_value,
        explanation,
    ]
    for column, value in enumerate(values, start=1):
        target = ws.cell(row, column, value)
        target.border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        target.alignment = Alignment(vertical="center", wrap_text=True)
    for column in (8, 10, 11, 13):
        value = ws.cell(row, column).value
        if isinstance(value, str) and value.startswith("="):
            ws.cell(row, column).fill = formula_fill


def _value_parts_text(calc: dict[str, object], signal_date: pd.Timestamp, ticker: str) -> str:
    market_cap = _at(calc["monthly_mktcap"], signal_date, ticker)
    quick = _at(calc["lagged_quick"], signal_date, ticker)
    debt = _at(calc["lagged_debt"], signal_date, ticker)
    return f"시총={market_cap}, 당좌={quick}, 부채={debt}"


def _format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    formula_fill = PatternFill("solid", fgColor="E2F0D9")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:200])
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 10), 30)

    _add_one_page_formulas(wb["one_page_KR"], formula_fill)
    _add_price_summary_formulas(wb["price_252d_summary"], wb["price_252d_raw_window"], formula_fill)
    _add_value_lag_trace_formulas(wb["value_lag_summary"], formula_fill)
    for sheet_name in ("A005930_calc_KR", "A000660_calc_KR"):
        _add_formulas(wb[sheet_name], formula_fill)
    _add_value_lag_trace_formulas(wb["value_lag_trace"], formula_fill)
    wb.save(path)


def _add_formulas(ws, formula_fill: PatternFill) -> None:
    headers = {cell.value: cell.column for cell in ws[1]}

    def col(header: str) -> str:
        return get_column_letter(headers[header])

    formula_headers = (
        "가격모멘텀_비율_엑셀수식",
        "이익성장률_엑셀수식",
        "배당수익률_엑셀수식",
        "배당증가_bonus_엑셀수식",
        "TEV_엑셀수식",
        "FCF_over_TEV_엑셀수식",
    )
    for row in range(2, ws.max_row + 1):
        ws[f"{col('가격모멘텀_비율_엑셀수식')}{row}"] = f"=IFERROR({col('종가')}{row}/{col('252일_종가고가')}{row},NA())"
        ws[f"{col('이익성장률_엑셀수식')}{row}"] = (
            f"=IFERROR(({col('12MF_OP')}{row}-{col('전월_12MF_OP')}{row})/ABS({col('전월_12MF_OP')}{row}),NA())"
        )
        ws[f"{col('배당수익률_엑셀수식')}{row}"] = f"=IFERROR({col('DPS_TTM')}{row}/{col('배당계산_종가')}{row},NA())"
        ws[f"{col('배당증가_bonus_엑셀수식')}{row}"] = (
            f"=IF(AND({col('배당증가_현금배당1')}{row}<{col('배당증가_현금배당2')}{row},"
            f"{col('배당증가_현금배당2')}{row}<{col('배당증가_현금배당3')}{row}),1,0)"
        )
        ws[f"{col('TEV_엑셀수식')}{row}"] = f"={col('시가총액')}{row}+{col('lagged_이자발생부채')}{row}-{col('lagged_당좌자산')}{row}"
        ws[f"{col('FCF_over_TEV_엑셀수식')}{row}"] = (
            f"=IFERROR(IF({col('TEV_엑셀수식')}{row}<=0,-1E+99,"
            f"{col('lagged_FCF')}{row}/{col('TEV_엑셀수식')}{row}),NA())"
        )
        for header in formula_headers:
            ws.cell(row, headers[header]).fill = formula_fill


def _add_one_page_formulas(ws, formula_fill: PatternFill) -> None:
    headers = {cell.value: cell.column for cell in ws[1]}

    def col(header: str) -> str:
        return get_column_letter(headers[header])

    formula_headers = (
        "가격_비율_수식",
        "이익성장률_수식",
        "배당수익률_수식",
        "가치_TEV_수식",
        "가치_FCF/TEV_수식",
    )
    for row in range(2, ws.max_row + 1):
        ws.cell(row, headers["가격_비율_수식"]).value = f"=IFERROR({col('가격_종가')}{row}/{col('가격_252일고가')}{row},NA())"
        ws.cell(row, headers["이익성장률_수식"]).value = (
            f"=IFERROR(({col('이익_12MF_OP')}{row}-{col('이익_전월_12MF_OP')}{row})/"
            f"ABS({col('이익_전월_12MF_OP')}{row}),NA())"
        )
        ws.cell(row, headers["배당수익률_수식"]).value = f"=IFERROR({col('배당_DPS_TTM')}{row}/{col('배당_종가')}{row},NA())"
        ws.cell(row, headers["가치_TEV_수식"]).value = (
            f"={col('가치_시가총액')}{row}+{col('가치_이자발생부채')}{row}-{col('가치_당좌자산')}{row}"
        )
        ws.cell(row, headers["가치_FCF/TEV_수식"]).value = (
            f"=IFERROR(IF({col('가치_TEV_수식')}{row}<=0,-1E+99,"
            f"{col('가치_FCF')}{row}/{col('가치_TEV_수식')}{row}),NA())"
        )
        for header in formula_headers:
            ws.cell(row, headers[header]).fill = formula_fill


def _add_price_summary_formulas(ws, raw_ws, formula_fill: PatternFill) -> None:
    headers = {cell.value: cell.column for cell in ws[1]}
    raw_headers = {cell.value: cell.column for cell in raw_ws[1]}

    def col(header: str) -> str:
        return get_column_letter(headers[header])

    raw_ticker_col = raw_headers["티커"]
    raw_close_col = get_column_letter(raw_headers["raw_close"])
    high_formula_header = "252일_종가고가_엑셀수식"
    ratio_formula_header = "종가/고가_엑셀수식"

    for row in range(2, ws.max_row + 1):
        ticker = ws.cell(row, headers["티커"]).value
        raw_rows = [idx for idx in range(2, raw_ws.max_row + 1) if raw_ws.cell(idx, raw_ticker_col).value == ticker]
        if not raw_rows:
            continue
        start, end = raw_rows[0], raw_rows[-1]
        ws.cell(row, headers[high_formula_header]).value = f"=MAX('price_252d_raw_window'!{raw_close_col}{start}:{raw_close_col}{end})"
        ws.cell(row, headers[ratio_formula_header]).value = f"=IFERROR({col('신호일_종가')}{row}/{col(high_formula_header)}{row},NA())"
        ws.cell(row, headers[high_formula_header]).fill = formula_fill
        ws.cell(row, headers[ratio_formula_header]).fill = formula_fill


def _add_value_lag_trace_formulas(ws, formula_fill: PatternFill) -> None:
    headers = {cell.value: cell.column for cell in ws[1]}

    def col(header: str) -> str:
        return get_column_letter(headers[header])

    fcf_header = "lagged_FCF" if "lagged_FCF" in headers else "원천_FCF"
    debt_header = "lagged_이자발생부채" if "lagged_이자발생부채" in headers else "원천_이자발생부채"
    quick_header = "lagged_당좌자산" if "lagged_당좌자산" in headers else "원천_당좌자산"
    tev_header = "TEV_엑셀수식"
    metric_header = "FCF_over_TEV_엑셀수식"
    for row in range(2, ws.max_row + 1):
        ws.cell(row, headers[tev_header]).value = (
            f"={col('시가총액')}{row}+{col(debt_header)}{row}-{col(quick_header)}{row}"
        )
        ws.cell(row, headers[metric_header]).value = (
            f"=IFERROR(IF({col(tev_header)}{row}<=0,-1E+99,{col(fcf_header)}{row}/{col(tev_header)}{row}),NA())"
        )
        ws.cell(row, headers[tev_header]).fill = formula_fill
        ws.cell(row, headers[metric_header]).fill = formula_fill


if __name__ == "__main__":
    main()
