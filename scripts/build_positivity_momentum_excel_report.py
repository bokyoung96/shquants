from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from root import ROOT

from backtesting.data import ParquetStore
from backtesting.strategies.positivity import positivity_score


INPUT_DIR = ROOT.results_path / "pos_research" / "momentum_horizon_high_low_2020"
OUTPUT_DIR = ROOT.results_path / "pos_research" / "positivity_momentum_report"
REPORT_PATH = OUTPUT_DIR / "positivity_momentum_2020_report.xlsx"
FALLBACK_REPORT_PATH = OUTPUT_DIR / "positivity_momentum_2020_report_with_q5_subplots.xlsx"
HORIZON_ORDER = ["1M", "3M", "6M", "12M", "3-1M", "6-1M", "12-1M"]
FACTOR_ORDER = ["positivity", "return_momentum", "high_sharpe", "high_low"]
FACTOR_LABELS = {
    "positivity": "Positivity",
    "return_momentum": "Return momentum",
    "high_sharpe": "High Sharpe",
    "high_low": "High-low channel",
}
COLORS = {
    "positivity": "#2563eb",
    "return_momentum": "#dc2626",
    "high_sharpe": "#16a34a",
    "high_low": "#9333ea",
}


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    summary = pd.read_csv(INPUT_DIR / "summary.csv")
    daily_returns = pd.read_csv(INPUT_DIR / "daily_returns.csv", parse_dates=["date"]).set_index("date")
    report_data = {
        "metrics": _build_metrics_sheet(summary),
        "winner_summary": _same_horizon_winner_summary(summary),
        "tail_diagnostics": _tail_diagnostics(daily_returns),
        "latest_portfolio": _latest_2026_positivity_12m_portfolio(),
    }
    histogram_paths = {
        "q5": OUTPUT_DIR / "q5_return_histograms.png",
        "spread": OUTPUT_DIR / "spread_return_histograms.png",
    }
    _plot_histogram_subplots(daily_returns, histogram_paths["q5"], leg="q5", title="Q5 daily return histograms")
    _plot_histogram_subplots(
        daily_returns,
        histogram_paths["spread"],
        leg="q5_minus_q1",
        title="Q5-Q1 daily return histograms",
    )
    report_path = write_workbook(report_data, histogram_paths)
    print(report_path)


def _build_metrics_sheet(summary: pd.DataFrame) -> pd.DataFrame:
    frame = summary.copy()
    frame["factor"] = pd.Categorical(frame["factor"], FACTOR_ORDER, ordered=True)
    frame["horizon"] = pd.Categorical(frame["horizon"], HORIZON_ORDER, ordered=True)
    frame = frame.sort_values(["leg", "horizon", "factor"])
    columns = [
        "leg",
        "horizon",
        "factor",
        "portfolio",
        "observations",
        "total_return",
        "cagr",
        "mdd",
        "sharpe",
        "daily_win_rate",
        "avg_daily_return",
        "daily_vol",
    ]
    return frame[columns]


def _same_horizon_winner_summary(summary: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for leg in ["q5", "q5_minus_q1"]:
        leg_frame = summary.loc[summary["leg"].eq(leg)].copy()
        for horizon in HORIZON_ORDER:
            sub = leg_frame.loc[leg_frame["horizon"].eq(horizon)].sort_values(
                ["sharpe", "cagr"],
                ascending=False,
            )
            if sub.empty:
                continue
            best = sub.iloc[0]
            positivity = sub.loc[sub["factor"].eq("positivity")].iloc[0]
            rows.append(
                {
                    "leg": leg,
                    "horizon": horizon,
                    "best_factor_by_sharpe": best["factor"],
                    "best_sharpe": best["sharpe"],
                    "best_cagr": best["cagr"],
                    "best_mdd": best["mdd"],
                    "positivity_rank_by_sharpe": int(sub.reset_index(drop=True).index[sub["factor"].eq("positivity")][0] + 1),
                    "positivity_cagr": positivity["cagr"],
                    "positivity_mdd": positivity["mdd"],
                    "positivity_sharpe": positivity["sharpe"],
                    "positivity_daily_win_rate": positivity["daily_win_rate"],
                }
            )
    return pd.DataFrame(rows)


def _tail_diagnostics(daily_returns: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    focus = ["6M", "12M", "6-1M", "12-1M"]
    for leg in ["q5", "q5_minus_q1"]:
        for horizon in focus:
            for factor in FACTOR_ORDER:
                column = f"{factor}_{horizon}_{leg}"
                if column not in daily_returns.columns:
                    continue
                series = daily_returns[column].dropna()
                q05 = float(series.quantile(0.05))
                rows.append(
                    {
                        "leg": leg,
                        "horizon": horizon,
                        "factor": factor,
                        "mean_daily_return": float(series.mean()),
                        "daily_vol": float(series.std(ddof=0)),
                        "win_rate": float(series.gt(0).mean()),
                        "skew": float(series.skew()),
                        "excess_kurtosis": float(series.kurt()),
                        "p05_daily_return": q05,
                        "cvar_5pct": float(series.loc[series.le(q05)].mean()),
                        "worst_daily_return": float(series.min()),
                    }
                )
    return pd.DataFrame(rows)


def _latest_2026_positivity_12m_portfolio() -> dict[str, pd.DataFrame | pd.Timestamp]:
    store = ParquetStore(ROOT.parquet_path)
    close = store.read("qw_adj_c").astype(float)
    membership = store.read("qw_k200_yn").reindex(index=close.index, columns=close.columns).fillna(False).astype(bool)
    names = pd.read_parquet(ROOT.parquet_path / "map__ticker_name_gics_sector_map.parquet")
    name_by_ticker = dict(zip(names["TICKER"], names["NAME"], strict=False))

    returns = close.pct_change(fill_method=None)
    score = positivity_score(returns, lookback=252, min_periods=252).where(membership)
    ranks = score.rank(axis=1, method="first", pct=True)
    q5_mask = ranks.gt(0.8) & score.notna() & membership
    counts = q5_mask.sum(axis=1)
    weights = q5_mask.div(counts.where(counts.gt(0)), axis=0).fillna(0.0)
    active_2026 = weights.loc["2026":].sum(axis=1).gt(0.0)
    latest_date = weights.loc["2026":].loc[active_2026].index.max()

    latest = _portfolio_frame(
        weights=weights,
        score=score,
        ranks=ranks,
        date=latest_date,
        name_by_ticker=name_by_ticker,
    )
    month_end_dates = weights.loc["2026":].loc[active_2026]
    month_end_dates = month_end_dates.loc[~month_end_dates.index.to_period("M").duplicated(keep="last")].index
    snapshots = pd.concat(
        [
            _portfolio_frame(
                weights=weights,
                score=score,
                ranks=ranks,
                date=date,
                name_by_ticker=name_by_ticker,
            )
            for date in month_end_dates
        ],
        ignore_index=True,
    )
    return {"latest_date": latest_date, "latest": latest, "snapshots": snapshots}


def _portfolio_frame(
    *,
    weights: pd.DataFrame,
    score: pd.DataFrame,
    ranks: pd.DataFrame,
    date: pd.Timestamp,
    name_by_ticker: dict[str, str],
) -> pd.DataFrame:
    row = weights.loc[date]
    tickers = row.loc[row.gt(0.0)].sort_values(ascending=False).index
    frame = pd.DataFrame(
        {
            "date": date.date().isoformat(),
            "종목코드": tickers,
            "종목명": [name_by_ticker.get(ticker, ticker) for ticker in tickers],
            "비중": [float(row.loc[ticker]) for ticker in tickers],
            "positivity_12m": [float(score.loc[date, ticker]) for ticker in tickers],
            "rank_pct": [float(ranks.loc[date, ticker]) for ticker in tickers],
        }
    )
    frame["종목코드_숫자"] = frame["종목코드"].str.replace("^A", "", regex=True)
    return frame[["date", "종목코드", "종목코드_숫자", "종목명", "비중", "positivity_12m", "rank_pct"]]


def _plot_histogram_subplots(daily_returns: pd.DataFrame, path: Path, *, leg: str, title: str) -> None:
    fig, axes = plt.subplots(3, 3, figsize=(18, 12))
    axes_list = list(axes.flat)
    for ax, horizon in zip(axes_list, HORIZON_ORDER, strict=False):
        for factor in FACTOR_ORDER:
            column = f"{factor}_{horizon}_{leg}"
            if column not in daily_returns.columns:
                continue
            values = daily_returns[column].dropna() * 100.0
            ax.hist(
                values,
                bins=45,
                alpha=0.28,
                density=True,
                color=COLORS[factor],
                label=FACTOR_LABELS[factor],
            )
        ax.axvline(0.0, color="#111111", lw=0.8, alpha=0.55)
        ax.set_title(horizon)
        ax.grid(True, alpha=0.2)
        ax.tick_params(axis="x", labelsize=8)
        ax.tick_params(axis="y", labelsize=8)
    for ax in axes_list[len(HORIZON_ORDER) :]:
        ax.axis("off")
    handles, labels = axes_list[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="lower center", ncols=4, frameon=False)
    fig.suptitle(title, fontsize=15)
    fig.tight_layout(rect=(0, 0.04, 1, 0.96))
    fig.savefig(path, dpi=180)
    plt.close(fig)


def write_workbook(report_data: dict[str, object], histogram_paths: dict[str, Path]) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    metrics_ws = wb.create_sheet("성과지표")
    hist_ws = wb.create_sheet("Return_Hist")
    q5_subplot_ws = wb.create_sheet("Q5_Equity_Drawdown")
    evidence_ws = wb.create_sheet("근거_진단")
    interpretation_ws = wb.create_sheet("경제적_해석")
    sources_ws = wb.create_sheet("Sources")
    portfolio_ws = wb.create_sheet("Portfolio_2026_12M")

    metrics = report_data["metrics"]
    winner_summary = report_data["winner_summary"]
    tail_diagnostics = report_data["tail_diagnostics"]
    portfolio = report_data["latest_portfolio"]
    assert isinstance(metrics, pd.DataFrame)
    assert isinstance(winner_summary, pd.DataFrame)
    assert isinstance(tail_diagnostics, pd.DataFrame)
    assert isinstance(portfolio, dict)

    _write_title(metrics_ws, "Positivity / Momentum / High-low 성과지표")
    _write_dataframe(metrics_ws, metrics, start_row=3)
    _format_sheet(metrics_ws, percent_columns={"total_return", "cagr", "mdd", "daily_win_rate", "avg_daily_return", "daily_vol"})

    _write_title(hist_ws, "Return histogram subplots")
    hist_ws["A3"] = "Q5 daily returns"
    hist_ws["A3"].font = Font(bold=True)
    hist_ws.add_image(Image(str(histogram_paths["q5"])), "A4")
    hist_ws["A70"] = "Q5-Q1 daily returns"
    hist_ws["A70"].font = Font(bold=True)
    hist_ws.add_image(Image(str(histogram_paths["spread"])), "A71")

    _write_title(q5_subplot_ws, "Q5 cumulative return and drawdown subplots")
    q5_subplot_ws["A3"] = "Q5 cumulative return by horizon"
    q5_subplot_ws["A3"].font = Font(bold=True)
    q5_subplot_ws.add_image(Image(str(INPUT_DIR / "q5_equity_subplots.png")), "A4")
    q5_subplot_ws["A70"] = "Q5 drawdown by horizon"
    q5_subplot_ws["A70"].font = Font(bold=True)
    q5_subplot_ws.add_image(Image(str(INPUT_DIR / "q5_drawdown_subplots.png")), "A71")

    latest_date = portfolio["latest_date"]
    latest = portfolio["latest"]
    snapshots = portfolio["snapshots"]
    assert isinstance(latest_date, pd.Timestamp)
    assert isinstance(latest, pd.DataFrame)
    assert isinstance(snapshots, pd.DataFrame)
    _write_title(portfolio_ws, f"2026 최신 12M Positivity Q5 포트폴리오: {latest_date.date().isoformat()}")
    _write_dataframe(portfolio_ws, latest, start_row=3)
    next_row = len(latest) + 6
    portfolio_ws.cell(next_row, 1, "2026 월말 스냅샷").font = Font(bold=True, size=13)
    _write_dataframe(portfolio_ws, snapshots, start_row=next_row + 2)
    _format_sheet(portfolio_ws, percent_columns={"비중", "positivity_12m", "rank_pct"})

    _write_title(evidence_ws, "Positivity가 좋은 이유: 실증 진단")
    evidence_ws["A3"] = "같은 lookback별 Sharpe 1위와 positivity 순위"
    evidence_ws["A3"].font = Font(bold=True)
    _write_dataframe(evidence_ws, winner_summary, start_row=5)
    tail_start = len(winner_summary) + 8
    evidence_ws.cell(tail_start, 1, "Tail / distribution diagnostics").font = Font(bold=True)
    _write_dataframe(evidence_ws, tail_diagnostics, start_row=tail_start + 2)
    _format_sheet(
        evidence_ws,
        percent_columns={
            "best_cagr",
            "best_mdd",
            "positivity_cagr",
            "positivity_mdd",
            "positivity_daily_win_rate",
            "mean_daily_return",
            "daily_vol",
            "win_rate",
            "p05_daily_return",
            "cvar_5pct",
            "worst_daily_return",
        },
    )

    _write_interpretation(interpretation_ws)
    _write_sources(sources_ws)
    for ws in wb.worksheets:
        ws.freeze_panes = "A3"
    try:
        wb.save(REPORT_PATH)
        return REPORT_PATH
    except PermissionError:
        wb.save(FALLBACK_REPORT_PATH)
        return FALLBACK_REPORT_PATH


def _write_title(ws, title: str) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=15)
    ws["A2"] = f"Generated: {datetime.now().isoformat(timespec='seconds')}"
    ws["A2"].font = Font(italic=True, color="666666")


def _write_dataframe(ws, frame: pd.DataFrame, *, start_row: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col_idx, column in enumerate(frame.columns, start=1):
        cell = ws.cell(start_row, col_idx, column)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_idx, (_, row) in enumerate(frame.iterrows(), start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
                value = None
            ws.cell(row_idx, col_idx, value)


def _format_sheet(ws, *, percent_columns: set[str] | None = None) -> None:
    percent_columns = percent_columns or set()
    header_row = None
    for row in range(1, min(ws.max_row, 20) + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        if any(value in percent_columns for value in values):
            header_row = row
            for col_idx, value in enumerate(values, start=1):
                if value in percent_columns:
                    for data_row in range(row + 1, ws.max_row + 1):
                        ws.cell(data_row, col_idx).number_format = "0.00%"
                elif isinstance(value, str) and value in {"sharpe", "best_sharpe", "positivity_sharpe"}:
                    for data_row in range(row + 1, ws.max_row + 1):
                        ws.cell(data_row, col_idx).number_format = "0.000"
            break
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 10
        for cell in ws[column_letter]:
            if cell.value is not None:
                max_length = max(max_length, min(len(str(cell.value)) + 2, 40))
        ws.column_dimensions[column_letter].width = max_length
    if header_row is not None:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"


def _write_interpretation(ws) -> None:
    _write_title(ws, "왜 한국 시장에서 positivity가 momentum보다 잘 작동할 수 있는가")
    rows = [
        ("핵심 결론", "같은 lookback 기준에서는 positivity가 대부분 Sharpe 1위다. 특히 Q5-Q1 spread에서는 12M, 6-1M, 12-1M에서 return/high-sharpe보다 MDD가 낮고 Sharpe가 높다."),
        ("Positivity란", "Chen, Jiang, Liu, Zhu(2026)는 positivity를 특정 기간 중 일간 수익률이 0 이상인 날의 비율로 정의한다. 예를 들어 12M positivity가 60%라면, 최근 252거래일 중 약 151일이 하락하지 않은 날이었다는 뜻이다. 이 지표는 가격이 얼마나 크게 올랐는지보다 상승 경로가 얼마나 꾸준했는지를 측정한다."),
        ("논문 abstract의 시사점", "해당 논문은 positivity가 과거 수익률이나 return consistency 같은 전통적 momentum 지표보다 장기 horizon에서 더 오래 지속되는 예측력을 보인다고 보고한다. 또한 high positivity winner는 투기적 glamour growth보다 상대적으로 덜 주목받는 value 성격, 견조한 펀더멘털, 높은 이익 성장의 조합으로 설명된다고 제시한다."),
        ("부호 기반 정보", "positivity는 수익률의 크기보다 상승일 빈도를 본다. 큰 한두 번의 급등보다 꾸준한 수급/정보 반영을 선호하므로 한국 시장의 테마성 급등락 노이즈를 덜 산다."),
        ("왜 high Sharpe와 다르나", "high Sharpe는 평균/변동성이라 극단 수익률과 변동성 추정에 민감하다. positivity는 하루 수익률 부호를 세기 때문에 jump magnitude를 덜 반영하고, 안정적 상승 경로를 더 직접적으로 잡는다."),
        ("왜 return momentum보다 낫나", "return momentum은 한 번의 큰 상승이 lookback 전체를 지배할 수 있다. 한국 시장처럼 개인투자자 거래, 테마 순환, 가격제한폭/공매도 제약이 강한 시장에서는 급등 후 반전 위험이 크므로 빈도 기반 필터가 유리할 수 있다."),
        ("경제적 해석", "positivity가 높은 종목은 정보가 여러 날에 걸쳐 점진적으로 반영되거나, 매도 압력이 낮고 수급이 꾸준한 상태일 가능성이 높다. 이는 단순 고수익률 momentum보다 crowding과 crash risk가 낮은 momentum 품질 신호로 해석할 수 있다."),
        ("주의점", "long-only Q5에서 12-1M은 high_sharpe/return_momentum이 더 높은 수익률을 보였다. positivity의 강점은 절대 수익률 극대화보다 같은 horizon 내 risk-adjusted consistency와 spread 안정성에 있다."),
    ]
    for idx, (title, text) in enumerate(rows, start=4):
        ws.cell(idx, 1, title).font = Font(bold=True)
        ws.cell(idx, 2, text)
        ws.cell(idx, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 110


def _write_sources(ws) -> None:
    _write_title(ws, "참고 문헌 / 외부 근거")
    rows = [
        {
            "source_id": "S0",
            "title": "Positivity and long-lasting momentum",
            "url": "https://ideas.repec.org/a/eee/empfin/v87y2026ics0927539826000095.html",
            "usage": "positivity를 non-negative return days의 비율로 정의하고, 장기 momentum 예측력과 펀더멘털 기반 해석을 제시한 핵심 참고문헌.",
        },
        {
            "source_id": "S1",
            "title": "Return Signal Momentum",
            "url": "https://ideas.repec.org/a/eee/jbfina/v124y2021ics0378426621000212.html",
            "usage": "과거 수익률의 부호/방향 정보가 기존 momentum보다 tail risk와 crash risk를 줄일 수 있다는 해석 근거.",
        },
        {
            "source_id": "S2",
            "title": "Stock market anomalies and individual investors in the Korean stock market",
            "url": "https://ideas.repec.org/a/eee/pacfin/v46y2017ipap141-157.html",
            "usage": "한국 시장에서 개인투자자 거래와 anomaly/과대평가/미래 수익률 관계를 해석하는 근거.",
        },
        {
            "source_id": "S3",
            "title": "OECD Economic Surveys: Korea 2024",
            "url": "https://www.oecd.org/en/publications/oecd-economic-surveys-korea-2024_bee6fecb-en.html",
            "usage": "한국 자본시장과 기업 밸류업/시장 구조 논의의 거시 배경.",
        },
        {
            "source_id": "S4",
            "title": "Korea short-selling restrictions / market microstructure references",
            "url": "https://www.fsc.go.kr/eng/pr010101/82465",
            "usage": "한국 시장의 공매도 제도와 제약이 momentum crash/overpricing 조정 속도에 영향을 줄 수 있다는 해석 배경.",
        },
    ]
    _write_dataframe(ws, pd.DataFrame(rows), start_row=4)
    _format_sheet(ws)


if __name__ == "__main__":
    main()
