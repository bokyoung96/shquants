import json
import subprocess
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from etfs import paths
from etfs.refresh.holdings_refresh import (
    DataguideExcelRefreshDriver,
    build_parser,
    extract_sheet_records,
    filter_refresh_targets,
    load_refresh_targets,
    load_refresh_targets_from_payloads,
    load_refresh_targets_from_ticker_workbook,
    refresh_targets_to_parquet_files_batch,
    refresh_targets_to_parquet_files,
    refresh_targets_to_parquet,
    ticker_output_path,
    write_records_by_ticker_parquet,
    write_records_parquet,
)


def test_load_refresh_targets_prefers_specs_products_and_formats_excel_codes(tmp_path: Path) -> None:
    specs_path = tmp_path / "methodology_specs.json"
    specs_path.write_text(
        json.dumps(
            {
                "indices": [
                    {
                        "index_code": "FI00.EXAMPLE.A",
                        "products": [
                            {"etf_code": "111111", "etf_name": "ETF A"},
                            {"etf_code": "A222222", "etf_name": "ETF B"},
                        ],
                    },
                    {
                        "index_code": "FI00.EXAMPLE.B",
                        "products": [{"etf_code": "111111", "etf_name": "ETF A duplicate"}],
                    },
                ]
            }
        ),
        encoding="utf-8",
    )
    rules_path = tmp_path / "rules.json"
    rules_path.write_text(
        json.dumps({"items": [{"code": "333333", "name": "Rules ETF", "status": "downloaded"}]}),
        encoding="utf-8",
    )

    targets = load_refresh_targets(specs_path=specs_path, rules_path=rules_path)

    assert [target.etf_code for target in targets] == ["111111", "222222"]
    assert [target.excel_code for target in targets] == ["A111111", "A222222"]
    assert targets[0].index_code == "FI00.EXAMPLE.A"
    assert targets[0].source == "methodology_specs"


def test_load_refresh_targets_falls_back_to_rules_when_specs_have_no_products(tmp_path: Path) -> None:
    specs_path = tmp_path / "methodology_specs.json"
    specs_path.write_text(json.dumps({"indices": [{"index_code": "FI00.EMPTY", "products": []}]}), encoding="utf-8")
    rules_path = tmp_path / "rules.json"
    rules_path.write_text(
        json.dumps(
            {
                "items": [
                    {"code": "333333", "name": "Downloaded ETF", "status": "downloaded"},
                    {"code": "444444", "name": "Missing PDF", "status": "not_found"},
                ]
            }
        ),
        encoding="utf-8",
    )

    targets = load_refresh_targets(specs_path=specs_path, rules_path=rules_path)

    assert [(target.etf_code, target.etf_name, target.source) for target in targets] == [
        ("333333", "Downloaded ETF", "rules")
    ]


def test_load_refresh_targets_from_ticker_workbook_uses_visible_rows_only(tmp_path: Path) -> None:
    path = tmp_path / "ticker.xlsx"
    wb = Workbook()
    ws = wb.active
    for _ in range(8):
        ws.append([])
    ws.append(["코드", "코드명", "유형", "아이템코드", "아이템명", "집계주기", "기초지수명"])
    ws.append(["A111111", "Visible ETF", "SSC", "S411000100", "ETF기초지수명", "일간", "FnGuide Visible"])
    ws.append(["A222222", "Hidden ETF", "SSC", "S411000100", "ETF기초지수명", "일간", "FnGuide Hidden"])
    ws.row_dimensions[11].hidden = True
    wb.save(path)

    targets = load_refresh_targets_from_ticker_workbook(path)

    assert [(target.etf_code, target.excel_code, target.etf_name, target.index_code, target.source) for target in targets] == [
        ("111111", "A111111", "Visible ETF", "FnGuide Visible", "ticker_workbook")
    ]


def test_extract_sheet_records_reads_a_to_h_from_row_7_and_normalizes_codes(tmp_path: Path) -> None:
    workbook_path = _write_dataguide_workbook(tmp_path)

    records = extract_sheet_records(workbook_path, index_code_by_etf={"0167A0": "FI00.WLT.KSS"})

    assert [record["as_of"] for record in records] == ["2026-06-16", "2026-06-16"]
    assert records[0] == {
        "as_of": "2026-06-16",
        "etf_code": "0167A0",
        "etf_code_raw": "A0167A0",
        "etf_name": "SOL AI반도체TOP2플러스",
        "index_code": "FI00.WLT.KSS",
        "security_code": "009150",
        "security_code_raw": "A009150",
        "security_name": "삼성전기",
        "quantity": 161.0,
        "market_value": 329728000.0,
        "weight": 0.2575,
        "weight_percent": 25.75,
        "is_cash": False,
    }
    assert records[1]["is_cash"] is True
    assert records[1]["security_code"] == ""


def test_write_records_parquet_round_trips_with_pyarrow(tmp_path: Path) -> None:
    records = [
        {
            "as_of": "2026-06-16",
            "etf_code": "0167A0",
            "etf_code_raw": "A0167A0",
            "etf_name": "SOL AI반도체TOP2플러스",
            "index_code": "FI00.WLT.KSS",
            "security_code": "009150",
            "security_code_raw": "A009150",
            "security_name": "삼성전기",
            "quantity": 161.0,
            "market_value": 329728000.0,
            "weight": 0.2575,
            "weight_percent": 25.75,
            "is_cash": False,
        }
    ]
    output_path = tmp_path / "holdings.parquet"

    write_records_parquet(records, output_path)

    frame = pd.read_parquet(output_path)
    assert frame.to_dict("records") == records


def test_write_records_by_ticker_parquet_writes_one_file_per_etf(tmp_path: Path) -> None:
    records = [
        _record("0167A0", "000660"),
        _record("0167A0", "005930"),
        _record("469150", "000660"),
    ]

    outputs = write_records_by_ticker_parquet(records, tmp_path / "files")

    assert sorted(outputs) == ["0167A0", "469150"]
    assert outputs["0167A0"] == tmp_path / "files" / "holdings_0167A0.parquet"
    assert outputs["469150"] == tmp_path / "files" / "holdings_469150.parquet"
    assert pd.read_parquet(outputs["0167A0"])["security_code"].tolist() == ["000660", "005930"]
    assert pd.read_parquet(outputs["469150"])["security_code"].tolist() == ["000660"]


def test_refresh_targets_to_parquet_uses_driver_and_writes_combined_results(tmp_path: Path) -> None:
    workbook_path = _write_dataguide_workbook(tmp_path)
    output_path = tmp_path / "refreshed.parquet"
    driver = FakeRefreshDriver(workbook_path)

    result_path = refresh_targets_to_parquet(
        targets=load_refresh_targets_from_codes(["0167A0", "123456"]),
        template_path=workbook_path,
        output_path=output_path,
        index_code_by_etf={"0167A0": "FI00.WLT.KSS", "123456": "FI00.EXAMPLE"},
        driver=driver,
    )

    assert result_path == output_path
    assert driver.calls == ["A0167A0", "A123456"]
    frame = pd.read_parquet(output_path)
    assert frame["etf_code"].tolist() == ["0167A0", "0167A0", "123456", "123456"]


def test_refresh_targets_to_parquet_files_writes_target_named_files(tmp_path: Path) -> None:
    workbook_path = _write_dataguide_workbook(tmp_path)
    output_dir = tmp_path / "files"
    driver = FakeRefreshDriver(workbook_path)

    outputs = refresh_targets_to_parquet_files(
        targets=load_refresh_targets_from_codes(["0167A0", "469150"]),
        template_path=workbook_path,
        output_dir=output_dir,
        index_code_by_etf={"0167A0": "FI00.WLT.KSS", "469150": "FI00.WLT.HBM"},
        driver=driver,
    )

    assert driver.calls == ["A0167A0", "A469150"]
    assert outputs == {
        "0167A0": output_dir / "holdings_0167A0.parquet",
        "469150": output_dir / "holdings_469150.parquet",
    }
    assert all(path.exists() for path in outputs.values())


def test_refresh_targets_to_parquet_files_batch_reuses_driver_batch_call(tmp_path: Path) -> None:
    workbook_path = _write_dataguide_workbook(tmp_path)
    output_dir = tmp_path / "files"
    targets = load_refresh_targets_from_codes(["0167A0", "469150"])
    driver = FakeBatchRefreshDriver(workbook_path, tmp_path / "batch_workbooks")

    outputs = refresh_targets_to_parquet_files_batch(
        targets=targets,
        template_path=workbook_path,
        output_dir=output_dir,
        index_code_by_etf={"0167A0": "FI00.WLT.KSS", "469150": "FI00.WLT.HBM"},
        driver=driver,
    )

    assert driver.calls == [["A0167A0", "A469150"]]
    assert outputs == {
        "0167A0": output_dir / "holdings_0167A0.parquet",
        "469150": output_dir / "holdings_469150.parquet",
    }
    assert pd.read_parquet(outputs["0167A0"])["etf_code"].unique().tolist() == ["0167A0"]
    assert pd.read_parquet(outputs["469150"])["etf_code"].unique().tolist() == ["469150"]


def test_excel_refresh_driver_invokes_powershell_dataguide_hyperlink_follow(tmp_path: Path) -> None:
    workbook_path = _write_dataguide_workbook(tmp_path)
    calls: list[list[str]] = []

    def fake_runner(command: list[str]) -> subprocess.CompletedProcess[str]:
        calls.append(command)
        return subprocess.CompletedProcess(command, 0, "", "")

    driver = DataguideExcelRefreshDriver(work_dir=tmp_path / "work", visible=True, runner=fake_runner)

    refreshed_path = driver.refresh(template_path=workbook_path, excel_code="A0167A0")

    assert refreshed_path.exists()
    assert calls and calls[0][0] == "powershell"
    script_text = (tmp_path / "work" / "refresh.ps1").read_text(encoding="utf-8")
    assert 'GetActiveObject' in script_text
    assert 'ScreenTip' in script_text
    assert 'DataGuide6' in script_text
    assert 'COMAddIns' in script_text
    assert 'Connect = $true' in script_text
    assert '.Follow(' in script_text
    assert "-ExcelCode" in calls[0]
    assert "A0167A0" in calls[0]
    assert "-Visible" in calls[0]


def test_excel_refresh_driver_batch_invokes_one_powershell_run_and_writes_targets(tmp_path: Path) -> None:
    workbook_path = _write_dataguide_workbook(tmp_path)
    calls: list[list[str]] = []

    def fake_runner(command: list[str]) -> subprocess.CompletedProcess[str]:
        calls.append(command)
        targets_path = Path(command[command.index("-TargetsPath") + 1])
        targets = json.loads(targets_path.read_text(encoding="utf-8"))
        for target in targets:
            Path(target["output_path"]).parent.mkdir(parents=True, exist_ok=True)
            Path(target["output_path"]).touch()
        return subprocess.CompletedProcess(command, 0, "", "")

    driver = DataguideExcelRefreshDriver(work_dir=tmp_path / "work", visible=True, runner=fake_runner)
    targets = load_refresh_targets_from_codes(["0167A0", "469150"])

    refreshed = driver.refresh_many(template_path=workbook_path, targets=targets)

    assert len(calls) == 1
    assert sorted(refreshed) == ["0167A0", "469150"]
    assert all(path.exists() for path in refreshed.values())
    script_text = (tmp_path / "work" / "refresh.ps1").read_text(encoding="utf-8")
    assert "ConvertFrom-Json" in script_text
    assert "SaveCopyAs" in script_text
    assert "-TargetsPath" in calls[0]
    assert "-Visible" in calls[0]


def test_filter_refresh_targets_accepts_raw_or_excel_codes_and_limit() -> None:
    targets = load_refresh_targets_from_codes(["0167A0", "123456", "0182R0"])

    filtered = filter_refresh_targets(targets, tickers=["A123456", "0182R0"], limit=1)

    assert [target.etf_code for target in filtered] == ["123456"]


def test_holdings_refresh_parser_defaults_to_fnguide_specs_and_validation_output() -> None:
    args = build_parser().parse_args([])

    assert args.template == "etfs/refresh/pdf.xlsx"
    assert args.specs == paths.FNGUIDE_METHODOLOGY_SPECS_JSON.as_posix()
    assert args.rules == paths.FNGUIDE_RULES_JSON.as_posix()
    assert args.output_dir == paths.REFRESHED_HOLDINGS_FILES_DIR.as_posix()
    assert args.refresh_mode == "excel"
    assert args.mode == "batch"
    assert args.tickers == []
    assert args.limit is None
    assert args.hidden is False
    assert args.work_dir == "etfs/refresh/work"


def test_ticker_output_path_is_stable_and_sanitized() -> None:
    assert ticker_output_path(Path("files"), "A0167A0") == Path("files/holdings_0167A0.parquet")


def load_refresh_targets_from_codes(codes: list[str]):
    specs = {
        "indices": [
            {
                "index_code": "FI00.TEST",
                "products": [{"etf_code": code, "etf_name": f"ETF {code}"} for code in codes],
            }
        ]
    }
    path = Path("unused.json")
    return load_refresh_targets_from_payloads(specs_payload=specs, rules_payload={}, specs_path=path, rules_path=path)


def _record(etf_code: str, security_code: str) -> dict[str, object]:
    return {
        "as_of": "2026-06-16",
        "etf_code": etf_code,
        "etf_code_raw": f"A{etf_code}",
        "etf_name": f"ETF {etf_code}",
        "index_code": "FI00.TEST",
        "security_code": security_code,
        "security_code_raw": f"A{security_code}",
        "security_name": f"Security {security_code}",
        "quantity": 1.0,
        "market_value": 100.0,
        "weight": 0.1,
        "weight_percent": 10.0,
        "is_cash": False,
    }


def _write_dataguide_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "pdf_A0167A0.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Refresh", "Last Updated: 2026-06-17 10:17:54"])
    ws.append(["ETF 구성종목"])
    ws.append(["ETF", "A0167A0", "SOL AI반도체TOP2플러스"])
    ws.append(["출력주기", "일간", "오름차순"])
    ws.append(["조회기간", 20260317, "최근일자(20260616)"])
    ws.append(["날짜", "ETF코드", "ETF명", "구성종목코드", "구성종목", "주식수(계약수)", "금액", "금액기준 구성비중(%)"])
    ws.append([datetime(2026, 6, 16), "A0167A0", "SOL AI반도체TOP2플러스", "A009150", "삼성전기", 161, 329728000, 25.75])
    ws.append([datetime(2026, 6, 16), "A0167A0", "SOL AI반도체TOP2플러스", None, "원화현금", 10, 10, 0.13])
    wb.save(path)
    return path


class FakeRefreshDriver:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self.calls: list[str] = []

    def refresh(self, *, template_path: Path, excel_code: str) -> Path:
        assert template_path == self.workbook_path
        self.calls.append(excel_code)
        wb = load_workbook(self.workbook_path)
        ws = wb.active
        code = excel_code[1:] if excel_code.startswith("A") else excel_code
        ws["B3"] = excel_code
        ws["C3"] = f"ETF {code}"
        for row in range(7, ws.max_row + 1):
            ws.cell(row=row, column=2).value = excel_code
            ws.cell(row=row, column=3).value = f"ETF {code}"
        wb.save(self.workbook_path)
        return self.workbook_path


class FakeBatchRefreshDriver:
    def __init__(self, workbook_path: Path, output_dir: Path) -> None:
        self.workbook_path = workbook_path
        self.output_dir = output_dir
        self.calls: list[list[str]] = []

    def refresh_many(self, *, template_path: Path, targets) -> dict[str, Path]:
        assert template_path == self.workbook_path
        target_list = list(targets)
        self.calls.append([target.excel_code for target in target_list])
        outputs: dict[str, Path] = {}
        self.output_dir.mkdir(parents=True, exist_ok=True)
        for target in target_list:
            output_path = self.output_dir / f"{target.etf_code}.xlsx"
            wb = load_workbook(self.workbook_path)
            ws = wb.active
            ws["B3"] = target.excel_code
            ws["C3"] = f"ETF {target.etf_code}"
            for row in range(7, ws.max_row + 1):
                ws.cell(row=row, column=2).value = target.excel_code
                ws.cell(row=row, column=3).value = f"ETF {target.etf_code}"
            wb.save(output_path)
            outputs[target.etf_code] = output_path
        return outputs
