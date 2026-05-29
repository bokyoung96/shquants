import pandas as pd
from openpyxl import Workbook

from scripts.build_mfbt_factor_signal_audit import ACCOUNTING_FORMAT, PERCENT_FORMAT, SCORE_FORMAT, FACTOR_SHEET_NAMES, _apply_number_formats, _price_window_sheet


def test_workbook_uses_one_sheet_per_factor_only() -> None:
    assert FACTOR_SHEET_NAMES == (
        "price_momentum",
        "earnings_momentum",
        "dividend_yield",
        "retail_flow",
        "value",
    )


def test_price_window_sheet_keeps_raw_rows_focused_on_rolling_window() -> None:
    signal_date = pd.Timestamp("2024-12-31")
    index = pd.bdate_range(end=signal_date, periods=252)
    close = pd.DataFrame({"A005930": range(1, 253)}, index=index, dtype=float)
    factors = {
        "price_momentum": pd.DataFrame({"A005930": [1.0]}, index=[signal_date]),
    }

    sheet = _price_window_sheet(["A005930"], signal_date, {"close": close}, factors)

    assert len(sheet) == 252
    assert list(sheet.columns) == ["종목명", "티커", "raw_date", "raw_close", "누적_종가고가", "최종고가일"]
    assert "신호일_종가" not in sheet.columns


def test_factor_audit_number_formats_follow_column_meaning() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["종목명", "종가", "배당수익률 수식", "score 코드"])
    ws.append(["삼성전자", 100000.0, "=B2/1000000", 4.0])

    _apply_number_formats(ws)

    assert ws["B2"].number_format == ACCOUNTING_FORMAT
    assert ws["C2"].number_format == PERCENT_FORMAT
    assert ws["D2"].number_format == SCORE_FORMAT
