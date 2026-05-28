import pandas as pd
from openpyxl import load_workbook

from scripts.build_mfbt_signal_matrix_excel import build_mfbt_factor_signals, signal_rows_only, write_signal_matrix_workbook


def test_signal_rows_only_drops_non_signal_daily_rows() -> None:
    frame = pd.DataFrame(
        {
            "A": [None, 1.0, None],
            "B": [None, None, 0.0],
        },
        index=pd.to_datetime(["2024-01-30", "2024-01-31", "2024-02-29"]),
    )

    output = signal_rows_only(frame)

    assert list(output.index) == [pd.Timestamp("2024-01-31"), pd.Timestamp("2024-02-29")]


def test_build_mfbt_factor_signals_returns_date_by_ticker_frames(monkeypatch) -> None:
    index = pd.to_datetime(["2024-01-31", "2024-02-29"])
    columns = ["A", "B"]
    expected = {
        "price_momentum": pd.DataFrame([[1.0, 0.0], [0.0, 1.0]], index=index, columns=columns),
        "value": pd.DataFrame([[0.0, 1.0], [1.0, 0.0]], index=index, columns=columns),
    }

    class FakeSignalProducer:
        def build(self, market):
            class Bundle:
                meta = expected

            return Bundle()

    class FakeStrategy:
        datasets = ()
        signal_producer = FakeSignalProducer()

    class FakeCatalog:
        pass

    class FakeRunner:
        catalog = FakeCatalog()

    class FakeLoader:
        def __init__(self, catalog, store):
            pass

        def load(self, request):
            class Market:
                frames = {"k200_yn": pd.DataFrame(True, index=index, columns=columns)}
                universe = None

            return Market()

    monkeypatch.setattr("scripts.build_mfbt_signal_matrix_excel.build_strategy", lambda name: FakeStrategy())
    monkeypatch.setattr("scripts.build_mfbt_signal_matrix_excel.BacktestRunner", FakeRunner)
    monkeypatch.setattr("scripts.build_mfbt_signal_matrix_excel.DataLoader", FakeLoader)

    factors = build_mfbt_factor_signals(start="2024-01-01", end="2024-02-29")

    assert factors == expected
    for frame in factors.values():
        assert list(frame.index) == list(index)
        assert list(frame.columns) == columns


def test_write_signal_matrix_workbook_writes_one_sheet_per_factor(tmp_path) -> None:
    index = pd.to_datetime(["2024-01-31", "2024-02-29"])
    factors = {
        "price_momentum": pd.DataFrame([[1.0, None], [0.0, 1.0]], index=index, columns=["A", "B"]),
        "value": pd.DataFrame([[None, None], [2.0, 3.0]], index=index, columns=["A", "B"]),
    }
    path = tmp_path / "signals.xlsx"

    write_signal_matrix_workbook(path, factors)

    wb = load_workbook(path, read_only=True, data_only=True)
    assert wb.sheetnames == ["price_momentum", "value"]
    assert len(list(wb["price_momentum"].iter_rows())) == 3
    assert len(list(wb["value"].iter_rows())) == 2
    assert [cell.value for cell in next(wb["price_momentum"].iter_rows(max_row=1))] == ["date", "A", "B"]
